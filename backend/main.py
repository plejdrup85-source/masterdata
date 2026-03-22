"""FastAPI application for masterdata quality check."""

import asyncio
import json
import logging
import os
import random
import shutil
import time
import uuid
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Optional

from fastapi import Body, FastAPI, File, HTTPException, Query, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse

from backend.ai_scorer import enrich_product_async, review_suggestions_async, score_product_async
from backend.analyzer import analyze_product
from backend.enricher import (
    apply_ai_review_to_suggestions,
    apply_enrichment_suggestions,
    enrich_product,
    final_quality_gate,
)
from backend.excel_handler import create_output_excel, read_article_numbers
from backend.image_analyzer import analyze_product_images
from backend.jeeves_loader import JeevesIndex, load_jeeves
from backend.manufacturer import (
    generate_improvement_suggestions,
    search_manufacturer_info,
    search_norengros,
)
from backend.models import (
    AnalysisJob, AnalysisMode, BatchMode, ImageSuggestion, JobStatus,
    ProductAnalysis, ProductData, QualityStatus, VerificationStatus,
)
from backend.scoring import score_product_areas, FOCUS_AREAS, ALL_AREAS, AREA_LABELS, ANALYSIS_PRESETS
from backend.pdf_enricher import run_enrichment_pipeline
from backend.scraper import (
    scrape_product, _load_sitemap, _sitemap_loaded, _sku_to_url,
    build_full_index, scan_index_incremental, get_index_stats,
    find_batch_products_in_sitemap,
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
logger = logging.getLogger(__name__)

app = FastAPI(
    title="Masterdata Kvalitetssjekk",
    description="Kvalitetsanalyse av produktkatalog mot onemed.no",
    version="2.0.0",
)

# CORS - allow Render domain and localhost for dev
ALLOWED_ORIGINS = os.environ.get(
    "ALLOWED_ORIGINS",
    "http://localhost:8000,http://localhost:3000,http://127.0.0.1:8000"
).split(",")

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "DELETE"],
    allow_headers=["*"],
)


def _find_jeeves_file() -> Optional[str]:
    """Find Jeeves Excel file from env var or well-known paths."""
    if JEEVES_FILE_PATH and Path(JEEVES_FILE_PATH).exists():
        return JEEVES_FILE_PATH
    # Auto-detect from common locations
    candidates = [
        Path("Masterdata 2103.xlsx"),
        Path("masterdata/Masterdata 2103.xlsx"),
        Path("/tmp/masterdata_cache/Masterdata_2103.xlsx"),
    ]
    for p in candidates:
        if p.exists():
            return str(p)
    return None


@app.on_event("startup")
async def preload_data():
    """Pre-load sitemap and Jeeves data on startup."""
    global _jeeves_index

    # Load Jeeves ERP data
    jeeves_path = _find_jeeves_file()
    if jeeves_path:
        try:
            _jeeves_index = load_jeeves(jeeves_path)
            logger.info(f"Jeeves data loaded: {_jeeves_index.count} products from {jeeves_path}")
        except Exception as e:
            logger.warning(f"Failed to load Jeeves data from {jeeves_path}: {e}")
    else:
        logger.warning(
            "Jeeves Excel file not found. Set JEEVES_FILE_PATH env var or place "
            "'Masterdata 2103.xlsx' in the project root. Two-source comparison disabled."
        )

    # Pre-load sitemap XML and cached SKU→URL index on startup.
    # This downloads ONE sitemap XML file and loads the cached SKU index from disk.
    # It does NOT scan/crawl product pages — that only happens in discovery mode.
    try:
        import httpx
        async with httpx.AsyncClient() as client:
            await _load_sitemap(client)
        logger.info("Sitemap index pre-loaded on startup (no page scanning)")
    except Exception as e:
        logger.warning(f"Failed to pre-load sitemap on startup: {e}")

    # Auto-start full index build if index coverage is low.
    # This runs in the background and doesn't block startup.
    idx = get_index_stats()
    if idx["sitemap_url_count"] > 0 and idx["coverage_pct"] < 80:
        logger.info(
            f"SKU index coverage is {idx['coverage_pct']}% "
            f"({idx['sku_index_count']}/{idx['sitemap_url_count']}). "
            f"Starting background index build..."
        )

        async def _auto_build():
            global _index_build_status, _index_build_task
            _index_build_status = {
                "status": "running",
                "started_at": datetime.now().isoformat(),
                "checked": 0, "total": 0, "new_indexed": 0,
                "trigger": "auto_startup",
            }

            async def _progress(checked, total, new_in_batch):
                _index_build_status["checked"] = checked
                _index_build_status["total"] = total
                _index_build_status["new_indexed"] += new_in_batch

            try:
                result = await build_full_index(on_progress=_progress)
                _index_build_status.update({"status": "completed", **result})
                logger.info(f"Auto index build completed: {result}")
            except Exception as exc:
                _index_build_status["status"] = "failed"
                _index_build_status["error"] = str(exc)
                logger.error(f"Auto index build failed: {exc}")

        _index_build_task = asyncio.create_task(_auto_build())
        _background_tasks.add(_index_build_task)
        _index_build_task.add_done_callback(_background_tasks.discard)
    else:
        logger.info(
            f"SKU index coverage OK: {idx['coverage_pct']}% "
            f"({idx['sku_index_count']}/{idx['sitemap_url_count']})"
        )

# In-memory job storage
jobs: dict[str, AnalysisJob] = {}

# Keep references to background tasks so they don't get GC-ed
_background_tasks: set[asyncio.Task] = set()

# Output directory - use /tmp on deployed environments for reliability
OUTPUT_DIR = Path(os.environ.get("OUTPUT_DIR", "/tmp/masterdata_output"))
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# Persistent history directory — survives job cleanup
HISTORY_DIR = Path(os.environ.get("HISTORY_DIR", "data/history"))
HISTORY_DIR.mkdir(parents=True, exist_ok=True)
HISTORY_EXCEL_DIR = HISTORY_DIR / "exports"
HISTORY_EXCEL_DIR.mkdir(parents=True, exist_ok=True)
HISTORY_INDEX_FILE = HISTORY_DIR / "index.json"
HISTORY_MAX_ENTRIES = 100


def _load_history() -> list[dict]:
    """Load job history index from disk."""
    if not HISTORY_INDEX_FILE.exists():
        return []
    try:
        return json.loads(HISTORY_INDEX_FILE.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError) as e:
        logger.warning(f"Failed to load history index: {e}")
        return []


def _save_history(entries: list[dict]) -> None:
    """Save job history index to disk (FIFO, max HISTORY_MAX_ENTRIES).

    When entries exceed the limit, the oldest are evicted and their
    Excel files are deleted from disk to reclaim space.
    """
    if len(entries) > HISTORY_MAX_ENTRIES:
        evicted = entries[:-HISTORY_MAX_ENTRIES]
        entries = entries[-HISTORY_MAX_ENTRIES:]
        for e in evicted:
            old_file = HISTORY_EXCEL_DIR / e.get("excel_filename", "")
            if old_file.exists():
                try:
                    old_file.unlink()
                    logger.info(f"History FIFO: deleted evicted Excel {old_file.name}")
                except OSError:
                    pass
    try:
        HISTORY_INDEX_FILE.write_text(
            json.dumps(entries, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
    except OSError as e:
        logger.error(f"Failed to save history index: {e}")


def _add_to_history(job: "AnalysisJob", source_filename: str = "") -> None:
    """Add a completed job to the persistent history.

    Copies the Excel file to the history directory and writes a metadata entry.
    """
    if job.status != JobStatus.COMPLETED or not job.output_file:
        return

    src = Path(job.output_file)
    if not src.exists():
        return

    # Copy Excel to persistent history directory
    history_filename = f"kvalitetssjekk_{job.job_id}.xlsx"
    dest = HISTORY_EXCEL_DIR / history_filename
    try:
        shutil.copy2(str(src), str(dest))
    except OSError as e:
        logger.error(f"Failed to copy Excel to history: {e}")
        return

    # Compute summary stats
    total = len(job.results)
    avg_score = 0.0
    critical_count = 0
    if total > 0:
        scores = []
        for r in job.results:
            s = r.total_score
            if r.ai_score and "overall_score" in r.ai_score:
                s = r.ai_score["overall_score"]
            scores.append(s)
            if s < 40:
                critical_count += 1
        avg_score = round(sum(scores) / len(scores), 1)

    entry = {
        "job_id": job.job_id,
        "timestamp": datetime.fromtimestamp(job.created_at).isoformat(),
        "created_at": job.created_at,
        "source_filename": source_filename,
        "excel_filename": history_filename,
        "product_count": total,
        "analysis_mode": job.analysis_mode,
        "focus_areas": job.focus_areas,
        "batch_info": job.batch_info or "",
        "avg_score": avg_score,
        "critical_count": critical_count,
    }

    entries = _load_history()
    entries.append(entry)
    _save_history(entries)

# Jeeves ERP data file path — auto-detected from repo or configured via env
JEEVES_FILE_PATH = os.environ.get("JEEVES_FILE_PATH", "")
_jeeves_index: Optional[JeevesIndex] = None

# Concurrency control - be polite to onemed.no
SCRAPE_CONCURRENCY = 3
SCRAPE_DELAY = 1.0  # seconds between requests

# Limits
MAX_FILE_SIZE = 5 * 1024 * 1024  # 5 MB
MAX_ARTICLES = 1000
MAX_CONCURRENT_JOBS = 3
JOB_TTL_SECONDS = 2 * 60 * 60  # 2 hours

# Rate limiting (simple per-IP)
_rate_limit: dict[str, list[float]] = defaultdict(list)
RATE_LIMIT_WINDOW = 60  # seconds
RATE_LIMIT_MAX = 5  # max uploads per window


def _cleanup_old_jobs() -> None:
    """Remove jobs older than JOB_TTL_SECONDS."""
    now = time.time()
    expired = [
        jid for jid, job in jobs.items()
        if now - job.created_at > JOB_TTL_SECONDS
    ]
    for jid in expired:
        job = jobs.pop(jid, None)
        if job and job.output_file:
            try:
                Path(job.output_file).unlink(missing_ok=True)
            except Exception:
                pass
        logger.info(f"Cleaned up expired job {jid}")


def _check_rate_limit(client_ip: str) -> bool:
    """Return True if request is within rate limits."""
    now = time.time()
    timestamps = _rate_limit[client_ip]
    # Remove old entries
    _rate_limit[client_ip] = [t for t in timestamps if now - t < RATE_LIMIT_WINDOW]
    if len(_rate_limit[client_ip]) >= RATE_LIMIT_MAX:
        return False
    _rate_limit[client_ip].append(now)
    return True


@app.get("/", response_class=HTMLResponse)
async def root():
    """Serve the frontend."""
    frontend_path = Path("frontend/index.html")
    if frontend_path.exists():
        return HTMLResponse(content=frontend_path.read_text(encoding="utf-8"))
    return HTMLResponse(content="<h1>Frontend not found</h1>", status_code=404)


@app.get("/api/health")
async def health_check():
    """Health check endpoint with feature status."""
    idx = get_index_stats()
    return {
        "status": "ok",
        "version": "2.1.0",
        "active_jobs": len(jobs),
        "jeeves_loaded": _jeeves_index is not None and _jeeves_index.loaded,
        "jeeves_product_count": _jeeves_index.count if _jeeves_index else 0,
        "ai_scoring_available": bool(os.environ.get("ANTHROPIC_API_KEY")),
        "sku_index": idx,
        "analysis_modes": [
            {"value": m.value, "label": l}
            for m, l in [
                (AnalysisMode.FULL_ENRICHMENT, "Full berikelse og forslag"),
                (AnalysisMode.AUDIT_ONLY, "Kun kvalitetsrevisjon"),
                (AnalysisMode.FOCUSED_SCAN, "Fokusert områdesjekk"),
            ]
        ],
        "focus_areas": [
            {"value": a, "label": AREA_LABELS[a]} for a in FOCUS_AREAS
        ],
        "presets": {
            k: {"label": v["label"], "description": v["description"],
                "analysis_mode": v["analysis_mode"], "focus_areas": v["focus_areas"]}
            for k, v in ANALYSIS_PRESETS.items()
        },
    }


# ── SKU Index management ──

# Track active index build so we don't run two at once
_index_build_task: Optional[asyncio.Task] = None
_index_build_status: dict = {}


@app.post("/api/build-index")
async def build_index():
    """Start a full SKU index build as a background task.

    Scans all sitemap product pages (~9500) and maps article numbers to URLs.
    Takes ~15-25 minutes but runs in the background. Results are persisted to
    disk and used automatically by all subsequent analyses.

    Returns immediately with status. Poll GET /api/index-status for progress.
    """
    global _index_build_task, _index_build_status

    if _index_build_task and not _index_build_task.done():
        return {
            "status": "already_running",
            "message": "Indeksbygging kjører allerede. Se /api/index-status for fremdrift.",
            **_index_build_status,
        }

    _index_build_status = {
        "status": "running",
        "started_at": datetime.now().isoformat(),
        "checked": 0,
        "total": 0,
        "new_indexed": 0,
    }

    async def _progress_callback(checked, total, new_in_batch):
        _index_build_status["checked"] = checked
        _index_build_status["total"] = total
        _index_build_status["new_indexed"] += new_in_batch

    async def _run_build():
        try:
            result = await build_full_index(on_progress=_progress_callback)
            _index_build_status.update({
                "status": "completed",
                "completed_at": datetime.now().isoformat(),
                **result,
            })
            logger.info(f"[build-index] Background build completed: {result}")
        except Exception as e:
            _index_build_status.update({
                "status": "failed",
                "error": str(e),
            })
            logger.error(f"[build-index] Background build failed: {e}")

    _index_build_task = asyncio.create_task(_run_build())
    _background_tasks.add(_index_build_task)
    _index_build_task.add_done_callback(_background_tasks.discard)

    return {
        "status": "started",
        "message": "Indeksbygging startet i bakgrunnen. Se /api/index-status for fremdrift.",
    }


@app.get("/api/index-status")
async def index_status():
    """Get current SKU index status and build progress."""
    idx = get_index_stats()
    build = dict(_index_build_status) if _index_build_status else {"status": "never_run"}

    return {
        "index": idx,
        "build": build,
    }


def _apply_batch_selection(
    articles: list[str],
    batch_mode: str,
    range_start: Optional[int],
    range_end: Optional[int],
    sample_size: Optional[int],
    sample_seed: Optional[int],
    specific_articles: Optional[str],
) -> tuple[list[str], str]:
    """Apply batch selection to article list.

    Returns (selected_articles, batch_info_description).
    """
    total = len(articles)

    if batch_mode == BatchMode.RANGE.value:
        # 1-based row indices
        start = max(1, range_start or 1) - 1  # convert to 0-based
        end = min(total, range_end or total)
        selected = articles[start:end]
        info = f"Rader {start + 1}\u2013{end} av {total} (utvalg: {len(selected)})"
        return selected, info

    elif batch_mode == BatchMode.RANDOM.value:
        n = min(sample_size or 100, total)
        seed = sample_seed if sample_seed is not None else int(time.time())
        rng = random.Random(seed)
        selected = rng.sample(articles, n)
        info = f"Tilfeldig utvalg: {n} av {total} (seed={seed})"
        return selected, info

    elif batch_mode == BatchMode.SPECIFIC.value:
        if specific_articles:
            # Parse comma or newline separated list
            requested = {
                a.strip()
                for a in specific_articles.replace("\n", ",").split(",")
                if a.strip()
            }
            # Keep only articles that exist in the uploaded file
            selected = [a for a in articles if a in requested]
            # Also add any specified articles not in the file (user might want to check them anyway)
            in_file = set(articles)
            extra = [a for a in requested if a not in in_file]
            selected.extend(extra)
            info = f"Spesifikke artikler: {len(selected)} valgt ({len(extra)} ikke i fil)"
            return selected, info
        return articles, f"Alle {total} artikler (ingen spesifikke angitt)"

    # FULL mode (default)
    return articles, f"Alle {total} artikler"


@app.post("/api/upload")
async def upload_excel(
    request: Request,
    file: UploadFile = File(...),
    skip_cache: bool = Query(False, description="Skip cache and re-scrape all products"),
    batch_mode: str = Query(BatchMode.FULL.value, description="Batch mode: full, range, random, specific"),
    range_start: Optional[int] = Query(None, description="Start row (1-based) for range mode"),
    range_end: Optional[int] = Query(None, description="End row (inclusive) for range mode"),
    sample_size: Optional[int] = Query(None, description="Number of random samples"),
    sample_seed: Optional[int] = Query(None, description="Random seed for reproducible sampling"),
    specific_articles: Optional[str] = Query(None, description="Comma-separated article numbers for specific mode"),
    analysis_mode: str = Query(AnalysisMode.FULL_ENRICHMENT.value, description="Analysis mode: full_enrichment, audit_only, focused_scan"),
    focus_areas: Optional[str] = Query(None, description="Comma-separated focus areas for focused_scan mode"),
):
    """Upload an Excel file and start analysis with batch selection support."""
    # Cleanup old jobs first
    _cleanup_old_jobs()

    # Rate limiting per client IP
    client_ip = request.client.host if request.client else "unknown"
    if not _check_rate_limit(client_ip):
        raise HTTPException(
            429,
            f"For mange opplastinger. Maks {RATE_LIMIT_MAX} per {RATE_LIMIT_WINDOW} sekunder."
        )

    if not file.filename:
        raise HTTPException(400, "Ingen fil valgt")

    ext = Path(file.filename).suffix.lower()
    if ext != ".xlsx":
        raise HTTPException(
            400,
            f"Filformatet {ext} st\u00f8ttes ikke. Bruk .xlsx (Excel 2007+)."
        )

    # Check file size before reading
    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(
            400,
            f"Filen er for stor ({len(content) / 1024 / 1024:.1f} MB). Maks {MAX_FILE_SIZE / 1024 / 1024:.0f} MB."
        )

    # Check concurrent job limit
    active_jobs = sum(
        1 for j in jobs.values()
        if j.status in (JobStatus.PENDING, JobStatus.RUNNING)
    )
    if active_jobs >= MAX_CONCURRENT_JOBS:
        raise HTTPException(
            429,
            f"For mange samtidige analyser ({active_jobs}). Vent til en er ferdig."
        )

    try:
        article_numbers, detected_column = read_article_numbers(content, file.filename)
    except Exception as e:
        logger.error(f"Failed to read Excel: {e}")
        raise HTTPException(400, f"Kunne ikke lese Excel-filen: {e}")

    if not article_numbers:
        raise HTTPException(400, "Ingen artikkelnumre funnet i filen")

    if len(article_numbers) > MAX_ARTICLES:
        raise HTTPException(
            400,
            f"For mange artikkelnumre ({len(article_numbers)}). Maks {MAX_ARTICLES} per analyse."
        )

    # Remove duplicates while preserving order
    seen = set()
    unique_articles = []
    for a in article_numbers:
        if a not in seen:
            seen.add(a)
            unique_articles.append(a)

    # Apply batch selection
    selected_articles, batch_info = _apply_batch_selection(
        unique_articles, batch_mode, range_start, range_end,
        sample_size, sample_seed, specific_articles,
    )

    if not selected_articles:
        raise HTTPException(400, "Ingen artikler valgt etter batch-filtrering")

    # Validate analysis_mode
    valid_modes = [m.value for m in AnalysisMode]
    if analysis_mode not in valid_modes:
        raise HTTPException(400, f"Ugyldig analysemodus: {analysis_mode}. Gyldige: {valid_modes}")

    # Parse focus_areas
    parsed_focus_areas = []
    if analysis_mode == AnalysisMode.FOCUSED_SCAN.value:
        if not focus_areas:
            raise HTTPException(400, "Fokusert modus krever minst ett valgt område (focus_areas)")
        parsed_focus_areas = [a.strip() for a in focus_areas.split(",") if a.strip()]
        valid_focus = set(ALL_AREAS)
        invalid = [a for a in parsed_focus_areas if a not in valid_focus]
        if invalid:
            raise HTTPException(400, f"Ugyldige fokusområder: {invalid}. Gyldige: {list(valid_focus)}")
        if not parsed_focus_areas:
            raise HTTPException(400, "Minst ett fokusområde må velges")

    # Create job
    job_id = str(uuid.uuid4())[:8]
    job = AnalysisJob(
        job_id=job_id,
        status=JobStatus.PENDING,
        total_products=len(selected_articles),
        created_at=time.time(),
        batch_mode=batch_mode,
        batch_info=batch_info,
        analysis_mode=analysis_mode,
        focus_areas=parsed_focus_areas,
        source_filename=file.filename or "",
    )
    jobs[job_id] = job

    # Start analysis in background - store task reference to prevent GC
    task = asyncio.create_task(
        _run_analysis(job_id, selected_articles, skip_cache,
                      analysis_mode=analysis_mode, focus_areas=parsed_focus_areas)
    )
    _background_tasks.add(task)
    task.add_done_callback(_background_tasks.discard)

    # Estimate time: audit mode is faster (no enrichment/AI), focused is even faster
    n = len(selected_articles)
    if analysis_mode == AnalysisMode.AUDIT_ONLY.value:
        est_seconds = n * 3 // SCRAPE_CONCURRENCY
    elif analysis_mode == AnalysisMode.FOCUSED_SCAN.value:
        est_seconds = n * 2 // SCRAPE_CONCURRENCY
    else:
        est_seconds = n * 5 // SCRAPE_CONCURRENCY
    if est_seconds < 60:
        est_str = f"~{est_seconds} sekunder"
    elif est_seconds < 3600:
        est_str = f"~{est_seconds // 60} minutter"
    else:
        est_str = f"~{est_seconds // 3600} timer {(est_seconds % 3600) // 60} minutter"

    jeeves_status = f", Jeeves: {_jeeves_index.count} produkter" if _jeeves_index else ", Jeeves: ikke lastet"

    mode_labels = {
        AnalysisMode.FULL_ENRICHMENT.value: "Full berikelse",
        AnalysisMode.AUDIT_ONLY.value: "Kvalitetsrevisjon",
        AnalysisMode.FOCUSED_SCAN.value: "Fokusert sjekk",
    }
    mode_label = mode_labels.get(analysis_mode, analysis_mode)
    focus_label = ""
    if parsed_focus_areas:
        focus_label = " (" + ", ".join(AREA_LABELS.get(a, a) for a in parsed_focus_areas) + ")"

    return {
        "job_id": job_id,
        "total_products": len(selected_articles),
        "detected_column": detected_column,
        "batch_mode": batch_mode,
        "batch_info": batch_info,
        "analysis_mode": analysis_mode,
        "focus_areas": parsed_focus_areas,
        "estimated_time": est_str,
        "jeeves_loaded": _jeeves_index is not None and _jeeves_index.loaded,
        "message": f"{mode_label}{focus_label}: {n} produkter ({batch_info}). Estimert tid: {est_str}{jeeves_status}",
    }


@app.post("/api/analyze-catalog")
async def analyze_from_catalog(
    request: Request,
    data: dict = Body(...),
):
    """Start analysis from the internal Jeeves catalog without Excel upload.

    Supports multiple input modes:
    - catalog_full: all products in the catalog
    - catalog_random: random sample from catalog
    - manual_articles: user-provided article numbers

    Body: {
        "source_mode": "catalog_full"|"catalog_random"|"manual_articles",
        "sample_size": 100,       # for catalog_random
        "article_numbers": [...], # for manual_articles
        "analysis_mode": "full_enrichment"|"audit_only"|"focused_scan",
        "focus_areas": [...],     # for focused_scan
        "skip_cache": false
    }
    """
    _cleanup_old_jobs()

    client_ip = request.client.host if request.client else "unknown"
    if not _check_rate_limit(client_ip):
        raise HTTPException(
            429,
            f"For mange analyser. Maks {RATE_LIMIT_MAX} per {RATE_LIMIT_WINDOW} sekunder."
        )

    if not _jeeves_index or not _jeeves_index.loaded:
        raise HTTPException(400, "Jeeves-katalog er ikke lastet inn. Kan ikke kjøre katalogbasert analyse.")

    active_jobs = sum(
        1 for j in jobs.values()
        if j.status in (JobStatus.PENDING, JobStatus.RUNNING)
    )
    if active_jobs >= MAX_CONCURRENT_JOBS:
        raise HTTPException(
            429,
            f"For mange samtidige analyser ({active_jobs}). Vent til en er ferdig."
        )

    source_mode = data.get("source_mode", "catalog_full")
    analysis_mode = data.get("analysis_mode", AnalysisMode.FULL_ENRICHMENT.value)
    focus_areas_raw = data.get("focus_areas", [])
    skip_cache = data.get("skip_cache", False)

    # Validate analysis_mode
    valid_modes = [m.value for m in AnalysisMode]
    if analysis_mode not in valid_modes:
        raise HTTPException(400, f"Ugyldig analysemodus: {analysis_mode}")

    parsed_focus_areas = []
    if analysis_mode == AnalysisMode.FOCUSED_SCAN.value:
        parsed_focus_areas = [a.strip() for a in focus_areas_raw if a.strip()]
        if not parsed_focus_areas:
            raise HTTPException(400, "Fokusert modus krever minst ett valgt område")

    all_catalog_articles = _jeeves_index.all_article_numbers()
    catalog_total = len(all_catalog_articles)

    if source_mode == "catalog_full":
        selected_articles = all_catalog_articles
        batch_info = f"Hele katalogen: {catalog_total} artikler"
        source_filename = "jeeves_katalog_full"

    elif source_mode == "catalog_random":
        sample_size = data.get("sample_size", 100)
        if sample_size < 1:
            raise HTTPException(400, "sample_size må være minst 1")
        sample_size = min(sample_size, catalog_total)
        seed = int(time.time())
        rng = random.Random(seed)
        selected_articles = rng.sample(all_catalog_articles, sample_size)
        batch_info = f"Tilfeldig utvalg: {sample_size} av {catalog_total} (seed={seed})"
        source_filename = f"jeeves_katalog_random_{sample_size}"

    elif source_mode == "manual_articles":
        raw_articles = data.get("article_numbers", [])
        if not raw_articles:
            raise HTTPException(400, "Ingen artikkelnumre oppgitt")
        # Deduplicate and normalize
        from backend.identifiers import normalize_identifier
        seen = set()
        selected_articles = []
        for art in raw_articles:
            norm = normalize_identifier(art)
            if norm and norm not in seen:
                seen.add(norm)
                selected_articles.append(norm)
        if not selected_articles:
            raise HTTPException(400, "Ingen gyldige artikkelnumre etter validering")
        batch_info = f"Manuelt oppgitt: {len(selected_articles)} artikler"
        source_filename = f"manuell_{len(selected_articles)}_artikler"

    else:
        raise HTTPException(400, f"Ugyldig source_mode: {source_mode}. Bruk: catalog_full, catalog_random, manual_articles")

    if not selected_articles:
        raise HTTPException(400, "Ingen artikler å analysere")

    if len(selected_articles) > MAX_ARTICLES:
        raise HTTPException(
            400,
            f"For mange artikler ({len(selected_articles)}). Maks {MAX_ARTICLES} per analyse. Bruk tilfeldig utvalg."
        )

    # Create job
    job_id = str(uuid.uuid4())[:8]
    job = AnalysisJob(
        job_id=job_id,
        status=JobStatus.PENDING,
        total_products=len(selected_articles),
        created_at=time.time(),
        batch_mode=source_mode,
        batch_info=batch_info,
        analysis_mode=analysis_mode,
        focus_areas=parsed_focus_areas,
        source_filename=source_filename,
    )
    jobs[job_id] = job

    task = asyncio.create_task(
        _run_analysis(job_id, selected_articles, skip_cache,
                      analysis_mode=analysis_mode, focus_areas=parsed_focus_areas)
    )
    _background_tasks.add(task)
    task.add_done_callback(_background_tasks.discard)

    n = len(selected_articles)
    if analysis_mode == AnalysisMode.AUDIT_ONLY.value:
        est_seconds = n * 3 // SCRAPE_CONCURRENCY
    elif analysis_mode == AnalysisMode.FOCUSED_SCAN.value:
        est_seconds = n * 2 // SCRAPE_CONCURRENCY
    else:
        est_seconds = n * 5 // SCRAPE_CONCURRENCY
    if est_seconds < 60:
        est_str = f"~{est_seconds} sekunder"
    elif est_seconds < 3600:
        est_str = f"~{est_seconds // 60} minutter"
    else:
        est_str = f"~{est_seconds // 3600} timer {(est_seconds % 3600) // 60} minutter"

    mode_labels = {
        AnalysisMode.FULL_ENRICHMENT.value: "Full berikelse",
        AnalysisMode.AUDIT_ONLY.value: "Kvalitetsrevisjon",
        AnalysisMode.FOCUSED_SCAN.value: "Fokusert sjekk",
    }
    mode_label = mode_labels.get(analysis_mode, analysis_mode)

    return {
        "job_id": job_id,
        "total_products": n,
        "source_mode": source_mode,
        "batch_info": batch_info,
        "analysis_mode": analysis_mode,
        "focus_areas": parsed_focus_areas,
        "estimated_time": est_str,
        "catalog_total": catalog_total,
        "jeeves_loaded": True,
        "message": f"{mode_label}: {n} produkter ({batch_info}). Estimert tid: {est_str}",
    }


@app.get("/api/catalog-info")
async def get_catalog_info():
    """Return info about the internal catalog for the frontend."""
    if not _jeeves_index or not _jeeves_index.loaded:
        return {
            "loaded": False,
            "count": 0,
            "sample": [],
        }
    all_arts = _jeeves_index.all_article_numbers()
    return {
        "loaded": True,
        "count": len(all_arts),
        "sample": all_arts[:5],
    }


@app.get("/api/status/{job_id}")
async def get_status(job_id: str):
    """Get the status of an analysis job."""
    job = jobs.get(job_id)
    if not job:
        raise HTTPException(404, "Jobb ikke funnet. Den kan ha utl\u00f8pt.")

    return {
        "job_id": job.job_id,
        "status": job.status.value,
        "total_products": job.total_products,
        "processed_products": job.processed_products,
        "current_product": job.current_product,
        "current_step": job.current_step,
        "progress_percent": round(
            job.processed_products / job.total_products * 100, 1
        ) if job.total_products > 0 else 0,
        "errors": job.errors[-10:],
        "output_file": job.output_file,
        "batch_mode": job.batch_mode,
        "batch_info": job.batch_info,
        "analysis_mode": job.analysis_mode,
        "focus_areas": job.focus_areas,
    }


@app.delete("/api/cancel/{job_id}")
async def cancel_job(job_id: str):
    """Cancel a running analysis job."""
    job = jobs.get(job_id)
    if not job:
        raise HTTPException(404, "Jobb ikke funnet")

    if job.status not in (JobStatus.PENDING, JobStatus.RUNNING):
        raise HTTPException(400, "Jobben er allerede ferdig")

    job.cancelled = True
    job.status = JobStatus.FAILED
    job.errors.append("Analyse avbrutt av bruker")
    job.current_product = None
    logger.info(f"[{job_id}] Job cancelled by user")

    return {"message": "Analyse avbrutt", "job_id": job_id}


@app.get("/api/download/{job_id}")
async def download_result(
    job_id: str,
    threshold: Optional[int] = Query(None, description="Score threshold — export only products below this score"),
):
    """Download the analysis result Excel file.

    If threshold is set, regenerates the export with only products scoring below it.
    """
    job = jobs.get(job_id)
    if not job:
        raise HTTPException(404, "Jobb ikke funnet")

    if job.status != JobStatus.COMPLETED:
        raise HTTPException(400, "Analysen er ikke fullført enda")

    # If threshold is set, regenerate filtered export
    if threshold is not None and 0 < threshold <= 100:
        filtered_results = []
        for r in job.results:
            # Use area score if available, fall back to total_score
            area_score = (r.ai_score or {}).get("area_scores", {}).get("overall_score")
            effective_score = area_score if area_score is not None else r.total_score
            if effective_score < threshold:
                filtered_results.append(r)

        if not filtered_results:
            raise HTTPException(400, f"Ingen produkter har score under {threshold}")

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filtered_filename = f"masterdata_filtrert_{job_id}_{timestamp}.xlsx"
        filtered_path = str(OUTPUT_DIR / filtered_filename)
        create_output_excel(
            filtered_results, filtered_path,
            analysis_mode=job.analysis_mode, focus_areas=job.focus_areas,
        )
        return FileResponse(
            filtered_path,
            filename=f"masterdata_kvalitetssjekk_{job_id}_under{threshold}.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    if not job.output_file or not Path(job.output_file).exists():
        raise HTTPException(404, "Resultatfilen finnes ikke lenger. Den kan ha blitt slettet.")

    return FileResponse(
        job.output_file,
        filename=f"masterdata_kvalitetssjekk_{job_id}.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.get("/api/results/{job_id}")
async def get_results(job_id: str):
    """Get analysis results as JSON."""
    job = jobs.get(job_id)
    if not job:
        raise HTTPException(404, "Jobb ikke funnet")

    if job.status not in (JobStatus.COMPLETED, JobStatus.RUNNING):
        raise HTTPException(400, "Ingen resultater tilgjengelig enda")

    return {
        "job_id": job.job_id,
        "status": job.status.value,
        "total_products": job.total_products,
        "processed_products": job.processed_products,
        "analysis_mode": job.analysis_mode,
        "focus_areas": job.focus_areas,
        "results": [
            {
                "article_number": r.article_number,
                "product_name": r.product_data.product_name,
                "found": r.product_data.found_on_onemed,
                "score": r.total_score,
                "status": r.overall_status.value,
                "comment": r.overall_comment,
                "manufacturer": r.product_data.manufacturer,
                "category": r.product_data.category,
                "requires_followup": r.requires_manufacturer_contact,
                "manual_review": r.manual_review_needed,
                "auto_fix": r.auto_fix_possible,
                "image_quality": _format_image_quality(r.image_quality),
                "pdf_available": r.pdf_available,
                "pdf_url": r.pdf_url,
                "enrichment_count": len([
                    e for e in r.enrichment_results
                    if e.match_status != "NOT_FOUND"
                ]),
                "enrichment_conflicts": len([
                    e for e in r.enrichment_results
                    if e.match_status == "FOUND_IN_BOTH_CONFLICT"
                ]),
                "priority_level": (r.ai_score or {}).get("area_scores", {}).get("priority_level", ""),
                "why_low": (r.ai_score or {}).get("area_scores", {}).get("why_low", ""),
                # Source traceability: which external sources were attempted and found
                "sources_used": {
                    "manufacturer": {
                        "searched": r.manufacturer_lookup.searched,
                        "found": r.manufacturer_lookup.found,
                        "source_url": r.manufacturer_lookup.source_url,
                        "notes": r.manufacturer_lookup.notes,
                    } if r.manufacturer_lookup.searched else None,
                    "norengros": {
                        "searched": r.norengros_lookup.searched,
                        "found": r.norengros_lookup.found,
                        "source_url": r.norengros_lookup.source_url,
                        "notes": r.norengros_lookup.notes,
                    } if r.norengros_lookup and r.norengros_lookup.searched else None,
                    "pdf": {
                        "available": r.pdf_available,
                        "url": r.pdf_url,
                    },
                },
                "ai_score": r.ai_score,
                "ai_enrichment": r.ai_enrichment,
                "enrichment_suggestions": [
                    {
                        "field_name": es.field_name,
                        "current_value": es.current_value,
                        "suggested_value": es.suggested_value,
                        "source": es.source,
                        "source_url": es.source_url,
                        "evidence": es.evidence,
                        "confidence": es.confidence,
                        "review_required": es.review_required,
                    }
                    for es in r.enrichment_suggestions
                ],
                "field_analyses": [
                    {
                        "field_name": fa.field_name,
                        "current_value": fa.current_value,
                        "status": fa.status.value,
                        "comment": fa.comment,
                        "suggested_value": fa.suggested_value,
                        "source": fa.source,
                        "confidence": fa.confidence,
                    }
                    for fa in r.field_analyses
                ],
            }
            for r in job.results
        ],
    }


# ── Job History API ──


@app.get("/api/history")
async def get_history():
    """Return list of past completed jobs (newest first)."""
    entries = _load_history()
    entries.reverse()
    return {"entries": entries}


@app.get("/api/history/{job_id}/download")
async def download_history_file(job_id: str):
    """Download an Excel file from job history."""
    entries = _load_history()
    entry = next((e for e in entries if e["job_id"] == job_id), None)
    if not entry:
        raise HTTPException(404, "Jobb ikke funnet i historikk")

    filepath = HISTORY_EXCEL_DIR / entry["excel_filename"]
    if not filepath.exists():
        raise HTTPException(404, "Excel-filen er ikke lenger tilgjengelig")

    download_name = entry.get("source_filename", "")
    if download_name:
        stem = Path(download_name).stem
        download_name = f"{stem}_kvalitetssjekk_{job_id}.xlsx"
    else:
        download_name = f"kvalitetssjekk_{job_id}.xlsx"

    return FileResponse(
        str(filepath),
        filename=download_name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.delete("/api/history/{job_id}")
async def delete_history_entry(job_id: str):
    """Delete a single job from history."""
    entries = _load_history()
    entry = next((e for e in entries if e["job_id"] == job_id), None)
    if not entry:
        raise HTTPException(404, "Jobb ikke funnet i historikk")

    # Delete Excel file
    filepath = HISTORY_EXCEL_DIR / entry["excel_filename"]
    filepath.unlink(missing_ok=True)

    # Remove from index
    entries = [e for e in entries if e["job_id"] != job_id]
    _save_history(entries)

    return {"message": "Historikkoppføring slettet", "job_id": job_id}


# ── Family / Relationship Analysis API ──
# Persisted results: stored per job_id so results survive page refreshes
_family_results: dict[str, dict] = {}  # job_id/source_id → full family results with review state

FAMILY_OUTPUT_DIR = OUTPUT_DIR / "families"
FAMILY_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def _serialize_family_member(m):
    return {
        "article_number": m.article_number,
        "role": m.role,
        "product_name": m.product_name,
        "specification": m.specification,
        "child_specific_title": m.child_specific_title,
        "variant_dimensions": [
            {"name": d.dimension_name, "value": d.value, "source": d.source}
            for d in m.variant_dimensions
        ],
        "safety_attributes": m.safety_attributes if m.safety_attributes else {},
    }


def _run_family_detection(products: list[dict], source_id: str, data_source: str = "unknown") -> dict:
    """Run family detection and persist results."""
    from backend.family_detector import detect_families
    families, all_members = detect_families(products)

    family_list = []
    for f in sorted(families, key=lambda x: (-x.confidence, -len(x.members))):
        family_list.append({
            "family_id": f.family_id,
            "family_name": f.family_name,
            "member_count": len(f.members),
            "variant_dimensions": f.variant_dimension_names,
            "mother_article": f.mother_article,
            "confidence": f.confidence,
            "review_required": f.review_required,
            "grouping_reason": f.grouping_reason,
            "review_status": "pending",  # pending / accepted / rejected / needs_review
            "review_comment": "",
            "members": [_serialize_family_member(m) for m in f.members],
        })

    standalone_list = [
        {
            "article_number": m.article_number,
            "product_name": m.product_name,
            "specification": m.specification,
            "grouping_reason": m.grouping_reason,
            "confidence": m.confidence,
        }
        for m in all_members if m.role == "standalone"
    ]

    families_with_dims = sum(1 for f in families if f.variant_dimension_names)
    strong = sum(1 for f in families if f.confidence >= 0.65)

    result = {
        "source_id": source_id,
        "data_source": data_source,
        "total_products": len(all_members),
        "total_families": len(families),
        "strong_families": strong,
        "families_with_dimensions": families_with_dims,
        "families_without_dimensions": len(families) - families_with_dims,
        "standalone_products": len(standalone_list),
        "families": family_list,
        "standalone": standalone_list,
    }

    # Persist in memory
    _family_results[source_id] = result

    # Persist to disk
    try:
        import json
        disk_path = FAMILY_OUTPUT_DIR / f"{source_id}.json"
        disk_path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
        logger.info(f"Family results persisted to {disk_path}")
    except Exception as e:
        logger.warning(f"Failed to persist family results: {e}")

    return result


@app.get("/api/families/{job_id}")
async def get_families(job_id: str):
    """Get or compute family detection results for a completed analysis job.

    Results are cached: first call computes, subsequent calls return persisted data.
    """
    # Check memory cache first
    if job_id in _family_results:
        return _family_results[job_id]

    # Check disk cache
    import json
    disk_path = FAMILY_OUTPUT_DIR / f"{job_id}.json"
    if disk_path.exists():
        try:
            result = json.loads(disk_path.read_text(encoding="utf-8"))
            _family_results[job_id] = result
            return result
        except Exception:
            pass

    # Compute from analysis job
    from backend.family_detector import products_from_analyses
    job = jobs.get(job_id)
    if not job:
        raise HTTPException(404, "Jobb ikke funnet")
    if job.status != JobStatus.COMPLETED:
        raise HTTPException(400, "Analyse ikke fullført ennå")
    if not job.results:
        raise HTTPException(400, "Ingen resultater tilgjengelig")

    product_dicts = products_from_analyses(job.results)
    return _run_family_detection(product_dicts, job_id, data_source="quality_analysis")


@app.post("/api/families/analyze-jeeves")
async def analyze_jeeves_families():
    """Run family detection directly on the loaded Jeeves ERP data.

    Does NOT require a completed quality analysis job — works independently
    using the Jeeves dataset that was loaded at startup.
    """
    from backend.family_detector import products_from_jeeves_index

    if not _jeeves_index or not _jeeves_index.loaded:
        raise HTTPException(400, "Jeeves-data er ikke lastet inn")

    source_id = "jeeves-direct"

    # Return cached if available
    if source_id in _family_results:
        return _family_results[source_id]

    product_dicts = products_from_jeeves_index(_jeeves_index)
    return _run_family_detection(product_dicts, source_id, data_source="jeeves_direct")


_ARTICLE_SEARCH_TERMS = {
    "artikkelnummer", "artikkel", "artikkelnr", "artnr", "artnummer",
    "varenummer", "varenr", "article", "articlenumber", "article_number",
    "sku", "item", "itemnumber", "produktnummer",
}
_ARTICLE_SEARCH_NORMALIZED = {
    t.replace(".", "").replace(" ", "").replace("_", "")
    for t in _ARTICLE_SEARCH_TERMS
}


def _detect_article_column(headers: list[str]) -> int | None:
    """Return 0-based index of the auto-detected article-number column, or None."""
    for idx, h in enumerate(headers):
        if h:
            normalized = h.lower().replace(".", "").replace(" ", "").replace("_", "")
            if normalized in _ARTICLE_SEARCH_NORMALIZED:
                return idx
    return None


def _preview_sheet(ws) -> dict:
    """Extract preview data from a worksheet: headers, first 5 rows, total data rows."""
    rows = []
    total_data_rows = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        cell_values = [str(v).strip() if v is not None else "" for v in row]
        if i < 6:  # header + 5 data rows
            rows.append(cell_values)
        if i > 0:
            # Count non-empty data rows (at least one non-empty cell)
            if any(c for c in cell_values):
                total_data_rows += 1

    if not rows:
        return {"headers": [], "preview_rows": [], "detected_column": None,
                "total_columns": 0, "total_data_rows": 0}

    headers = rows[0]
    preview_rows = rows[1:] if len(rows) > 1 else []
    detected_col = _detect_article_column(headers)

    return {
        "headers": headers,
        "preview_rows": preview_rows,
        "detected_column": detected_col,
        "total_columns": len(headers),
        "total_data_rows": total_data_rows,
    }


@app.post("/api/families/preview-upload")
async def preview_upload_for_families(
    file: UploadFile = File(...),
    sheet_name: Optional[str] = Query(None, description="Sheet name to preview (default: active sheet)"),
):
    """Preview an uploaded Excel file: return headers, first 5 rows, and sheet list.

    Used by the relationship module to let users pick the sheet and article-number
    column before running analysis.
    """
    from openpyxl import load_workbook
    from io import BytesIO

    content = await file.read()
    try:
        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Kunne ikke lese Excel-fil: {e}")

    sheet_names = wb.sheetnames

    # Select requested sheet or default to active
    if sheet_name and sheet_name in sheet_names:
        ws = wb[sheet_name]
    elif sheet_name and sheet_name not in sheet_names:
        wb.close()
        raise HTTPException(400, f"Arket '{sheet_name}' finnes ikke i filen")
    else:
        ws = wb.active

    selected_title = ws.title
    result = _preview_sheet(ws)
    wb.close()

    if not result["headers"]:
        raise HTTPException(400, "Arket er tomt eller har ingen overskrifter")

    result["sheet_names"] = sheet_names
    result["active_sheet"] = selected_title
    return result


@app.post("/api/families/validate-upload-column")
async def validate_upload_column(
    file: UploadFile = File(...),
    column_index: int = Query(..., description="Column index for article numbers (0-based)"),
    sheet_name: Optional[str] = Query(None, description="Sheet name to read from"),
):
    """Validate a specific column in an uploaded Excel file.

    Extracts all values from the selected column, normalizes and deduplicates them,
    and returns the validated count and sample — without running family analysis.
    Used by the frontend to show a trustworthy article count before the user commits
    to running the analysis.
    """
    from openpyxl import load_workbook
    from io import BytesIO
    from backend.scraper import normalize_identifier

    content = await file.read()
    try:
        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Kunne ikke lese Excel-fil: {e}")

    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    raw_count = 0
    article_numbers = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # Skip header
        if row and len(row) > column_index:
            val = row[column_index]
            if val is not None and str(val).strip():
                raw_count += 1
            normalized = normalize_identifier(val)
            if normalized:
                article_numbers.append(normalized)
    wb.close()

    # Deduplicate
    seen = set()
    unique = []
    for a in article_numbers:
        if a not in seen:
            seen.add(a)
            unique.append(a)

    return {
        "raw_values": raw_count,
        "valid_unique": len(unique),
        "sample": unique[:10],
        "duplicates_removed": len(article_numbers) - len(unique),
    }


@app.post("/api/families/analyze-upload")
async def analyze_upload_families(
    file: UploadFile = File(...),
    column_index: Optional[int] = Query(None, description="Column index for article numbers (0-based)"),
    sheet_name: Optional[str] = Query(None, description="Sheet name to read from"),
):
    """Run family detection on an uploaded Excel file directly.

    Reads article numbers from the file, looks up Jeeves data for each,
    and runs family detection without requiring a full quality analysis.
    If column_index is provided, uses that column instead of auto-detection.
    If sheet_name is provided, reads from that sheet instead of the active sheet.
    """
    if not _jeeves_index or not _jeeves_index.loaded:
        raise HTTPException(400, "Jeeves-data er ikke lastet inn — kreves for familieanalyse")

    content = await file.read()

    if column_index is not None:
        # Manual column/sheet selection — read from specified column
        from openpyxl import load_workbook
        from io import BytesIO
        from backend.scraper import normalize_identifier

        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        article_numbers = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue  # Skip header
            if row and len(row) > column_index:
                val = row[column_index]
                normalized = normalize_identifier(val)
                if normalized:
                    article_numbers.append(normalized)
        wb.close()
        # Deduplicate
        seen = set()
        unique = []
        for a in article_numbers:
            if a not in seen:
                seen.add(a)
                unique.append(a)
        article_numbers = unique
    else:
        article_numbers, _ = read_article_numbers(content, file.filename or "upload.xlsx")

    if not article_numbers:
        raise HTTPException(400, "Ingen artikkelnumre funnet i valgt kolonne")

    # Build product dicts from Jeeves for the uploaded articles only
    products = []
    for artnr in article_numbers:
        j = _jeeves_index.get(artnr)
        if j:
            products.append({
                "article_number": j.article_number,
                "product_name": j.item_description or j.web_title or "",
                "brand": j.product_brand or "",
                "supplier": j.supplier or "",
                "specification": j.specification or "",
                "technical_details": {},
                "category": "",
            })

    if not products:
        raise HTTPException(400, "Ingen produkter funnet i Jeeves for de oppgitte artiklene")

    import hashlib
    source_id = f"upload-{hashlib.md5(content).hexdigest()[:8]}"
    return _run_family_detection(products, source_id, data_source="upload")


@app.post("/api/families/analyze-articles")
async def analyze_articles_families(data: dict):
    """Run family detection on a list of pasted article numbers.

    Body: {"article_numbers": ["12345", "67890", ...]}

    Looks up each article in Jeeves and runs family detection.
    """
    if not _jeeves_index or not _jeeves_index.loaded:
        raise HTTPException(400, "Jeeves-data er ikke lastet inn — kreves for familieanalyse")

    raw_articles = data.get("article_numbers", [])
    if not raw_articles:
        raise HTTPException(400, "Ingen artikkelnumre oppgitt")

    # Deduplicate and normalize
    seen = set()
    article_numbers = []
    for art in raw_articles:
        art = str(art).strip()
        if art and art not in seen:
            seen.add(art)
            article_numbers.append(art)

    if not article_numbers:
        raise HTTPException(400, "Ingen gyldige artikkelnumre etter validering")

    # Build product dicts from Jeeves
    products = []
    not_found = []
    for artnr in article_numbers:
        j = _jeeves_index.get(artnr)
        if j:
            products.append({
                "article_number": j.article_number,
                "product_name": j.item_description or j.web_title or "",
                "brand": j.product_brand or "",
                "supplier": j.supplier or "",
                "specification": j.specification or "",
                "technical_details": {},
                "category": "",
            })
        else:
            not_found.append(artnr)

    if not products:
        raise HTTPException(400, f"Ingen produkter funnet i Jeeves for de oppgitte artiklene")

    import hashlib
    hash_input = ",".join(sorted(seen))
    source_id = f"articles-{hashlib.md5(hash_input.encode()).hexdigest()[:8]}"
    result = _run_family_detection(products, source_id, data_source="pasted_articles")
    result["not_found_articles"] = not_found
    result["input_count"] = len(article_numbers)
    return result


@app.put("/api/families/{source_id}/review/{family_id}")
async def update_family_review(source_id: str, family_id: str, data: dict):
    """Update review status and comment for a single family.

    Body: {"review_status": "accepted"|"rejected"|"needs_review"|"pending",
           "review_comment": "optional note",
           "family_name": "optional edited name"}
    """
    result = _load_family_results(source_id)
    family = next((f for f in result["families"] if f["family_id"] == family_id), None)
    if not family:
        raise HTTPException(404, f"Familie {family_id} ikke funnet")

    valid_statuses = {"pending", "accepted", "rejected", "needs_review"}
    new_status = data.get("review_status")
    if new_status and new_status not in valid_statuses:
        raise HTTPException(400, f"Ugyldig status: {new_status}. Gyldige: {valid_statuses}")

    if new_status:
        family["review_status"] = new_status
    if "review_comment" in data:
        family["review_comment"] = data["review_comment"]
    if "family_name" in data and data["family_name"]:
        family["family_name"] = data["family_name"]

    _persist_family_to_disk(source_id)
    return {"status": "ok", "family_id": family_id, "review_status": family.get("review_status")}


@app.get("/api/families/{source_id}/export")
async def export_families(source_id: str):
    """Export family results as an Excel file."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    data = _load_family_results(source_id)
    wb = Workbook()

    # Sheet 1: Families
    ws = wb.active
    ws.title = "Produktfamilier"
    headers = [
        "Familie_ID", "Familienavn", "Antall", "Variantdimensjoner",
        "Mor_Artikkel", "Konfidensgrad", "Gjennomgang_status",
        "Kommentar", "Grupperingsgrunn",
    ]
    hfont = Font(bold=True, color="FFFFFF", size=10)
    hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = hfont
        c.fill = hfill
    for row_idx, f in enumerate(data["families"], 2):
        ws.cell(row=row_idx, column=1, value=f["family_id"])
        ws.cell(row=row_idx, column=2, value=f["family_name"])
        ws.cell(row=row_idx, column=3, value=f["member_count"])
        ws.cell(row=row_idx, column=4, value=", ".join(f.get("variant_dimensions", [])))
        ws.cell(row=row_idx, column=5, value=f.get("mother_article") or "(abstrakt)")
        ws.cell(row=row_idx, column=6, value=f["confidence"])
        ws.cell(row=row_idx, column=7, value=f.get("review_status", "pending"))
        ws.cell(row=row_idx, column=8, value=f.get("review_comment", ""))
        ws.cell(row=row_idx, column=9, value=f["grouping_reason"])
    ws.freeze_panes = "A2"

    # Sheet 2: Members
    ws2 = wb.create_sheet("Familiemedlemmer")
    mheaders = [
        "Familie_ID", "Familienavn", "Artikkelnummer", "Rolle",
        "Produktnavn", "Spesifikasjon", "Barnespesifikt", "Varianter",
    ]
    for i, h in enumerate(mheaders, 1):
        c = ws2.cell(row=1, column=i, value=h)
        c.font = hfont
        c.fill = hfill
    row = 2
    for f in data["families"]:
        for m in f["members"]:
            ws2.cell(row=row, column=1, value=f["family_id"])
            ws2.cell(row=row, column=2, value=f["family_name"])
            ws2.cell(row=row, column=3, value=m["article_number"])
            ws2.cell(row=row, column=4, value=m["role"])
            ws2.cell(row=row, column=5, value=m.get("product_name", ""))
            ws2.cell(row=row, column=6, value=m.get("specification", ""))
            ws2.cell(row=row, column=7, value=m.get("child_specific_title", ""))
            dims = ", ".join(f'{d["name"]}={d["value"]}' for d in m.get("variant_dimensions", []))
            ws2.cell(row=row, column=8, value=dims)
            row += 1
    ws2.freeze_panes = "A2"

    # Sheet 3: Standalone
    ws3 = wb.create_sheet("Frittstående")
    sheaders = ["Artikkelnummer", "Produktnavn", "Spesifikasjon", "Årsak", "Konfidensgrad"]
    for i, h in enumerate(sheaders, 1):
        c = ws3.cell(row=1, column=i, value=h)
        c.font = hfont
        c.fill = hfill
    for row_idx, s in enumerate(data.get("standalone", []), 2):
        ws3.cell(row=row_idx, column=1, value=s["article_number"])
        ws3.cell(row=row_idx, column=2, value=s.get("product_name", ""))
        ws3.cell(row=row_idx, column=3, value=s.get("specification", ""))
        ws3.cell(row=row_idx, column=4, value=s.get("grouping_reason", ""))
        ws3.cell(row=row_idx, column=5, value=s.get("confidence", ""))
    ws3.freeze_panes = "A2"

    # Save and return
    from io import BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    from starlette.responses import StreamingResponse
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=produktfamilier_{source_id}.xlsx"},
    )


def _load_family_results(source_id: str) -> dict:
    """Load family results from memory or disk. Raises HTTPException(404) if not found."""
    import json
    if source_id not in _family_results:
        disk_path = FAMILY_OUTPUT_DIR / f"{source_id}.json"
        if disk_path.exists():
            _family_results[source_id] = json.loads(disk_path.read_text(encoding="utf-8"))
        else:
            raise HTTPException(404, "Familieresultater ikke funnet")
    return _family_results[source_id]


def _persist_family_to_disk(source_id: str):
    """Write current in-memory family results to disk."""
    import json
    if source_id not in _family_results:
        return
    try:
        disk_path = FAMILY_OUTPUT_DIR / f"{source_id}.json"
        disk_path.write_text(
            json.dumps(_family_results[source_id], ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
    except Exception as e:
        logger.warning(f"Failed to persist family results: {e}")


@app.put("/api/families/{source_id}/bulk-review")
async def bulk_review_families(source_id: str, data: dict):
    """Bulk update review status for multiple families at once.

    Body: {"family_ids": ["FAM-abc12345", ...],
           "review_status": "accepted"|"rejected"|"needs_review"|"pending",
           "review_comment": "optional shared comment"}
    """
    result = _load_family_results(source_id)

    family_ids = data.get("family_ids", [])
    if not family_ids:
        raise HTTPException(400, "Ingen familie-IDer oppgitt")

    valid_statuses = {"pending", "accepted", "rejected", "needs_review"}
    new_status = data.get("review_status")
    if not new_status or new_status not in valid_statuses:
        raise HTTPException(400, f"Ugyldig status: {new_status}. Gyldige: {valid_statuses}")

    comment = data.get("review_comment")
    updated = 0
    for fam in result["families"]:
        if fam["family_id"] in family_ids:
            fam["review_status"] = new_status
            if comment is not None:
                fam["review_comment"] = comment
            updated += 1

    _persist_family_to_disk(source_id)
    return {"status": "ok", "updated": updated, "review_status": new_status}


@app.put("/api/families/{source_id}/move-member")
async def move_family_member(source_id: str, data: dict):
    """Move a product between families or to/from standalone.

    Body: {"article_number": "12345",
           "from_family_id": "FAM-xxx" or "standalone",
           "to_family_id": "FAM-yyy" or "standalone"}

    Records the move as a manual override for traceability.
    """
    result = _load_family_results(source_id)

    article = data.get("article_number")
    from_id = data.get("from_family_id")
    to_id = data.get("to_family_id")
    if not article or not from_id or not to_id:
        raise HTTPException(400, "article_number, from_family_id og to_family_id er påkrevd")
    if from_id == to_id:
        raise HTTPException(400, "Kilde og mål kan ikke være like")

    # Find the member to move
    member = None
    if from_id == "standalone":
        idx = next((i for i, s in enumerate(result.get("standalone", [])) if s["article_number"] == article), None)
        if idx is None:
            raise HTTPException(404, f"Produkt {article} ikke funnet i frittstående")
        s = result["standalone"].pop(idx)
        member = {
            "article_number": s["article_number"],
            "role": "child",
            "product_name": s.get("product_name", ""),
            "specification": s.get("specification", ""),
            "child_specific_title": "",
            "variant_dimensions": [],
            "manually_moved": True,
        }
    else:
        src_family = next((f for f in result["families"] if f["family_id"] == from_id), None)
        if not src_family:
            raise HTTPException(404, f"Kildefamilie {from_id} ikke funnet")
        midx = next((i for i, m in enumerate(src_family["members"]) if m["article_number"] == article), None)
        if midx is None:
            raise HTTPException(404, f"Produkt {article} ikke funnet i familie {from_id}")
        member = src_family["members"].pop(midx)
        member["manually_moved"] = True
        member["role"] = "child"  # reset role when moving
        src_family["member_count"] = len(src_family["members"])
        # If source family is now empty, remove it
        if len(src_family["members"]) == 0:
            result["families"] = [f for f in result["families"] if f["family_id"] != from_id]

    # Place the member in destination
    if to_id == "standalone":
        result.setdefault("standalone", []).append({
            "article_number": member["article_number"],
            "product_name": member.get("product_name", ""),
            "specification": member.get("specification", ""),
            "grouping_reason": "Manuelt flyttet til frittstående",
            "confidence": 0,
            "manually_moved": True,
        })
    else:
        dst_family = next((f for f in result["families"] if f["family_id"] == to_id), None)
        if not dst_family:
            raise HTTPException(404, f"Målfamilie {to_id} ikke funnet")
        dst_family["members"].append(member)
        dst_family["member_count"] = len(dst_family["members"])

    # Update standalone count
    result["standalone_products"] = len(result.get("standalone", []))
    result["total_families"] = len(result["families"])

    # Track manual overrides
    result.setdefault("manual_overrides", []).append({
        "article_number": article,
        "from": from_id,
        "to": to_id,
        "action": "move_member",
    })

    _persist_family_to_disk(source_id)
    return {
        "status": "ok",
        "article_number": article,
        "moved_from": from_id,
        "moved_to": to_id,
    }


@app.put("/api/families/{source_id}/merge")
async def merge_families(source_id: str, data: dict):
    """Merge one family into another.

    Body: {"source_family_id": "FAM-xxx", "target_family_id": "FAM-yyy"}

    All members of the source family are moved into the target family.
    The source family is removed. The operation is recorded as an undoable
    override event.
    """
    result = _load_family_results(source_id)

    src_id = data.get("source_family_id")
    tgt_id = data.get("target_family_id")
    if not src_id or not tgt_id:
        raise HTTPException(400, "source_family_id og target_family_id er påkrevd")
    if src_id == tgt_id:
        raise HTTPException(400, "Kilde og mål kan ikke være like")

    src = next((f for f in result["families"] if f["family_id"] == src_id), None)
    tgt = next((f for f in result["families"] if f["family_id"] == tgt_id), None)
    if not src:
        raise HTTPException(404, f"Kildefamilie {src_id} ikke funnet")
    if not tgt:
        raise HTTPException(404, f"Målfamilie {tgt_id} ikke funnet")

    # Snapshot for undo
    moved_articles = [m["article_number"] for m in src["members"]]
    import copy
    override_event = {
        "action": "merge",
        "source_family_id": src_id,
        "target_family_id": tgt_id,
        "moved_articles": moved_articles,
        "source_family_snapshot": copy.deepcopy(src),
    }

    # Move all members
    for m in src["members"]:
        m["manually_moved"] = True
        tgt["members"].append(m)
    tgt["member_count"] = len(tgt["members"])

    # Remove source family
    result["families"] = [f for f in result["families"] if f["family_id"] != src_id]
    result["total_families"] = len(result["families"])

    result.setdefault("manual_overrides", []).append(override_event)
    _persist_family_to_disk(source_id)
    return {
        "status": "ok",
        "merged": src_id,
        "into": tgt_id,
        "members_moved": len(moved_articles),
    }


@app.put("/api/families/{source_id}/split")
async def split_family(source_id: str, data: dict):
    """Split selected members out of a family into a new family or standalone.

    Body: {"family_id": "FAM-xxx",
           "article_numbers": ["ART001", "ART002"],
           "target": "new_family" or "standalone",
           "new_family_name": "optional name for new family"}
    """
    result = _load_family_results(source_id)

    fam_id = data.get("family_id")
    articles = data.get("article_numbers", [])
    target = data.get("target", "new_family")
    new_name = data.get("new_family_name", "")

    if not fam_id or not articles:
        raise HTTPException(400, "family_id og article_numbers er påkrevd")

    fam = next((f for f in result["families"] if f["family_id"] == fam_id), None)
    if not fam:
        raise HTTPException(404, f"Familie {fam_id} ikke funnet")

    articles_set = set(articles)
    to_split = [m for m in fam["members"] if m["article_number"] in articles_set]
    if not to_split:
        raise HTTPException(404, "Ingen av de angitte artiklene ble funnet i familien")
    remaining = [m for m in fam["members"] if m["article_number"] not in articles_set]

    if not remaining:
        raise HTTPException(400, "Kan ikke flytte alle medlemmer — bruk sletting eller flytt i stedet")

    import copy
    override_event = {
        "action": "split",
        "source_family_id": fam_id,
        "split_articles": list(articles_set),
        "target": target,
        "pre_split_members": copy.deepcopy(fam["members"]),
    }

    # Update source family
    fam["members"] = remaining
    fam["member_count"] = len(remaining)

    new_family_id = None
    if target == "standalone":
        for m in to_split:
            result.setdefault("standalone", []).append({
                "article_number": m["article_number"],
                "product_name": m.get("product_name", ""),
                "specification": m.get("specification", ""),
                "grouping_reason": f"Manuelt splittet fra {fam_id}",
                "confidence": 0,
                "manually_moved": True,
            })
    else:
        # Create new family
        import hashlib
        hash_input = f"manual-split-{fam_id}-{'-'.join(sorted(articles_set))}"
        new_family_id = f"FAM-{hashlib.sha256(hash_input.encode()).hexdigest()[:16]}"
        new_fam = {
            "family_id": new_family_id,
            "family_name": new_name or f"{fam['family_name']} (splittet)",
            "member_count": len(to_split),
            "variant_dimensions": [],
            "mother_article": None,
            "confidence": fam["confidence"],
            "review_required": True,
            "grouping_reason": f"Manuelt splittet fra {fam_id}",
            "review_status": "needs_review",
            "review_comment": "",
            "manually_created": True,
            "members": [],
        }
        for m in to_split:
            m["manually_moved"] = True
            m["role"] = "child"
            new_fam["members"].append(m)
        result["families"].append(new_fam)

    override_event["new_family_id"] = new_family_id
    result["standalone_products"] = len(result.get("standalone", []))
    result["total_families"] = len(result["families"])
    result.setdefault("manual_overrides", []).append(override_event)
    _persist_family_to_disk(source_id)
    return {
        "status": "ok",
        "split_from": fam_id,
        "articles_moved": len(to_split),
        "target": target,
        "new_family_id": new_family_id,
    }


@app.put("/api/families/{source_id}/undo")
async def undo_last_override(source_id: str):
    """Undo the most recent manual override action.

    Reverses the last entry in the manual_overrides list by replaying the
    stored snapshot data. Supports undo of: move_member, merge, split.
    """
    result = _load_family_results(source_id)
    overrides = result.get("manual_overrides", [])
    if not overrides:
        raise HTTPException(400, "Ingen manuelle endringer å angre")

    last = overrides.pop()
    action = last.get("action")

    if action == "move_member":
        article = last["article_number"]
        original_from = last["from"]
        original_to = last["to"]

        # Reverse: remove from original_to, place back in original_from
        member = None
        if original_to == "standalone":
            idx = next((i for i, s in enumerate(result.get("standalone", []))
                        if s["article_number"] == article), None)
            if idx is not None:
                s = result["standalone"].pop(idx)
                member = {
                    "article_number": s["article_number"],
                    "role": "child",
                    "product_name": s.get("product_name", ""),
                    "specification": s.get("specification", ""),
                    "child_specific_title": "",
                    "variant_dimensions": [],
                }
        else:
            dst_fam = next((f for f in result["families"] if f["family_id"] == original_to), None)
            if dst_fam:
                midx = next((i for i, m in enumerate(dst_fam["members"])
                             if m["article_number"] == article), None)
                if midx is not None:
                    member = dst_fam["members"].pop(midx)
                    member.pop("manually_moved", None)
                    dst_fam["member_count"] = len(dst_fam["members"])

        if member:
            if original_from == "standalone":
                result.setdefault("standalone", []).append({
                    "article_number": member["article_number"],
                    "product_name": member.get("product_name", ""),
                    "specification": member.get("specification", ""),
                    "grouping_reason": member.get("grouping_reason", ""),
                    "confidence": member.get("confidence", 0),
                })
            else:
                src_fam = next((f for f in result["families"] if f["family_id"] == original_from), None)
                if src_fam:
                    member.pop("manually_moved", None)
                    src_fam["members"].append(member)
                    src_fam["member_count"] = len(src_fam["members"])

    elif action == "merge":
        # Restore the source family from snapshot
        snapshot = last.get("source_family_snapshot")
        tgt_id = last["target_family_id"]
        moved_articles = set(last.get("moved_articles", []))
        if snapshot:
            # Remove moved members from target
            tgt = next((f for f in result["families"] if f["family_id"] == tgt_id), None)
            if tgt:
                tgt["members"] = [m for m in tgt["members"]
                                  if m["article_number"] not in moved_articles]
                tgt["member_count"] = len(tgt["members"])
            # Restore source family
            for m in snapshot.get("members", []):
                m.pop("manually_moved", None)
            result["families"].append(snapshot)

    elif action == "split":
        src_fam_id = last["source_family_id"]
        pre_members = last.get("pre_split_members", [])
        split_articles = set(last.get("split_articles", []))
        target = last.get("target")
        new_fam_id = last.get("new_family_id")

        # Restore source family members
        src_fam = next((f for f in result["families"] if f["family_id"] == src_fam_id), None)
        if src_fam and pre_members:
            for m in pre_members:
                m.pop("manually_moved", None)
            src_fam["members"] = pre_members
            src_fam["member_count"] = len(pre_members)

        # Remove new family if it was created
        if target == "new_family" and new_fam_id:
            result["families"] = [f for f in result["families"] if f["family_id"] != new_fam_id]
        elif target == "standalone":
            result["standalone"] = [
                s for s in result.get("standalone", [])
                if s["article_number"] not in split_articles
            ]

    # Update counts
    result["standalone_products"] = len(result.get("standalone", []))
    result["total_families"] = len(result["families"])
    _persist_family_to_disk(source_id)
    return {"status": "ok", "undone_action": action}


@app.get("/api/families/{source_id}/standalone")
async def get_standalone_page(
    source_id: str,
    offset: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=200),
    search: str = Query("", description="Search in article_number or product_name"),
):
    """Paginated standalone products with optional search."""
    result = _load_family_results(source_id)
    standalone = result.get("standalone", [])

    if search:
        q = search.lower()
        standalone = [
            s for s in standalone
            if q in s.get("article_number", "").lower()
            or q in s.get("product_name", "").lower()
            or q in s.get("specification", "").lower()
        ]

    total = len(standalone)
    page = standalone[offset:offset + limit]
    return {
        "total": total,
        "offset": offset,
        "limit": limit,
        "items": page,
    }


@app.post("/api/score-product")
async def score_product_endpoint(data: dict):
    """Score a single product using Claude AI.

    Input: {product_name, description, specification, category, packaging}
    Returns: {overall_score, field_scores, issues, improvement_suggestions}
    """
    result = await score_product_async(
        product_name=data.get("product_name"),
        description=data.get("description"),
        specification=data.get("specification"),
        category=data.get("category"),
        packaging=data.get("packaging"),
    )
    if result is None:
        raise HTTPException(
            503,
            "AI-scoring er ikke tilgjengelig. Sjekk at ANTHROPIC_API_KEY er satt."
        )
    return result


@app.post("/api/enrich-product")
async def enrich_product_endpoint(data: dict):
    """Get AI-powered enrichment suggestions for a product.

    Input: {product_name, description, specification, category, packaging}
    Returns: {improved_description, missing_specifications, suggested_category, packaging_suggestions}
    """
    result = await enrich_product_async(
        product_name=data.get("product_name"),
        description=data.get("description"),
        specification=data.get("specification"),
        category=data.get("category"),
        packaging=data.get("packaging"),
    )
    if result is None:
        raise HTTPException(
            503,
            "AI-berikelse er ikke tilgjengelig. Sjekk at ANTHROPIC_API_KEY er satt."
        )
    return result


@app.post("/api/batch-evaluate")
async def batch_evaluate_endpoint(data: dict):
    """Evaluate a batch of products and return scores + flags.

    Input: {products: [{product_name, description, specification, category, packaging}, ...]}
    Returns: {results: [{article_number, score, flag, issues, ...}, ...]}
    """
    products = data.get("products", [])
    if not products:
        raise HTTPException(400, "Ingen produkter å evaluere")
    if len(products) > 50:
        raise HTTPException(400, "Maks 50 produkter per batch-evaluering")

    results = []
    for product in products:
        # Rule-based score (from analyzer)
        rule_score = 0.0
        issues = []

        # Quick rule-based checks
        if not product.get("product_name"):
            issues.append("Produktnavn mangler")
        if not product.get("description"):
            issues.append("Beskrivelse mangler")
        if not product.get("specification"):
            issues.append("Spesifikasjon mangler")
        if not product.get("category"):
            issues.append("Kategori mangler")
        if not product.get("packaging"):
            issues.append("Pakningsinfo mangler")

        fields_present = 5 - len(issues)
        rule_score = fields_present / 5 * 100

        # AI score (if available)
        ai_result = await score_product_async(
            product_name=product.get("product_name"),
            description=product.get("description"),
            specification=product.get("specification"),
            category=product.get("category"),
            packaging=product.get("packaging"),
        )

        ai_score = ai_result["overall_score"] if ai_result else None
        ai_issues = ai_result.get("issues", []) if ai_result else []
        ai_suggestions = ai_result.get("improvement_suggestions", []) if ai_result else []

        # Combined score: weight rule-based 40%, AI 60% (if available)
        if ai_score is not None:
            combined = rule_score * 0.4 + ai_score * 0.6
        else:
            combined = rule_score

        # Flag
        if combined >= 70:
            flag = "OK"
        elif combined >= 40:
            flag = "Needs review"
        else:
            flag = "Critical"

        results.append({
            "article_number": product.get("article_number", ""),
            "product_name": product.get("product_name"),
            "rule_score": round(rule_score, 1),
            "ai_score": round(ai_score, 1) if ai_score is not None else None,
            "combined_score": round(combined, 1),
            "flag": flag,
            "issues": issues + ai_issues,
            "improvement_suggestions": ai_suggestions,
        })

    return {"results": results}


def _build_image_suggestion(
    product_data: ProductData,
    image_summary,
    mfr_data,
    norengros_data,
) -> Optional[ImageSuggestion]:
    """Build an image improvement suggestion if the current image is weak/missing.

    Priority: manufacturer image > Norengros image.
    Manufacturer images can be auto-suggested; Norengros always requires review.
    """
    current_url = product_data.image_url
    current_status = "ok"

    # Determine current image status
    if not image_summary or not image_summary.main_image_exists:
        current_status = "missing"
    elif image_summary.image_analyses:
        main = image_summary.image_analyses[0]
        score = main.overall_score
        if score < 40:
            current_status = "low_quality"
        elif score < 70:
            issues = main.issues or []
            if any("background" in str(i).lower() for i in issues):
                current_status = "poor_background"
            elif any("resolution" in str(i).lower() for i in issues):
                current_status = "low_quality"
            else:
                current_status = "review"
        # OK images don't need suggestions
        else:
            return None

    if current_status == "ok":
        return None

    # Check manufacturer image first (preferred)
    if mfr_data and mfr_data.found and mfr_data.image_url:
        return ImageSuggestion(
            current_image_url=current_url,
            current_image_status=current_status,
            suggested_image_url=mfr_data.image_url,
            suggested_source="manufacturer",
            suggested_source_url=mfr_data.source_url,
            confidence=mfr_data.confidence * 0.8,
            review_required=True,
            reason=f"Produsentbilde funnet ({current_status}). Verifiser produktmatch.",
        )

    # Fallback: Norengros image (always conservative)
    if norengros_data and norengros_data.found and norengros_data.image_url:
        return ImageSuggestion(
            current_image_url=current_url,
            current_image_status=current_status,
            suggested_image_url=norengros_data.image_url,
            suggested_source="norengros",
            suggested_source_url=norengros_data.source_url,
            confidence=0.25,
            review_required=True,
            reason=f"Norengros-bilde funnet ({current_status}). Konkurrentkilde — krever manuell godkjenning.",
        )

    # No better image available — report the issue with producer search suggestion
    if current_status != "ok":
        # Build a producer search hint using available catalog info
        search_hint = None
        producer = product_data.manufacturer
        supplier_item = product_data.manufacturer_article_number
        if producer:
            search_terms = [producer]
            if supplier_item:
                search_terms.append(supplier_item)
            elif product_data.product_name:
                search_terms.append(product_data.product_name)
            search_hint = " ".join(search_terms)

        reason = f"Bildestatus: {current_status}. Ingen bedre bildekilde funnet automatisk."
        if search_hint:
            reason += f" Foreslått søk hos produsent: \"{search_hint}\""

        # Always include the CDN image URL so the user has a reference.
        # For producer_search suggestions, suggested_image_url points to the
        # current CDN image (the best we have), and reason explains what to do.
        cdn_url = current_url or (
            f"https://res.onemed.com/NO/ARWebBig/{product_data.article_number}.jpg"
        )

        return ImageSuggestion(
            current_image_url=cdn_url,
            current_image_status=current_status,
            suggested_image_url=cdn_url,  # Best available image URL
            suggested_source="producer_search" if producer else "current_cdn",
            suggested_source_url=product_data.product_url,  # Link to product page if available
            confidence=0.0,
            review_required=True,
            reason=reason,
        )

    return None


def _apply_enrichment_to_analysis(analysis: ProductAnalysis) -> None:
    """Apply enrichment results to field analyses.

    Enrichment from internal PDF takes priority over manufacturer sources.
    Only applies suggestions where confidence is sufficient and no conflict.
    """
    if not analysis.enrichment_results:
        return

    # Map enrichment field names to FieldAnalysis field names
    field_map = {
        "product_name": "Produktnavn",
        "description": "Beskrivelse",
        "manufacturer": "Produsent",
        "manufacturer_article_number": "Produsentens varenummer",
        "packaging_info": "Pakningsinformasjon",
        "specifications": "Spesifikasjon",
        "category": "Kategori",
    }

    for er in analysis.enrichment_results:
        if not er.suggested_value or er.match_status == "NOT_FOUND":
            continue

        fa_name = field_map.get(er.field_name)
        if not fa_name:
            continue

        # Only apply if confidence >= 0.6 and not a conflict requiring review
        if er.confidence < 0.6 or er.review_status == "conflict":
            continue

        # Find matching field analysis
        for fa in analysis.field_analyses:
            if fa.field_name == fa_name:
                # Only suggest if field is currently missing/poor and enrichment has value
                if fa.status in (QualityStatus.MISSING, QualityStatus.WEAK, QualityStatus.SHOULD_IMPROVE, QualityStatus.PROBABLE_ERROR, QualityStatus.MANUAL_REVIEW):
                    # Don't overwrite existing manufacturer suggestion with lower-confidence PDF
                    if fa.suggested_value and fa.confidence and fa.confidence > er.confidence:
                        continue
                    # PDF source takes priority
                    if er.source_level == "internal_product_sheet":
                        fa.suggested_value = er.suggested_value
                        fa.source = f"Produktdatablad ({er.source_url})"
                        fa.confidence = er.confidence
                    elif not fa.suggested_value:
                        fa.suggested_value = er.suggested_value
                        fa.source = f"Produsent ({er.source_url})"
                        fa.confidence = er.confidence
                break


def _format_image_quality(iq: Optional[dict]) -> dict:
    """Format image quality data for API response."""
    if not iq:
        return {"score": 0, "status": "MISSING", "count": 0, "main_exists": False, "issues": ""}
    return {
        "score": iq.get("avg_image_score", 0),
        "status": iq.get("image_quality_status", "MISSING"),
        "count": iq.get("image_count_found", 0),
        "main_exists": iq.get("main_image_exists", False),
        "issues": iq.get("image_issue_summary", ""),
    }


async def _run_analysis(
    job_id: str,
    article_numbers: list[str],
    skip_cache: bool = False,
    analysis_mode: str = AnalysisMode.FULL_ENRICHMENT.value,
    focus_areas: Optional[list[str]] = None,
) -> None:
    """Run the analysis pipeline for all products.

    Pipeline steps depend on analysis_mode:
    - full_enrichment: all steps (scrape, images, mfr, PDF, enrichment, AI)
    - audit_only: scrape + images + quality analysis + scoring (no enrichment/AI suggestions)
    - focused_scan: only evaluate selected focus_areas, skip irrelevant steps
    """
    is_audit = analysis_mode == AnalysisMode.AUDIT_ONLY.value
    is_focused = analysis_mode == AnalysisMode.FOCUSED_SCAN.value
    skip_enrichment = is_audit or is_focused
    job = jobs[job_id]
    job.status = JobStatus.RUNNING

    # --- Deduplicate article numbers to avoid duplicate HTTP requests ---
    # Preserve original order and count for output mapping
    unique_articles: list[str] = []
    seen: set[str] = set()
    duplicate_count = 0
    for art in article_numbers:
        if art not in seen:
            seen.add(art)
            unique_articles.append(art)
        else:
            duplicate_count += 1

    if duplicate_count:
        logger.info(
            f"[{job_id}] Dedup: {len(article_numbers)} rows → "
            f"{len(unique_articles)} unique articles "
            f"({duplicate_count} duplicate fetches prevented)"
        )

    # --- Batch product discovery ---
    # Before analysis, find product page URLs for all products not yet in the
    # SKU index. This scans sitemap pages looking specifically for the products
    # in this batch. Stops as soon as all targets are found or sitemap is exhausted.
    # Also builds the general index as a side effect for future runs.
    try:
        job.current_step = "finding_product_pages"
        missing_from_index = {
            a for a in unique_articles if a.strip() not in _sku_to_url
        }
        if missing_from_index:
            logger.info(
                f"[{job_id}] {len(missing_from_index)}/{len(unique_articles)} products "
                f"not in SKU index — scanning sitemap to find their pages"
            )
            found_urls = await find_batch_products_in_sitemap(missing_from_index)
            logger.info(
                f"[{job_id}] Batch discovery found {len(found_urls)} product pages"
            )
        else:
            logger.info(f"[{job_id}] All {len(unique_articles)} products already in SKU index")
    except Exception as e:
        logger.warning(f"[{job_id}] Batch product discovery failed (non-fatal): {e}")

    # Scope logging — verify strict input processing
    try:
        in_index = sum(1 for a in unique_articles if a.strip() in _sku_to_url)
    except Exception:
        in_index = 0
    logger.info(
        f"[{job_id}] SCOPE: "
        f"input_rows={len(article_numbers)} | "
        f"unique_articles={len(unique_articles)} | "
        f"in_sku_index={in_index} | "
        f"not_in_index={len(unique_articles) - in_index}"
    )

    semaphore = asyncio.Semaphore(SCRAPE_CONCURRENCY)
    _fetch_count = 0  # Track actual URL fetches for scope verification

    async def process_product(article_number: str) -> Optional[ProductAnalysis]:
        nonlocal _fetch_count
        # Check if job was cancelled
        if job.cancelled:
            return None

        async with semaphore:
            if job.cancelled:
                return None

            job.current_product = article_number
            try:
                # Step 1: Scrape from onemed.no (strict input mode — no sitemap scan)
                job.current_step = "scraping"
                logger.info(f"[{job_id}] Scraping {article_number}")
                product_data = await scrape_product(
                    article_number, use_cache=not skip_cache,
                    enable_discovery=False,
                )
                _fetch_count += 1

                # Step 2: Analyze product images with CV
                # Always run - CDN URLs use article numbers directly, no scraping needed
                image_quality_dict = None
                job.current_step = "image_analysis"
                logger.info(f"[{job_id}] Image analysis for {article_number}")
                image_summary = await analyze_product_images(article_number)
                image_quality_dict = image_summary.to_dict()

                # Update product_data with image info from CV analysis
                product_data.image_quality_ok = image_summary.main_image_exists
                if image_summary.main_image_exists and image_summary.image_analyses:
                    product_data.image_url = image_summary.image_analyses[0].image_url

                # If scraper failed but images exist on CDN, mark as CDN-only (weak verification)
                # This allows the enrichment pipeline to continue, but does NOT confirm identity.
                if not product_data.found_on_onemed and image_summary.main_image_exists:
                    product_data.found_on_onemed = True
                    product_data.error = None
                    # Guard: if product already has web-sourced data (product_url, name,
                    # description from page parsing), do NOT downgrade to CDN_ONLY.
                    # This prevents overwriting a page-verified status when the scraper
                    # found the page but marked found_on_onemed=False (e.g., mismatch).
                    has_web_data = bool(
                        product_data.product_url
                        or product_data.product_name
                        or product_data.description
                    )
                    # Only downgrade verification — never upgrade from a confirmed mismatch
                    # or overwrite when the product page was actually found
                    if product_data.verification_status not in (
                        VerificationStatus.MISMATCH,
                        VerificationStatus.EXACT_MATCH,
                        VerificationStatus.NORMALIZED_MATCH,
                        VerificationStatus.SKU_IN_PAGE,
                    ) and not has_web_data:
                        product_data.verification_status = VerificationStatus.CDN_ONLY
                        product_data.verification_evidence = (
                            f"Produktbilde funnet i bildekatalogen for '{article_number}'. "
                            f"Ingen produktside med detaljer ble funnet eller hentet. "
                            f"Produktidentiteten er usikker — vurder manuelt."
                        )
                    elif has_web_data:
                        logger.info(
                            f"[{job_id}] {article_number} has web data but found_on_onemed was False. "
                            f"Keeping verification status: {product_data.verification_status.value}"
                        )
                    logger.info(
                        f"[{job_id}] {article_number} found via CDN image "
                        f"(verification: {product_data.verification_status.value})"
                    )

                # Step 3-5: enrichment pipeline (skipped in audit/focused modes)
                mfr_data = None
                norengros_data = None
                enrichment_results = []
                pdf_exists = False
                pdf_url = None

                if not skip_enrichment:
                    # Step 3: Manufacturer lookup (only if product found and data incomplete)
                    # Check Jeeves first — if Jeeves has supplier info, skip mfr lookup for that
                    jeeves_check = _jeeves_index.get(article_number) if _jeeves_index else None
                    if product_data.found_on_onemed:
                        has_mfr = product_data.manufacturer or (jeeves_check and jeeves_check.supplier)
                        has_mfr_num = product_data.manufacturer_article_number or (jeeves_check and jeeves_check.supplier_item_no)
                        has_spec = product_data.specification or product_data.technical_details
                        needs_mfr = not has_mfr or not has_mfr_num or not has_spec
                        if needs_mfr:
                            job.current_step = "manufacturer_lookup"
                            logger.info(f"[{job_id}] Manufacturer lookup for {article_number}")
                            mfr_data = await search_manufacturer_info(product_data)

                    # Step 4: PDF enrichment - always try (PDF CDN uses article numbers directly)
                    job.current_step = "pdf_enrichment"
                    logger.info(f"[{job_id}] Enrichment pipeline for {article_number}")
                    pdf_exists, pdf_url, enrichment_results = await run_enrichment_pipeline(
                        article_number, product_data, manufacturer_data=mfr_data
                    )

                    # Step 4b: Norengros secondary lookup
                    # Triggered when primary enrichment sources are insufficient.
                    # Previous gate was too restrictive (required ALL of: no PDF, no mfr,
                    # no description) — meaning Norengros was almost never reached.
                    # Now: trigger if any key field is still missing after primary sources.
                    has_good_description = (
                        product_data.description and len(product_data.description or "") >= 50
                    )
                    has_good_spec = bool(
                        product_data.specification or product_data.technical_details
                    )
                    primary_enrichment_sufficient = (
                        has_good_description and has_good_spec
                        and (pdf_exists or (mfr_data and mfr_data.found))
                    )
                    if not primary_enrichment_sufficient and product_data.found_on_onemed:
                        job.current_step = "norengros_lookup"
                        logger.info(f"[{job_id}] Norengros secondary lookup for {article_number}")
                        try:
                            norengros_data = await search_norengros(product_data)
                        except Exception as e:
                            logger.debug(f"[{job_id}] Norengros lookup failed for {article_number}: {e}")
                else:
                    # Audit/focused: still check PDF existence for document scoring
                    # but don't run the full enrichment pipeline
                    job.current_step = "pdf_check"
                    try:
                        from backend.pdf_enricher import check_pdf_exists
                        pdf_exists, pdf_url = await check_pdf_exists(article_number)
                    except Exception:
                        pass  # PDF check is best-effort in audit mode

                # Step 5: Quality analysis (with all collected data + Jeeves)
                job.current_step = "quality_analysis"
                jeeves_data = _jeeves_index.get(article_number) if _jeeves_index else None
                analysis = analyze_product(
                    product_data, image_quality=image_quality_dict, jeeves=jeeves_data
                )
                analysis.image_quality = image_quality_dict
                analysis.enrichment_results = enrichment_results
                analysis.pdf_available = pdf_exists
                analysis.pdf_url = pdf_url

                # Step 5a: Centralized area scoring (all modes)
                job.current_step = "area_scoring"
                score_result = score_product_areas(
                    product_data, jeeves_data, image_quality_dict,
                    pdf_available=pdf_exists,
                    areas=focus_areas if is_focused else None,
                )
                # Store area scores in ai_score dict for backward-compatible transport
                # (existing results endpoint already serializes ai_score)
                if not analysis.ai_score:
                    analysis.ai_score = {}
                analysis.ai_score["area_scores"] = score_result.to_dict()

                if not skip_enrichment:
                    # Apply manufacturer data if found
                    if mfr_data:
                        analysis.manufacturer_lookup = mfr_data
                        if mfr_data.found:
                            suggestions = generate_improvement_suggestions(
                                product_data, mfr_data
                            )
                            for suggestion in suggestions:
                                for fa in analysis.field_analyses:
                                    if fa.field_name == suggestion["field"]:
                                        fa.suggested_value = suggestion["suggested"]
                                        fa.source = suggestion["source"]
                                        fa.confidence = suggestion["confidence"]
                                        break

                    # Store Norengros data if found
                    if norengros_data:
                        analysis.norengros_lookup = norengros_data

                    # Step 5b: Image suggestion logic
                    # Also pass Jeeves data to enrich producer search hints
                    analysis.image_suggestion = _build_image_suggestion(
                        product_data, image_summary, mfr_data, norengros_data,
                    )
                    # Step 5b+: For low-quality images without a suggested replacement,
                    # try to find a better image from the producer if we have catalog info
                    if (analysis.image_suggestion
                            and analysis.image_suggestion.current_image_status in ("low_quality", "missing")
                            and not analysis.image_suggestion.suggested_image_url
                            and jeeves_data):
                        # Use Jeeves supplier + supplier_item_no for a targeted search hint
                        producer = jeeves_data.supplier or product_data.manufacturer
                        supplier_item = jeeves_data.supplier_item_no or product_data.manufacturer_article_number
                        if producer and supplier_item:
                            analysis.image_suggestion.reason = (
                                f"Bildekvalitet lav ({analysis.image_suggestion.current_image_status}). "
                                f"Søk hos produsent anbefalt: {producer} (art.nr: {supplier_item}). "
                                f"Produsentens nettside bør prioriteres som bildekilde."
                            )
                            analysis.image_suggestion.suggested_source = "producer_search"

                    # Apply enrichment suggestions to field_analyses (PDF takes priority)
                    _apply_enrichment_to_analysis(analysis)

                    # Run source-priority enrichment engine
                    enrichment_suggestions = enrich_product(
                        analysis, enrichment_results, manufacturer_data=mfr_data
                    )

                    # Step 6: AI quality layer (if API key configured)
                    job.current_step = "ai_scoring"
                    try:
                        # 6a: AI scoring (quality evaluation)
                        ai_score_result = await score_product_async(
                            product_name=product_data.product_name,
                            description=product_data.description,
                            specification=product_data.specification,
                            category=product_data.category,
                            packaging=product_data.packaging_info or product_data.packaging_unit,
                        )
                        if ai_score_result:
                            # Merge with area scores already stored
                            analysis.ai_score.update(ai_score_result)
                            logger.info(
                                f"[{job_id}] AI score for {article_number}: "
                                f"{ai_score_result.get('overall_score', 'N/A')}"
                            )

                        # 6b: AI editorial review of enrichment suggestions
                        if enrichment_suggestions:
                            suggestions_for_review = [
                                {
                                    "field_name": s.field_name,
                                    "current_value": s.current_value,
                                    "suggested_value": s.suggested_value,
                                    "source": s.source,
                                    "evidence": s.evidence,
                                }
                                for s in enrichment_suggestions
                            ]
                            ai_reviews = await review_suggestions_async(
                                article_number=article_number,
                                product_name=product_data.product_name,
                                suggestions=suggestions_for_review,
                            )
                            if ai_reviews:
                                enrichment_suggestions = apply_ai_review_to_suggestions(
                                    enrichment_suggestions, ai_reviews
                                )

                        # 6c: Legacy AI enrichment removed (P0-1 safety fix).

                    except Exception as e:
                        logger.warning(f"[{job_id}] AI scoring/review failed for {article_number}: {e}")

                    # Step 7: Final quality gate
                    if enrichment_suggestions:
                        enrichment_suggestions = final_quality_gate(enrichment_suggestions)

                    # Apply surviving suggestions
                    if enrichment_suggestions:
                        analysis.enrichment_suggestions = enrichment_suggestions
                        apply_enrichment_suggestions(analysis, enrichment_suggestions)
                        logger.info(
                            f"[{job_id}] {article_number}: "
                            f"{len(enrichment_suggestions)} enrichment suggestion(s) after quality gate"
                        )

                return analysis

            except Exception as e:
                logger.error(f"[{job_id}] Error processing {article_number}: {e}")
                job.errors.append(f"{article_number}: {str(e)}")
                fallback_data = ProductData(
                    article_number=article_number,
                    error=str(e),
                )
                return ProductAnalysis(
                    article_number=article_number,
                    product_data=fallback_data,
                    overall_comment=f"Feil under analyse: {str(e)}",
                )
            finally:
                await asyncio.sleep(SCRAPE_DELAY)

    try:
        # Process only unique articles — duplicate rows reuse the same result
        tasks = [process_product(art) for art in unique_articles]
        result_map: dict[str, ProductAnalysis] = {}
        cache_hits = 0

        for coro in asyncio.as_completed(tasks):
            result = await coro
            if result is None:
                continue  # Job was cancelled
            result_map[result.article_number] = result
            job.processed_products += 1
            logger.info(
                f"[{job_id}] Progress: {job.processed_products}/{len(unique_articles)} "
                f"({result.article_number}: {result.overall_status.value})"
            )

        if job.cancelled:
            logger.info(f"[{job_id}] Analysis cancelled by user")
            return

        # Map results back to all original rows (including duplicates)
        for art in article_numbers:
            if art in result_map:
                job.results.append(result_map[art])

        # Update processed count to include duplicates for progress reporting
        job.processed_products = len(job.results)

        logger.info(
            f"[{job_id}] Fetch stats: {len(article_numbers)} input_rows, "
            f"{len(unique_articles)} unique_articles, "
            f"{_fetch_count} product_urls_fetched, "
            f"{duplicate_count} duplicates_prevented, "
            f"discovery=OFF"
        )
        # Scope safety check: fetched URLs should never exceed unique articles
        if _fetch_count > len(unique_articles) * 2:
            logger.error(
                f"[{job_id}] SCOPE VIOLATION: fetched {_fetch_count} URLs "
                f"for {len(unique_articles)} unique articles — possible scope leak"
            )

        # Results are already in input order from the article_numbers loop above.
        # No sort needed — duplicate rows preserve their original positions.

        # Selector health check — warn if key website fields missing in >50% of found products
        found_results = [r for r in job.results if r.product_data.found_on_onemed]
        if len(found_results) >= 3:
            no_desc = sum(1 for r in found_results if not r.product_data.description)
            no_spec = sum(1 for r in found_results if not r.product_data.specification and not r.product_data.technical_details)
            no_cat = sum(1 for r in found_results if not r.product_data.category and not r.product_data.category_breadcrumb)
            threshold = len(found_results) * 0.5
            if no_desc > threshold:
                msg = f"SELECTOR WARNING: {no_desc}/{len(found_results)} products missing description — website structure may have changed"
                logger.error(f"[{job_id}] {msg}")
                job.errors.append(msg)
            if no_spec > threshold:
                msg = f"SELECTOR WARNING: {no_spec}/{len(found_results)} products missing specification — website structure may have changed"
                logger.error(f"[{job_id}] {msg}")
                job.errors.append(msg)
            if no_cat > threshold:
                msg = f"SELECTOR WARNING: {no_cat}/{len(found_results)} products missing category — website structure may have changed"
                logger.error(f"[{job_id}] {msg}")
                job.errors.append(msg)

        # Generate output Excel
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"masterdata_kvalitetssjekk_{job_id}_{timestamp}.xlsx"
        output_path = str(OUTPUT_DIR / output_filename)

        create_output_excel(job.results, output_path,
                           analysis_mode=analysis_mode, focus_areas=focus_areas)
        job.output_file = output_path
        job.status = JobStatus.COMPLETED
        job.current_product = None

        # Persist to history for later retrieval
        _add_to_history(job, source_filename=job.source_filename)

        logger.info(f"[{job_id}] Analysis completed. Output: {output_path}")

    except Exception as e:
        logger.error(f"[{job_id}] Analysis failed: {e}")
        job.status = JobStatus.FAILED
        job.errors.append(f"Fatal error: {str(e)}")


if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 8000))
    uvicorn.run(app, host="0.0.0.0", port=port)
