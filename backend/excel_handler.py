"""Excel import/export handler for masterdata quality check."""

import logging
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from backend.content_validator import (
    get_best_producer_info,
    translate_to_norwegian_if_needed,
    validate_suggestion_output,
)
from backend.diff_display import build_field_diff, detect_change_scope, summarize_change_type
from backend.identifiers import normalize_identifier, normalize_identifier_strict
from backend.models import EnrichmentSuggestion, JeevesData, ProductAnalysis, QualityStatus

logger = logging.getLogger(__name__)

# Color scheme for statuses — visually distinct for user prioritization
STATUS_COLORS = {
    # Good (greens)
    QualityStatus.STRONG: "92D050",              # Bright green — excellent quality
    QualityStatus.OK: "C6EFCE",                  # Light green — acceptable, no action

    # Actionable (blues/teals)
    QualityStatus.IMPROVEMENT_READY: "BDE0FE",   # Light blue — improvement available
    QualityStatus.WEAK: "FFF2CC",                # Light yellow — present but thin
    QualityStatus.SOURCE_CONFLICT: "FFD6A5",     # Light orange — sources disagree

    # Issues (yellows/reds)
    QualityStatus.SHOULD_IMPROVE: "FFEB9C",      # Yellow — quality issues
    QualityStatus.PROBABLE_ERROR: "FF6B6B",      # Dark red — likely wrong
    QualityStatus.MISSING: "FFC7CE",             # Red — field absent

    # Blocked (purples/blues)
    QualityStatus.NO_RELIABLE_SOURCE: "E8DAEF",  # Lavender — can't evaluate
    QualityStatus.MANUAL_REVIEW: "E2BFFF",       # Purple — human must decide
    QualityStatus.REQUIRES_MANUFACTURER: "B4C7E7",  # Blue — needs manufacturer
}

STATUS_FONT_COLORS = {
    QualityStatus.STRONG: "006100",
    QualityStatus.OK: "006100",
    QualityStatus.IMPROVEMENT_READY: "1D4ED8",   # Blue text
    QualityStatus.WEAK: "9C6500",
    QualityStatus.SOURCE_CONFLICT: "9A3412",     # Dark orange text
    QualityStatus.SHOULD_IMPROVE: "9C6500",
    QualityStatus.PROBABLE_ERROR: "FFFFFF",
    QualityStatus.MISSING: "9C0006",
    QualityStatus.NO_RELIABLE_SOURCE: "6B21A8",  # Dark purple text
    QualityStatus.MANUAL_REVIEW: "4B0082",
    QualityStatus.REQUIRES_MANUFACTURER: "003380",
}


def read_article_numbers(file_content: bytes, filename: str) -> tuple[list[str], str]:
    """Read article numbers from an uploaded Excel file.

    Returns a tuple of (article_numbers, detected_column_name).
    """
    wb = load_workbook(BytesIO(file_content), read_only=True, data_only=True)
    ws = wb.active

    article_numbers = []
    article_col = 0
    detected_column = "Kolonne A (standard)"

    # Check header row for article number column
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if header_row:
        search_terms = {
            "artikkelnummer", "artikkel", "artikkelnr", "artnr", "artnummer",
            "varenummer", "varenr",
            "article", "articlenumber", "article_number",
            "sku", "item", "itemnumber", "produktnummer",
        }
        normalized_search = {
            t.replace(".", "").replace(" ", "").replace("_", "")
            for t in search_terms
        }
        for idx, cell_value in enumerate(header_row):
            if cell_value:
                normalized = str(cell_value).strip().lower().replace(".", "").replace(" ", "").replace("_", "")
                if normalized in normalized_search:
                    article_col = idx
                    detected_column = str(cell_value).strip()
                    logger.info(f"Found article number column: '{cell_value}' at index {idx}")
                    break

    # Read article numbers — normalize to prevent float coercion (e.g., 12345.0 → "12345")
    # This safely handles re-imported Excel files where numeric article numbers were
    # previously exported without text formatting. Old "12345.0" values are corrected
    # to "12345" on the next run via normalize_identifier().
    start_row = 2 if header_row else 1
    float_coercion_count = 0
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        if row and len(row) > article_col:
            value = row[article_col]
            normalized = normalize_identifier(value)
            if normalized:
                # Track float-to-string corrections for diagnostics
                if isinstance(value, float) and normalized != str(value):
                    float_coercion_count += 1
                article_numbers.append(normalized)

    wb.close()
    if float_coercion_count:
        logger.warning(
            f"Corrected {float_coercion_count} float-formatted article numbers "
            f"in {filename} (e.g., 12345.0 → '12345')"
        )
    logger.info(f"Read {len(article_numbers)} article numbers from {filename} (column: {detected_column})")
    return article_numbers, detected_column


def create_output_excel(
    results: list[ProductAnalysis],
    output_path: str,
    analysis_mode: str = "full_enrichment",
    focus_areas: list[str] | None = None,
) -> None:
    """Create a structured Excel output file with analysis results.

    Output structure adapts to analysis_mode:
    - full_enrichment: all sheets (existing behavior)
    - audit_only: summary + audit scoring + overview (no enrichment sheets)
    - focused_scan: summary + focused area scores (minimal, relevant sheets only)
    """
    from backend.scoring import AREA_LABELS
    wb = Workbook()

    is_audit = analysis_mode == "audit_only"
    is_focused = analysis_mode == "focused_scan"

    # Sheet 1: Summary (always included)
    ws_summary = wb.active
    ws_summary.title = "Summary"
    _create_summary_sheet(ws_summary, results, analysis_mode=analysis_mode, focus_areas=focus_areas)

    # Sheet 2: Area Scores (audit and focused modes)
    if is_audit or is_focused:
        ws_audit = wb.create_sheet("Omr\u00e5descorer")
        _create_area_scores_sheet(ws_audit, results, focus_areas)

    if not is_focused:
        # Full overview for audit and enrichment modes
        ws_overview = wb.create_sheet("Oversikt")
        _create_overview_sheet(ws_overview, results)

        ws_detail = wb.create_sheet("Feltanalyse")
        _create_detail_sheet(ws_detail, results)

    if not is_audit and not is_focused:
        # P2 FIX: Simplified sheet structure for full_enrichment mode.
        # Removed: Comparison_And_Enrichment (merged into Feltanalyse with traceability)
        # Removed: Debug_Log (merged into Feltanalyse with traceability)
        # Removed: Kildekonflikter (conflicts shown in Forbedringsforslag)

        ws_improvements = wb.create_sheet("Forbedringsforslag")
        _create_improvements_sheet(ws_improvements, results)

        ws_quick = wb.create_sheet("Quick Wins")
        _create_quick_wins_sheet(ws_quick, results)

        ws_inriver = wb.create_sheet("Inriver Import")
        _create_inriver_import_sheet(ws_inriver, results)

        # Produsentoppfølging: only include if there are actual manufacturer-contact items
        mfr_contact_results = [
            r for r in results if r.requires_manufacturer_contact
        ]
        if mfr_contact_results:
            ws_manufacturer = wb.create_sheet("Produsentoppfølging")
            _create_manufacturer_sheet(ws_manufacturer, results)

    # Image sheet: single combined sheet (merged Bildeanalyse + Bildeproblemer)
    include_images = (not is_focused) or (focus_areas and "images" in focus_areas)
    if include_images:
        ws_images = wb.create_sheet("Bildeanalyse")
        _create_image_detail_sheet(ws_images, results)

    # Save directly to file
    wb.save(output_path)
    logger.info(f"Excel output saved to {output_path} (mode={analysis_mode})")


def _write_id_cell(ws, row: int, column: int, value, alignment=None):
    """Write an identifier value to a cell, formatted as text to prevent Excel auto-conversion.

    This ensures article numbers, GIDs, SKUs, and other identifiers are never
    displayed as scientific notation or truncated floats in Excel.
    """
    cell = ws.cell(row=row, column=column, value=normalize_identifier_strict(value))
    cell.number_format = "@"  # Excel text format — prevents auto-conversion to number
    if alignment:
        cell.alignment = alignment
    return cell


def _style_header(ws, row: int, num_cols: int) -> None:
    """Apply header styling to a row."""
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _apply_status_style(cell, status: QualityStatus) -> None:
    """Apply color styling based on quality status."""
    fill_color = STATUS_COLORS.get(status, "FFFFFF")
    font_color = STATUS_FONT_COLORS.get(status, "000000")
    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    cell.font = Font(color=font_color)


IMAGE_STATUS_COLORS = {
    "PASS": "C6EFCE",
    "PASS_WITH_NOTES": "FFEB9C",
    "REVIEW": "FFEB9C",
    "FAIL": "FF6B6B",
    "MISSING": "FFC7CE",
}

IMAGE_STATUS_FONT_COLORS = {
    "PASS": "006100",
    "PASS_WITH_NOTES": "9C6500",
    "REVIEW": "9C6500",
    "FAIL": "FFFFFF",
    "MISSING": "9C0006",
}


def _apply_quality_score_color(cell, score: int) -> None:
    """Apply color coding to a quality dimension score (0-100)."""
    if score >= 75:
        cell.font = Font(color="006100")  # Green
    elif score >= 50:
        cell.font = Font(color="9C6500")  # Orange
    else:
        cell.font = Font(color="9C0006", bold=True)  # Red


def _apply_image_status_style(cell, status: str) -> None:
    """Apply color styling based on image quality status."""
    fill_color = IMAGE_STATUS_COLORS.get(status, "FFFFFF")
    font_color = IMAGE_STATUS_FONT_COLORS.get(status, "000000")
    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    cell.font = Font(color=font_color)


def _create_overview_sheet(ws, results: list[ProductAnalysis]) -> None:
    """Create the overview sheet with one row per product."""
    headers = [
        "Artikkelnummer",
        "Produktnavn",
        "Produsent",
        "Produsentens varenummer",
        "Funnet p\u00e5 OneMed",
        "Verifiseringsstatus",
        "Verifiseringsbevis",
        "Total score (%)",
        "Innholdskvalitet (snitt)",
        "Samsvarskvalitet (snitt)",
        "Status",
        "Kommentar",
        "Kategori",
        "Bildescore",
        "Bildestatus",
        "Antall bilder",
        "PDF tilgjengelig",
        "Berikelser funnet",
        "Kildekonflikter",
        "Auto-fix",
        "Manuell vurdering",
        "Krever produsentkontakt",
        "Produkt-URL",
        "PDF-URL",
    ]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    _style_header(ws, 1, len(headers))

    for row_idx, result in enumerate(results, 2):
        pd = result.product_data
        iq = result.image_quality or {}
        enriched = [e for e in result.enrichment_results if e.match_status != "NOT_FOUND"]
        conflicts = [e for e in result.enrichment_results if e.match_status == "FOUND_IN_BOTH_CONFLICT"]

        # Get best producer info from all available sources
        producer, producer_artnr = get_best_producer_info(
            pd, result.jeeves_data, result.manufacturer_lookup
        )

        _write_id_cell(ws, row_idx, 1, result.article_number)
        ws.cell(row=row_idx, column=2, value=pd.product_name or "")
        ws.cell(row=row_idx, column=3, value=producer or "")
        ws.cell(row=row_idx, column=4, value=producer_artnr or "")
        ws.cell(row=row_idx, column=5, value="Ja" if pd.found_on_onemed else "Nei")
        # Verification status and evidence — business-friendly labels
        from backend.models import VerificationStatus as VS
        v_label = VS.business_label(pd.verification_status)
        v_cell = ws.cell(row=row_idx, column=6, value=v_label)
        # Color-code verification: green=exact, yellow=weak, red=mismatch
        if pd.verification_status in (VS.EXACT_MATCH, VS.NORMALIZED_MATCH):
            v_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif pd.verification_status in (VS.CDN_ONLY, VS.UNVERIFIED, VS.SKU_IN_PAGE):
            v_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        elif pd.verification_status in (VS.MISMATCH, VS.AMBIGUOUS):
            v_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        # Business-friendly evidence text
        v_evidence = VS.business_evidence(pd.verification_status, pd.verification_evidence)
        ws.cell(row=row_idx, column=7, value=v_evidence)
        ws.cell(row=row_idx, column=8, value=result.total_score)
        # Average content quality and conformity quality across fields
        cq_scores = [fa.content_quality for fa in result.field_analyses if fa.content_quality is not None]
        conf_scores = [fa.conformity_quality for fa in result.field_analyses if fa.conformity_quality is not None]
        avg_cq = round(sum(cq_scores) / len(cq_scores)) if cq_scores else ""
        avg_conf = round(sum(conf_scores) / len(conf_scores)) if conf_scores else ""
        cq_cell = ws.cell(row=row_idx, column=9, value=avg_cq)
        if isinstance(avg_cq, (int, float)):
            _apply_quality_score_color(cq_cell, avg_cq)
        conf_cell = ws.cell(row=row_idx, column=10, value=avg_conf)
        if isinstance(avg_conf, (int, float)):
            _apply_quality_score_color(conf_cell, avg_conf)
        status_cell = ws.cell(row=row_idx, column=11, value=result.overall_status.value)
        _apply_status_style(status_cell, result.overall_status)
        ws.cell(row=row_idx, column=12, value=result.overall_comment or "")
        # Show full breadcrumb path when available, fall back to leaf category
        cat_display = ""
        if pd.category_breadcrumb:
            cat_display = " > ".join(pd.category_breadcrumb)
        elif pd.category:
            cat_display = pd.category
        ws.cell(row=row_idx, column=13, value=cat_display)
        # Image quality columns
        img_score = iq.get("avg_image_score", 0)
        img_status = iq.get("image_quality_status", "MISSING")
        img_count = iq.get("image_count_found", 0)
        ws.cell(row=row_idx, column=14, value=round(img_score, 1) if img_score else 0)
        img_status_cell = ws.cell(row=row_idx, column=15, value=img_status)
        _apply_image_status_style(img_status_cell, img_status)
        ws.cell(row=row_idx, column=16, value=img_count)
        # Enrichment columns
        ws.cell(row=row_idx, column=17, value="Ja" if result.pdf_available else "Nei")
        ws.cell(row=row_idx, column=18, value=len(enriched))
        conflict_cell = ws.cell(row=row_idx, column=19, value=len(conflicts))
        if conflicts:
            conflict_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            conflict_cell.font = Font(color="9C6500")
        ws.cell(row=row_idx, column=20, value="Ja" if result.auto_fix_possible else "Nei")
        ws.cell(row=row_idx, column=21, value="Ja" if result.manual_review_needed else "Nei")
        ws.cell(row=row_idx, column=22, value="Ja" if result.requires_manufacturer_contact else "Nei")
        ws.cell(row=row_idx, column=23, value=pd.product_url or "")
        ws.cell(row=row_idx, column=24, value=result.pdf_url or "")

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = max(15, len(headers[col - 1]) + 5)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(results) + 1}"


def _create_detail_sheet(ws, results: list[ProductAnalysis]) -> None:
    """Create detailed field analysis sheet with traceability."""
    headers = [
        "Artikkelnummer",
        "Produktnavn",
        "Produsent",
        "Produsentens varenummer",
        "Felt",
        "Nåværende verdi",
        "Status",
        "Confidence",
        "Innholdskvalitet",
        "Samsvarskvalitet",
        "Kvalitetsvurdering",
        "Confidence-detaljer",
        "Statusårsak",
        "Verdikilde",
        "Nettside-verdi",
        "Jeeves-verdi",
        "Foreslått verdi",
        "Forslag-kilde",
        "Endringstype",
        "Diff",
    ]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    _style_header(ws, 1, len(headers))

    row_idx = 2
    for result in results:
        producer, producer_artnr = get_best_producer_info(
            result.product_data, result.jeeves_data, result.manufacturer_lookup
        )
        for fa in result.field_analyses:
            _write_id_cell(ws, row_idx, 1, result.article_number)
            ws.cell(row=row_idx, column=2, value=result.product_data.product_name or "")
            ws.cell(row=row_idx, column=3, value=producer or "")
            ws.cell(row=row_idx, column=4, value=producer_artnr or "")
            ws.cell(row=row_idx, column=5, value=fa.field_name)
            ws.cell(row=row_idx, column=6, value=fa.current_value or "")
            status_cell = ws.cell(row=row_idx, column=7, value=fa.status.value)
            _apply_status_style(status_cell, fa.status)
            # Confidence score with color coding
            conf_val = fa.confidence if fa.confidence is not None else ""
            conf_cell = ws.cell(row=row_idx, column=8, value=conf_val)
            if isinstance(conf_val, (int, float)):
                if conf_val >= 75:
                    conf_cell.font = Font(color="006100")
                elif conf_val >= 50:
                    conf_cell.font = Font(color="9C6500")
                else:
                    conf_cell.font = Font(color="9C0006", bold=True)
            # Two-dimensional quality scores
            cq_val = fa.content_quality if fa.content_quality is not None else ""
            cq_cell = ws.cell(row=row_idx, column=9, value=cq_val)
            if isinstance(cq_val, (int, float)):
                _apply_quality_score_color(cq_cell, cq_val)
            conf_q_val = fa.conformity_quality if fa.conformity_quality is not None else ""
            conf_q_cell = ws.cell(row=row_idx, column=10, value=conf_q_val)
            if isinstance(conf_q_val, (int, float)):
                _apply_quality_score_color(conf_q_cell, conf_q_val)
            ws.cell(row=row_idx, column=11, value=fa.quality_label or "")
            ws.cell(row=row_idx, column=12, value=fa.confidence_details or "")
            ws.cell(row=row_idx, column=13, value=fa.status_reason or fa.comment or "")
            ws.cell(row=row_idx, column=14, value=fa.value_origin or fa.source or "")
            ws.cell(row=row_idx, column=15, value=fa.website_value or "")
            ws.cell(row=row_idx, column=16, value=fa.jeeves_value or "")
            ws.cell(row=row_idx, column=17, value=fa.suggested_value or "")
            ws.cell(row=row_idx, column=18, value=fa.suggestion_source or fa.source or "")
            # Diff columns — only populated when a suggestion exists
            if fa.suggested_value:
                ws.cell(row=row_idx, column=19, value=summarize_change_type(fa.current_value, fa.suggested_value))
                ws.cell(row=row_idx, column=20, value=build_field_diff(fa.current_value, fa.suggested_value))
            row_idx += 1

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = max(15, len(headers[col - 1]) + 5)

    ws.freeze_panes = "A2"
    if row_idx > 2:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row_idx - 1}"


def _create_improvements_sheet(ws, results: list[ProductAnalysis]) -> None:
    """Create improvements suggestion sheet.

    Combines field analysis suggestions with enrichment engine suggestions.
    All suggestions are validated before writing — contact info, other SKUs,
    PDF noise, and wrong-language content is rejected or flagged.
    """
    # Structured evidence columns replace the old "Evidens / Begrunnelse" free-text
    headers = [
        "Artikkelnummer",
        "Produktnavn",
        "Produsent",
        "Produsentens varenummer",
        "Felt",
        "Nåværende verdi",
        "Foreslått verdi",
        "Kilde brukt",
        "Kilde-URL",
        "Kildespråk",
        "Oversatt",
        "Variant bekreftet",
        "Støy filtrert",
        "Feltklassifisering",
        "Confidence",
        "Inriver-klar",
        "Begrunnelse",
        "Krever gjennomgang",
        "Kildeverdi (før AI)",
        "Endret av AI",
        "Diff",
        "Endringstype",
        "Endringsomfang",
    ]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    _style_header(ws, 1, len(headers))

    row_idx = 2
    for result in results:
        enriched_fields = set()
        sku = result.article_number

        # Get best producer info
        producer, producer_artnr = get_best_producer_info(
            result.product_data, result.jeeves_data, result.manufacturer_lookup
        )

        # Enrichment suggestions — validated before writing
        for es in result.enrichment_suggestions:
            if not es.suggested_value:
                continue
            # Skip if suggestion equals current value (redundant)
            if es.current_value and es.suggested_value:
                cv = es.current_value.strip().lower()
                sv = es.suggested_value.strip().lower()
                if cv == sv:
                    continue

            # Validate suggestion output — reject contact info, other SKUs, PDF noise
            is_valid, reject_reason = validate_suggestion_output(
                es.suggested_value, es.field_name, sku
            )
            if not is_valid:
                logger.info(
                    f"[excel] Suggestion rejected for {sku}/{es.field_name}: {reject_reason}"
                )
                continue

            # Safety net: translate any remaining sv/da content to Norwegian
            output_value = es.suggested_value
            translated, _lang, was_translated = translate_to_norwegian_if_needed(output_value)
            if was_translated:
                output_value = translated

            # Extract structured evidence tags (or fall back to legacy evidence)
            ev = es.evidence_structured or {}

            _write_id_cell(ws, row_idx, 1, sku)
            ws.cell(row=row_idx, column=2, value=result.product_data.product_name or "")
            ws.cell(row=row_idx, column=3, value=producer or "")
            ws.cell(row=row_idx, column=4, value=producer_artnr or "")
            ws.cell(row=row_idx, column=5, value=es.field_name)
            ws.cell(row=row_idx, column=6, value=es.current_value or "")
            ws.cell(row=row_idx, column=7, value=output_value)
            # Structured evidence columns
            ws.cell(row=row_idx, column=8, value=ev.get("Kilde", es.source or ""))
            ws.cell(row=row_idx, column=9, value=ev.get("Kilde-URL", es.source_url or ""))
            ws.cell(row=row_idx, column=10, value=ev.get("Kildespråk", ""))
            ws.cell(row=row_idx, column=11, value=ev.get("Oversatt til norsk", "nei"))
            ws.cell(row=row_idx, column=12, value=ev.get("Variant sikkert identifisert", ""))
            ws.cell(row=row_idx, column=13, value=ev.get("Støy filtrert bort", "nei"))
            ws.cell(row=row_idx, column=14, value=ev.get("Felttypeklassifisering", ""))
            ws.cell(row=row_idx, column=15, value=es.confidence if es.confidence else "")
            ws.cell(row=row_idx, column=16, value=ev.get("Auto-egnet for Inriver", ""))
            ws.cell(row=row_idx, column=17, value=es.evidence or "")  # Full text begrunnelse
            ws.cell(row=row_idx, column=18, value="Ja" if es.review_required else "Nei")
            ws.cell(row=row_idx, column=19, value=es.original_suggested_value or "")
            ws.cell(row=row_idx, column=20, value="Ja" if es.ai_modified else "Nei")
            # Diff columns
            ws.cell(row=row_idx, column=21, value=build_field_diff(es.current_value, output_value))
            ws.cell(row=row_idx, column=22, value=summarize_change_type(es.current_value, output_value))
            ws.cell(row=row_idx, column=23, value=detect_change_scope(es.current_value, output_value))
            enriched_fields.add(es.field_name)
            row_idx += 1

        # Field analysis suggestions (for fields not covered by enricher)
        for fa in result.field_analyses:
            if fa.suggested_value and fa.field_name not in enriched_fields:
                # Validate before writing
                is_valid, reject_reason = validate_suggestion_output(
                    fa.suggested_value, fa.field_name, sku
                )
                if not is_valid:
                    logger.info(
                        f"[excel] Field suggestion rejected for {sku}/{fa.field_name}: {reject_reason}"
                    )
                    continue

                # Safety net: translate remaining sv/da
                fa_output = fa.suggested_value
                fa_translated, _fa_lang, fa_was_translated = translate_to_norwegian_if_needed(fa_output)
                if fa_was_translated:
                    fa_output = fa_translated

                _write_id_cell(ws, row_idx, 1, sku)
                ws.cell(row=row_idx, column=2, value=result.product_data.product_name or "")
                ws.cell(row=row_idx, column=3, value=producer or "")
                ws.cell(row=row_idx, column=4, value=producer_artnr or "")
                ws.cell(row=row_idx, column=5, value=fa.field_name)
                ws.cell(row=row_idx, column=6, value=fa.current_value or "")
                ws.cell(row=row_idx, column=7, value=fa_output)
                ws.cell(row=row_idx, column=8, value=fa.source or "")  # Kilde brukt
                ws.cell(row=row_idx, column=9, value="")   # Kilde-URL
                ws.cell(row=row_idx, column=10, value="")  # Kildespråk
                ws.cell(row=row_idx, column=11, value="")  # Oversatt
                ws.cell(row=row_idx, column=12, value="")  # Variant bekreftet
                ws.cell(row=row_idx, column=13, value="")  # Støy filtrert
                ws.cell(row=row_idx, column=14, value="")  # Feltklassifisering
                ws.cell(row=row_idx, column=15, value=fa.confidence if fa.confidence else "")
                ws.cell(row=row_idx, column=16, value="")  # Inriver-klar
                ws.cell(row=row_idx, column=17, value=fa.comment or "")
                ws.cell(row=row_idx, column=18, value="Ja")
                ws.cell(row=row_idx, column=19, value="")
                ws.cell(row=row_idx, column=20, value="Nei")
                # Diff columns
                ws.cell(row=row_idx, column=21, value=build_field_diff(fa.current_value, fa_output))
                ws.cell(row=row_idx, column=22, value=summarize_change_type(fa.current_value, fa_output))
                ws.cell(row=row_idx, column=23, value=detect_change_scope(fa.current_value, fa_output))
                row_idx += 1

    if row_idx == 2:
        ws.cell(row=2, column=1, value="Ingen forbedringsforslag generert")

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = max(15, len(headers[col - 1]) + 5)

    ws.freeze_panes = "A2"


def _create_quick_wins_sheet(ws, results: list[ProductAnalysis]) -> None:
    """Create Quick Wins sheet — low-risk, high-value improvements only.

    Shows ONLY suggestions that pass all quick win criteria:
    high confidence, trusted source, not medically sensitive,
    no source conflicts, not AI-generated.
    """
    from backend.quick_wins import is_quick_win

    headers = [
        "Artikkelnummer",
        "Produktnavn",
        "Produsent",
        "Felt",
        "Nåværende verdi",
        "Foreslått verdi",
        "Kilde",
        "Confidence",
        "Begrunnelse",
    ]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    _style_header(ws, 1, len(headers))

    # Green header band for quick wins
    qw_fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")

    row_idx = 2
    fa_map_cache = {}

    for result in results:
        fa_map = {fa.field_name: fa for fa in result.field_analyses}
        producer, _ = get_best_producer_info(
            result.product_data, result.jeeves_data, result.manufacturer_lookup
        )

        for es in (result.enrichment_suggestions or []):
            if not es.suggested_value:
                continue
            fa = fa_map.get(es.field_name)
            if not is_quick_win(es, fa):
                continue

            _write_id_cell(ws, row_idx, 1, result.article_number)
            ws.cell(row=row_idx, column=2, value=result.product_data.product_name or "")
            ws.cell(row=row_idx, column=3, value=producer or "")
            ws.cell(row=row_idx, column=4, value=es.field_name)
            ws.cell(row=row_idx, column=5, value=es.current_value or "")
            val_cell = ws.cell(row=row_idx, column=6, value=es.suggested_value)
            val_cell.fill = qw_fill
            ws.cell(row=row_idx, column=7, value=es.source or "")
            ws.cell(row=row_idx, column=8, value=es.confidence if es.confidence else "")
            ws.cell(row=row_idx, column=9, value=es.evidence or "")
            row_idx += 1

    if row_idx == 2:
        ws.cell(row=2, column=1, value="Ingen quick wins funnet")
        ws.cell(row=2, column=2, value="Alle forslag krever manuell vurdering eller har for lav confidence")

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = max(15, len(headers[col - 1]) + 5)

    ws.freeze_panes = "A2"
    if row_idx > 2:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row_idx - 1}"


def _classify_missing_gap(fa_name: str, fa_status, iq: dict = None) -> str:
    """Classify a missing/weak field into a follow-up category."""
    if fa_name == "Bildekvalitet":
        return "Mangler bilde"
    if fa_name == "Spesifikasjon":
        return "Mangler teknisk info"
    if fa_name == "Beskrivelse":
        return "Mangler produktbeskrivelse"
    if fa_name == "Produsentens varenummer":
        return "Mangler produsent art.nr"
    if fa_name == "Produsent":
        return "Mangler produsentinfo"
    if fa_name == "Kategori":
        return "Mangler kategorisering"
    if fa_name == "Pakningsinformasjon":
        return "Mangler pakningsinfo"
    return f"Mangler: {fa_name}"


def _sources_checked_for_product(result: "ProductAnalysis") -> list[str]:
    """List which sources were already checked for a product."""
    checked = []
    pd = result.product_data
    if pd.found_on_onemed:
        checked.append("onemed.no nettside")
    if result.pdf_available:
        checked.append(f"PDF datablad ({result.pdf_url or 'tilgjengelig'})")
    mfr = result.manufacturer_lookup
    if mfr and mfr.searched:
        if mfr.found:
            checked.append(f"Produsentens nettside ({mfr.source_url or 'funnet'})")
        else:
            checked.append("Produsentens nettside (ikke funnet)")
    if result.norengros_lookup and getattr(result.norengros_lookup, 'searched', False):
        checked.append("Norengros (sekundærkilde)")
    if result.jeeves_data:
        checked.append("Jeeves produktkatalog")
    if not checked:
        checked.append("Ingen kilder sjekket")
    return checked


def _build_email_draft(
    manufacturer: str,
    products: list["ProductAnalysis"],
) -> str:
    """Build a ready-to-send email draft for a manufacturer."""
    lines = [
        f"Emne: Forespørsel om produktinformasjon — {len(products)} produkt(er)",
        "",
        f"Hei {manufacturer},",
        "",
        "Vi holder på med en kvalitetsgjennomgang av produktdataene i nettbutikken vår",
        "og mangler informasjon for følgende produkter fra dere:",
        "",
    ]

    for result in products:
        producer, producer_artnr = get_best_producer_info(
            result.product_data, result.jeeves_data, result.manufacturer_lookup
        )
        name = result.product_data.product_name or result.article_number
        artnr_display = f" (deres art.nr: {producer_artnr})" if producer_artnr else ""

        # Categorize gaps
        gaps = {}
        for fa in result.field_analyses:
            if fa.status in (QualityStatus.MISSING, QualityStatus.REQUIRES_MANUFACTURER,
                             QualityStatus.WEAK, QualityStatus.NO_RELIABLE_SOURCE):
                category = _classify_missing_gap(fa.field_name, fa.status)
                gaps[category] = gaps.get(category, [])
                gaps[category].append(fa.field_name)

        # Check image separately
        iq = result.image_quality or {}
        img_status = iq.get("image_quality_status", "")
        if img_status in ("MISSING", "FAIL"):
            gaps["Mangler bilde"] = gaps.get("Mangler bilde", ["Bildekvalitet"])

        if not gaps:
            continue

        lines.append(f"  Produkt: {name}{artnr_display}")
        lines.append(f"  Vårt art.nr: {result.article_number}")
        for gap_category in sorted(gaps.keys()):
            lines.append(f"    → {gap_category}")
        lines.append("")

    lines.extend([
        "Vi ville satt stor pris på om dere kunne sende oss:",
        "",
    ])

    # Aggregate what's needed across all products
    all_gap_categories = set()
    for result in products:
        for fa in result.field_analyses:
            if fa.status in (QualityStatus.MISSING, QualityStatus.REQUIRES_MANUFACTURER,
                             QualityStatus.WEAK, QualityStatus.NO_RELIABLE_SOURCE):
                all_gap_categories.add(_classify_missing_gap(fa.field_name, fa.status))
        iq = result.image_quality or {}
        if iq.get("image_quality_status", "") in ("MISSING", "FAIL"):
            all_gap_categories.add("Mangler bilde")

    request_map = {
        "Mangler teknisk info": "Teknisk datablad / spesifikasjoner (materiale, størrelse, sterilitet, etc.)",
        "Mangler bilde": "Produktbilder i høy oppløsning (min. 800x800px, hvit bakgrunn)",
        "Mangler produsent art.nr": "Deres artikkelnummer for produktet",
        "Mangler produktbeskrivelse": "Produktbeskrivelse egnet for nettbutikk (norsk)",
        "Mangler produsentinfo": "Bekreftelse av produsentnavn",
        "Mangler kategorisering": "Produktkategori / bruksområde",
        "Mangler pakningsinfo": "Pakningsinformasjon (antall per eske/kartong)",
    }

    for gap_cat in sorted(all_gap_categories):
        request_text = request_map.get(gap_cat, gap_cat)
        lines.append(f"  • {request_text}")

    lines.extend([
        "",
        "Format: PDF-datablad, produktbilder som JPG/PNG, eller annet egnet format.",
        "",
        "Med vennlig hilsen,",
        "[Ditt navn]",
        "OneMed Norge",
    ])

    return "\n".join(lines)


def _create_manufacturer_sheet(ws, results: list[ProductAnalysis]) -> None:
    """Create manufacturer follow-up sheet — a practical 'send to producer' package.

    Two sections:
    1. Detail table: one row per product with categorized gaps, sources checked,
       and what the system attempted to find
    2. Email drafts: one per manufacturer, ready to copy-paste
    """
    # ── Section 1: Detail table ──
    headers = [
        "Produsent",
        "Produsentens varenummer",
        "Artikkelnummer",
        "Produktnavn",
        "Mangel-kategori",
        "Manglende felt",
        "Felt med svak verdi",
        "Kilder allerede sjekket",
        "Hva systemet forsøkte",
        "PDF tilgjengelig",
        "Bilde-status",
        "Confidence",
        "Status",
    ]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    _style_header(ws, 1, len(headers))

    # Group by manufacturer
    by_manufacturer: dict[str, list[ProductAnalysis]] = {}
    for result in results:
        if result.requires_manufacturer_contact:
            producer, _ = get_best_producer_info(
                result.product_data, result.jeeves_data, result.manufacturer_lookup
            )
            mfr = producer or "Ukjent produsent"
            if mfr not in by_manufacturer:
                by_manufacturer[mfr] = []
            by_manufacturer[mfr].append(result)

    wrap = Alignment(wrap_text=True, vertical="top")
    group_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    group_font = Font(bold=True, size=11)

    row_idx = 2
    for manufacturer in sorted(by_manufacturer.keys()):
        products = by_manufacturer[manufacturer]

        # Group header row
        cell = ws.cell(row=row_idx, column=1, value=f"{manufacturer} — {len(products)} produkt(er)")
        cell.fill = group_fill
        cell.font = group_font
        for col in range(2, len(headers) + 1):
            ws.cell(row=row_idx, column=col).fill = group_fill
        row_idx += 1

        for result in products:
            _, producer_artnr = get_best_producer_info(
                result.product_data, result.jeeves_data, result.manufacturer_lookup
            )
            pd = result.product_data
            iq = result.image_quality or {}

            # Categorize field gaps
            missing_fields = []
            weak_fields = []
            gap_categories = set()
            for fa in result.field_analyses:
                if fa.status in (QualityStatus.MISSING, QualityStatus.REQUIRES_MANUFACTURER,
                                 QualityStatus.NO_RELIABLE_SOURCE):
                    missing_fields.append(fa.field_name)
                    gap_categories.add(_classify_missing_gap(fa.field_name, fa.status))
                elif fa.status in (QualityStatus.WEAK, QualityStatus.SHOULD_IMPROVE):
                    weak_fields.append(fa.field_name)

            # Image gap
            img_status = iq.get("image_quality_status", "MISSING")
            if img_status in ("MISSING", "FAIL"):
                gap_categories.add("Mangler bilde")

            # Sources checked
            sources = _sources_checked_for_product(result)

            # What the system attempted
            attempts = []
            if pd.found_on_onemed:
                attempts.append("Scrapet produktside på onemed.no")
            if result.pdf_available:
                attempts.append("Ekstrahert data fra PDF-datablad")
            mfr_lookup = result.manufacturer_lookup
            if mfr_lookup and mfr_lookup.searched:
                attempts.append("Søkt etter produsent-nettside" + (" (funnet)" if mfr_lookup.found else " (ikke funnet)"))
            if result.enrichment_suggestions:
                attempts.append(f"Generert {len(result.enrichment_suggestions)} forbedringsforslag")
            if not attempts:
                attempts.append("Ingen automatisk berikelse utført")

            # Average confidence for this product
            confs = [fa.confidence for fa in result.field_analyses if fa.confidence is not None and fa.confidence > 0]
            avg_conf = round(sum(confs) / len(confs)) if confs else 0

            # Write row
            ws.cell(row=row_idx, column=1, value=manufacturer)
            ws.cell(row=row_idx, column=2, value=producer_artnr or "")
            _write_id_cell(ws, row_idx, 3, result.article_number)
            ws.cell(row=row_idx, column=4, value=pd.product_name or "")
            ws.cell(row=row_idx, column=5, value="\n".join(sorted(gap_categories))).alignment = wrap
            ws.cell(row=row_idx, column=6, value=", ".join(missing_fields) if missing_fields else "Ingen")
            ws.cell(row=row_idx, column=7, value=", ".join(weak_fields) if weak_fields else "Ingen")
            ws.cell(row=row_idx, column=8, value="\n".join(sources)).alignment = wrap
            ws.cell(row=row_idx, column=9, value="\n".join(attempts)).alignment = wrap
            ws.cell(row=row_idx, column=10, value="Ja" if result.pdf_available else "Nei")
            img_cell = ws.cell(row=row_idx, column=11, value=img_status)
            if img_status in ("MISSING", "FAIL"):
                img_cell.font = Font(color="9C0006", bold=True)
            ws.cell(row=row_idx, column=12, value=avg_conf)
            status_cell = ws.cell(row=row_idx, column=13, value=result.overall_status.value)
            _apply_status_style(status_cell, result.overall_status)
            row_idx += 1

    if row_idx == 2:
        ws.cell(row=2, column=1, value="Ingen produkter krever produsentkontakt")

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = max(18, len(headers[col - 1]) + 5)
    ws.column_dimensions["E"].width = 30  # Mangel-kategori
    ws.column_dimensions["H"].width = 35  # Kilder sjekket
    ws.column_dimensions["I"].width = 40  # Hva systemet forsøkte

    ws.freeze_panes = "A2"
    if row_idx > 2:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row_idx - 1}"

    # ── Section 2: Email drafts per manufacturer ──
    if not by_manufacturer:
        return

    email_start = row_idx + 3
    email_header_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    email_font = Font(bold=True, size=12)

    ws.cell(row=email_start - 1, column=1, value="E-postutkast per produsent").font = Font(bold=True, size=13, color="1F4E79")

    for manufacturer in sorted(by_manufacturer.keys()):
        products = by_manufacturer[manufacturer]

        # Manufacturer header
        header_cell = ws.cell(row=email_start, column=1, value=f"Til: {manufacturer}")
        header_cell.fill = email_header_fill
        header_cell.font = email_font
        for col in range(2, 6):
            ws.cell(row=email_start, column=col).fill = email_header_fill
        email_start += 1

        # Build email
        email_text = _build_email_draft(manufacturer, products)

        # Write email text
        email_cell = ws.cell(row=email_start, column=1, value=email_text)
        email_cell.alignment = Alignment(wrap_text=True, vertical="top")
        # Merge across columns for readability
        ws.merge_cells(
            start_row=email_start, start_column=1,
            end_row=email_start, end_column=len(headers),
        )
        # Set row height based on line count
        line_count = email_text.count("\n") + 1
        ws.row_dimensions[email_start].height = max(15, min(400, line_count * 14))

        email_start += 2  # Gap between emails


def _create_image_detail_sheet(ws, results: list[ProductAnalysis]) -> None:
    """Create detailed image analysis sheet with one row per image.

    Includes producer information and integrated improvement suggestions
    for products with image problems.
    """
    headers = [
        "Artikkelnummer",
        "Produktnavn",
        "Produsent",
        "Produsentens varenummer",
        "Bilde",
        "URL",
        "Finnes",
        "Filstr. (KB)",
        "Bredde",
        "H\u00f8yde",
        "Ratio",
        "Oppl\u00f8sning",
        "Skarphet (r\u00e5)",
        "Skarphet",
        "Lysstyrke",
        "Lysstyrke score",
        "Kontrast",
        "Kontrast score",
        "Hvit bakgr.",
        "Bakgrunn score",
        "Kantdeteksjon",
        "Kant score",
        "Produktfyll",
        "Fyll score",
        "Total score",
        "Status",
        "Problemer",
    ]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    _style_header(ws, 1, len(headers))

    row_idx = 2
    for result in results:
        iq = result.image_quality
        if not iq:
            continue

        producer, producer_artnr = get_best_producer_info(
            result.product_data, result.jeeves_data, result.manufacturer_lookup
        )

        for img in iq.get("image_analyses", []):
            _write_id_cell(ws, row_idx, 1, result.article_number)
            ws.cell(row=row_idx, column=2, value=result.product_data.product_name or "")
            ws.cell(row=row_idx, column=3, value=producer or "")
            ws.cell(row=row_idx, column=4, value=producer_artnr or "")
            ws.cell(row=row_idx, column=5, value=img.get("image_name", ""))
            ws.cell(row=row_idx, column=6, value=img.get("image_url", ""))
            ws.cell(row=row_idx, column=7, value="Ja" if img.get("exists") else "Nei")
            ws.cell(row=row_idx, column=8, value=img.get("file_size_kb", 0))
            ws.cell(row=row_idx, column=9, value=img.get("width", 0))
            ws.cell(row=row_idx, column=10, value=img.get("height", 0))
            ws.cell(row=row_idx, column=11, value=img.get("aspect_ratio", 0))
            ws.cell(row=row_idx, column=12, value=img.get("resolution_score", 0))
            ws.cell(row=row_idx, column=13, value=img.get("blur_score_raw", 0))
            ws.cell(row=row_idx, column=14, value=img.get("blur_score", 0))
            ws.cell(row=row_idx, column=15, value=img.get("brightness_mean", 0))
            ws.cell(row=row_idx, column=16, value=img.get("brightness_score", 0))
            ws.cell(row=row_idx, column=17, value=img.get("contrast_std", 0))
            ws.cell(row=row_idx, column=18, value=img.get("contrast_score", 0))
            ws.cell(row=row_idx, column=19, value=img.get("white_bg_ratio", 0))
            ws.cell(row=row_idx, column=20, value=img.get("background_score", 0))
            ws.cell(row=row_idx, column=21, value=img.get("edge_density", 0))
            ws.cell(row=row_idx, column=22, value=img.get("edge_score", 0))
            ws.cell(row=row_idx, column=23, value=img.get("product_fill_ratio", 0))
            ws.cell(row=row_idx, column=24, value=img.get("fill_score", 0))
            ws.cell(row=row_idx, column=25, value=img.get("overall_score", 0))
            status = img.get("status", "MISSING")
            status_cell = ws.cell(row=row_idx, column=26, value=status)
            _apply_image_status_style(status_cell, status)
            issues = img.get("issues", [])
            ws.cell(row=row_idx, column=27, value=", ".join(issues) if issues else "")
            row_idx += 1

    if row_idx == 2:
        ws.cell(row=2, column=1, value="Ingen bildedata tilgjengelig")

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = max(12, len(headers[col - 1]) + 3)

    ws.freeze_panes = "A2"
    if row_idx > 2:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row_idx - 1}"

    # ── Image Improvement Suggestions section ──
    # Separate table below main data with producer info and concrete suggestions
    sugg_start = row_idx + 2
    sugg_headers = [
        "Artikkelnummer",
        "Produktnavn",
        "Produsent",
        "Produsentens varenummer",
        "Bildeproblem",
        "Nåværende bilde-URL",
        "Nåværende status",
        "Foreslått bilde-URL",
        "Foreslått bildekilde",
        "Kilde-URL",
        "Bildesøk confidence",
        "Forbedringsforslag",
        "Krever manuell vurdering",
    ]
    for col, h in enumerate(sugg_headers, 1):
        ws.cell(row=sugg_start, column=col, value=h)
    _style_header(ws, sugg_start, len(sugg_headers))

    sugg_row = sugg_start + 1
    for result in results:
        iq = result.image_quality or {}
        img_sugg = result.image_suggestion
        img_status = iq.get("image_quality_status", "MISSING")
        img_issues = iq.get("image_issue_summary", "")

        # Include rows for ALL products with image problems (not just those with suggestions)
        has_problem = img_status in ("MISSING", "FAIL", "REVIEW")
        has_suggestion = img_sugg is not None

        if not has_problem and not has_suggestion:
            continue

        producer, producer_artnr = get_best_producer_info(
            result.product_data, result.jeeves_data, result.manufacturer_lookup
        )

        # Determine the image problem description
        if img_status == "MISSING":
            problem_desc = "Bilde mangler"
        elif img_status == "FAIL":
            problem_desc = f"Bildekvalitet for lav: {img_issues}" if img_issues else "Bildekvalitet for lav"
        elif img_status == "REVIEW":
            problem_desc = f"Bilde bør gjennomgås: {img_issues}" if img_issues else "Bilde bør gjennomgås"
        else:
            problem_desc = img_issues or ""

        # Build improvement suggestion text
        if has_suggestion and img_sugg.suggested_image_url:
            improvement = (
                f"Bedre bilde funnet hos {img_sugg.suggested_source or 'ekstern kilde'}. "
                f"URL: {img_sugg.suggested_image_url}"
            )
        elif has_problem:
            # No image suggestion found, but there's a problem — suggest manual action
            search_hint = ""
            if producer and producer_artnr:
                search_hint = f"Søk hos {producer} med varenummer {producer_artnr}"
            elif producer:
                search_hint = f"Søk hos {producer} med produktnavn"
            else:
                search_hint = "Kontakt produsent for produktbilde"
            improvement = f"Krever manuell bildesøk. {search_hint}"
        else:
            improvement = ""

        _write_id_cell(ws, sugg_row, 1, result.article_number)
        ws.cell(row=sugg_row, column=2, value=result.product_data.product_name or "")
        ws.cell(row=sugg_row, column=3, value=producer or "")
        ws.cell(row=sugg_row, column=4, value=producer_artnr or "")
        ws.cell(row=sugg_row, column=5, value=problem_desc)
        ws.cell(row=sugg_row, column=6, value=(img_sugg.current_image_url if img_sugg else result.product_data.image_url) or "")
        ws.cell(row=sugg_row, column=7, value=(img_sugg.current_image_status if img_sugg else img_status) or "")
        ws.cell(row=sugg_row, column=8, value=(img_sugg.suggested_image_url if img_sugg else "") or "")
        ws.cell(row=sugg_row, column=9, value=(img_sugg.suggested_source if img_sugg else "") or "")
        ws.cell(row=sugg_row, column=10, value=(img_sugg.suggested_source_url if img_sugg else "") or "")
        ws.cell(row=sugg_row, column=11, value=(img_sugg.confidence if img_sugg and img_sugg.confidence else 0))
        ws.cell(row=sugg_row, column=12, value=improvement)
        ws.cell(row=sugg_row, column=13, value="Ja" if not has_suggestion or (img_sugg and img_sugg.review_required) else "Nei")
        sugg_row += 1


def _create_image_issues_sheet(ws, results: list[ProductAnalysis]) -> None:
    """Create prioritized image issues sheet."""
    headers = [
        "Prioritet",
        "Artikkelnummer",
        "Produktnavn",
        "Bildestatus",
        "Bildescore",
        "Antall bilder",
        "Problemer",
    ]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    _style_header(ws, 1, len(headers))

    # Collect products with image issues, sorted by priority
    issue_rows = []
    for result in results:
        iq = result.image_quality or {}
        status = iq.get("image_quality_status", "MISSING")
        priority = iq.get("image_quality_priority", "none")
        if priority == "none":
            continue
        priority_order = {"high": 0, "medium": 1, "low": 2}
        issue_rows.append((
            priority_order.get(priority, 3),
            priority,
            result,
            iq,
        ))

    issue_rows.sort(key=lambda x: (x[0], -x[3].get("avg_image_score", 0)))

    row_idx = 2
    priority_labels = {"high": "H\u00f8y", "medium": "Medium", "low": "Lav"}
    for _, priority, result, iq in issue_rows:
        label = priority_labels.get(priority, priority)
        ws.cell(row=row_idx, column=1, value=label)
        _write_id_cell(ws, row_idx, 2, result.article_number)
        ws.cell(row=row_idx, column=3, value=result.product_data.product_name or "")
        status = iq.get("image_quality_status", "MISSING")
        status_cell = ws.cell(row=row_idx, column=4, value=status)
        _apply_image_status_style(status_cell, status)
        ws.cell(row=row_idx, column=5, value=round(iq.get("avg_image_score", 0), 1))
        ws.cell(row=row_idx, column=6, value=iq.get("image_count_found", 0))
        ws.cell(row=row_idx, column=7, value=iq.get("image_issue_summary", ""))
        row_idx += 1

    if row_idx == 2:
        ws.cell(row=2, column=1, value="Ingen bildeproblemer funnet")

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = max(15, len(headers[col - 1]) + 5)

    ws.freeze_panes = "A2"


ENRICHMENT_STATUS_COLORS = {
    "FOUND_IN_INTERNAL_PDF": "C6EFCE",
    "FOUND_IN_MANUFACTURER_SOURCE": "D9E2F3",
    "FOUND_IN_BOTH_MATCH": "C6EFCE",
    "FOUND_IN_BOTH_CONFLICT": "FFEB9C",
    "NOT_FOUND": "F2F2F2",
    "REVIEW_REQUIRED": "FFC7CE",
}


def _apply_enrichment_status_style(cell, status: str) -> None:
    fill_color = ENRICHMENT_STATUS_COLORS.get(status, "FFFFFF")
    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")


def _create_comparison_and_enrichment_sheet(ws, results: list[ProductAnalysis]) -> None:
    """Create the main comparison + enrichment sheet.

    One row per product. Shows Jeeves source values, website source values,
    field statuses, enrichment suggestions, and traceability columns.
    Every product is included — not just those with suggestions.
    """
    headers = [
        # IDENTIFIERS (1-2)
        "Artikkelnummer",
        "GID",
        # JEEVES SOURCE VALUES (3-9)
        "Jeeves_Item_description",
        "Jeeves_Specification",
        "Jeeves_Supplier",
        "Jeeves_Supplier_Item_no",
        "Jeeves_Product_Brand",
        "Jeeves_Web_Title",
        "Jeeves_Web_Text",
        # WEBSITE SOURCE VALUES (10-17)
        "Website_URL",
        "Website_Title",
        "Website_Breadcrumb",
        "Website_Description",
        "Website_Specification",
        "Website_Packaging",
        "Website_Image_Present",
        "Website_Datasheet_URL",
        # STATUS COLUMNS (18-26)
        "Status_Produktnavn",
        "Status_Beskrivelse",
        "Status_Spesifikasjon",
        "Status_Produsent",
        "Status_Produsent_Artnr",
        "Status_Merkevare",
        "Status_Kategori",
        "Status_Pakningsinformasjon",
        "Status_Bilde",
        # SUGGESTION COLUMNS (27-34)
        "Forslag_Produktnavn",
        "Forslag_Beskrivelse",
        "Forslag_Spesifikasjon",
        "Forslag_Produsent",
        "Forslag_Produsent_Artnr",
        "Forslag_Merkevare",
        "Forslag_Kategori",
        "Forslag_Pakningsinformasjon",
        # TRACEABILITY (35-38)
        "Kilde_for_forslag",
        "Confidence",
        "Review_Required",
        "Kommentar",
    ]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # Section-based header colors
    section_colors = {
        range(1, 3): "4472C4",     # Identifiers: blue
        range(3, 10): "548235",    # Jeeves: green
        range(10, 18): "BF8F00",   # Website: gold
        range(18, 27): "7030A0",   # Status: purple
        range(27, 35): "C55A11",   # Suggestions: orange
        range(35, 39): "404040",   # Traceability: dark gray
    }
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for col_range, color in section_colors.items():
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for col in col_range:
            cell = ws.cell(row=1, column=col)
            cell.fill = fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Fills for suggestion / status cells
    suggestion_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    review_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    wrap_align = Alignment(wrap_text=True, vertical="top")
    top_align = Alignment(vertical="top")

    row_idx = 2
    for result in results:
        product = result.product_data
        jeeves = result.jeeves_data

        # Build field status lookup
        fa_by_name = {fa.field_name: fa for fa in result.field_analyses}

        # Build suggestion lookup from enrichment_suggestions + field_analyses
        sugg_by_field = {}
        for es in result.enrichment_suggestions:
            sugg_by_field[es.field_name] = es
        for fa in result.field_analyses:
            if fa.suggested_value and fa.field_name not in sugg_by_field:
                sugg_by_field[fa.field_name] = EnrichmentSuggestion(
                    field_name=fa.field_name,
                    current_value=fa.current_value,
                    suggested_value=fa.suggested_value,
                    source=fa.source,
                    confidence=fa.confidence or 0.0,
                    review_required=True,
                )

        # Website-derived values
        web_spec = product.specification or ""
        if not web_spec and product.technical_details:
            web_spec = "; ".join(f"{k}: {v}" for k, v in product.technical_details.items())
        web_breadcrumb = " > ".join(product.category_breadcrumb) if product.category_breadcrumb else ""
        web_packaging = product.packaging_info or product.packaging_unit or ""
        web_image_present = "Ja" if product.image_url else "Nei"

        def _status_val(field_name):
            fa = fa_by_name.get(field_name)
            if not fa:
                return ""
            return fa.status.value

        def _sugg_val(field_name):
            es = sugg_by_field.get(field_name)
            return es.suggested_value if es else ""

        # Build comment and sources for traceability
        sources = set()
        comments = []
        min_confidence = 1.0
        any_review = False

        for es in sugg_by_field.values():
            if es.source:
                sources.add(es.source)
            if es.confidence is not None:
                min_confidence = min(min_confidence, es.confidence)
            if es.review_required:
                any_review = True

        # Generate per-field comments
        field_comment_parts = []
        for fa in result.field_analyses:
            if fa.field_name in ("Konsistens mellom felter",):
                continue
            src = fa.source or ""
            if fa.status in (QualityStatus.STRONG, QualityStatus.OK) and "Jeeves kun" in src:
                field_comment_parts.append(f"{fa.field_name}: present in Jeeves only")
            elif fa.status in (QualityStatus.STRONG, QualityStatus.OK) and "nettside kun" in src:
                field_comment_parts.append(f"{fa.field_name}: present on website only")
            elif fa.status == QualityStatus.MISSING:
                field_comment_parts.append(f"{fa.field_name}: missing in both sources")
            elif fa.status == QualityStatus.WEAK:
                field_comment_parts.append(f"{fa.field_name}: present but weak — {fa.comment}")
            elif fa.status == QualityStatus.SHOULD_IMPROVE:
                field_comment_parts.append(f"{fa.field_name}: {fa.comment}")

        # Check for Jeeves vs website conflicts
        if jeeves and product.product_name and jeeves.item_description:
            if product.product_name.lower().strip() != jeeves.item_description.lower().strip():
                if jeeves.web_title and product.product_name.lower().strip() != jeeves.web_title.lower().strip():
                    field_comment_parts.append(
                        "Produktnavn: website and Jeeves differ"
                    )
                    any_review = True

        comment_text = "; ".join(field_comment_parts) if field_comment_parts else ""

        # --- Write row ---
        c = 1
        # IDENTIFIERS — formatted as text to prevent Excel auto-conversion
        _write_id_cell(ws, row_idx, c, result.article_number, top_align); c += 1
        _write_id_cell(ws, row_idx, c, jeeves.gid if jeeves else "", top_align); c += 1
        # JEEVES SOURCE VALUES
        ws.cell(row=row_idx, column=c, value=jeeves.item_description if jeeves else "").alignment = top_align; c += 1
        ws.cell(row=row_idx, column=c, value=jeeves.specification if jeeves else "").alignment = wrap_align; c += 1
        ws.cell(row=row_idx, column=c, value=jeeves.supplier if jeeves else "").alignment = top_align; c += 1
        _write_id_cell(ws, row_idx, c, jeeves.supplier_item_no if jeeves else "", top_align); c += 1
        ws.cell(row=row_idx, column=c, value=jeeves.product_brand if jeeves else "").alignment = top_align; c += 1
        ws.cell(row=row_idx, column=c, value=jeeves.web_title if jeeves else "").alignment = top_align; c += 1
        ws.cell(row=row_idx, column=c, value=jeeves.web_text if jeeves else "").alignment = wrap_align; c += 1
        # WEBSITE SOURCE VALUES
        ws.cell(row=row_idx, column=c, value=product.product_url or "").alignment = top_align; c += 1
        ws.cell(row=row_idx, column=c, value=product.product_name or "").alignment = top_align; c += 1
        ws.cell(row=row_idx, column=c, value=web_breadcrumb).alignment = top_align; c += 1
        ws.cell(row=row_idx, column=c, value=product.description or "").alignment = wrap_align; c += 1
        ws.cell(row=row_idx, column=c, value=web_spec).alignment = wrap_align; c += 1
        ws.cell(row=row_idx, column=c, value=web_packaging).alignment = top_align; c += 1
        ws.cell(row=row_idx, column=c, value=web_image_present).alignment = top_align; c += 1
        ws.cell(row=row_idx, column=c, value=result.pdf_url or "").alignment = top_align; c += 1
        # STATUS COLUMNS
        status_fields = [
            "Produktnavn", "Beskrivelse", "Spesifikasjon", "Produsent",
            "Produsentens varenummer", "Merkevare", "Kategori",
            "Pakningsinformasjon", "Bildekvalitet",
        ]
        for sf in status_fields:
            status_text = _status_val(sf)
            cell = ws.cell(row=row_idx, column=c, value=status_text)
            cell.alignment = top_align
            fa = fa_by_name.get(sf)
            if fa:
                _apply_status_style(cell, fa.status)
            c += 1
        # SUGGESTION COLUMNS
        suggestion_fields = [
            "Produktnavn", "Beskrivelse", "Spesifikasjon", "Produsent",
            "Produsentens varenummer", "Merkevare", "Kategori", "Pakningsinformasjon",
        ]
        for sf in suggestion_fields:
            val = _sugg_val(sf) or ""
            cell = ws.cell(row=row_idx, column=c, value=val)
            cell.alignment = wrap_align
            if val:
                cell.fill = review_fill if any_review else suggestion_fill
            c += 1
        # TRACEABILITY
        ws.cell(row=row_idx, column=c, value="; ".join(sorted(sources)) if sources else "").alignment = top_align; c += 1
        conf_val = round(min_confidence, 2) if sugg_by_field else ""
        ws.cell(row=row_idx, column=c, value=conf_val).alignment = top_align; c += 1
        ws.cell(row=row_idx, column=c, value="Ja" if any_review else "Nei").alignment = top_align; c += 1
        ws.cell(row=row_idx, column=c, value=comment_text).alignment = wrap_align; c += 1

        # Verify column count matches header count
        assert c == len(headers) + 1, (
            f"Column count mismatch in Comparison_And_Enrichment: "
            f"wrote {c - 1} columns, expected {len(headers)}"
        )

        row_idx += 1

    # Column widths
    col_widths = {
        1: 16, 2: 12,                                    # identifiers
        3: 25, 4: 20, 5: 22, 6: 18, 7: 16, 8: 25, 9: 35,  # Jeeves
        10: 35, 11: 25, 12: 30, 13: 35, 14: 35, 15: 25, 16: 14, 17: 35,  # website
        18: 14, 19: 14, 20: 14, 21: 14, 22: 14, 23: 14, 24: 14, 25: 14, 26: 14,  # status
        27: 25, 28: 35, 29: 30, 30: 20, 31: 18, 32: 16, 33: 20, 34: 25,  # suggestions
        35: 25, 36: 12, 37: 14, 38: 50,  # traceability
    }
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "C2"
    if row_idx > 2:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row_idx - 1}"


def _create_conflicts_sheet(ws, results: list[ProductAnalysis]) -> None:
    """Create source conflict sheet - shows fields where PDF and manufacturer disagree."""
    headers = [
        "Artikkelnummer",
        "Produktnavn",
        "Felt",
        "N\u00e5v\u00e6rende verdi",
        "PDF-verdi",
        "Produsent-verdi",
        "Evidens",
        "Anbefaling",
    ]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    _style_header(ws, 1, len(headers))

    row_idx = 2
    for result in results:
        for er in result.enrichment_results:
            if er.match_status != "FOUND_IN_BOTH_CONFLICT":
                continue
            _write_id_cell(ws, row_idx, 1, result.article_number)
            ws.cell(row=row_idx, column=2, value=result.product_data.product_name or "")
            ws.cell(row=row_idx, column=3, value=er.field_name)
            ws.cell(row=row_idx, column=4, value=er.current_value or "")
            # Parse the evidence snippet to extract both values
            evidence = er.evidence_snippet or ""
            pdf_val = ""
            mfr_val = ""
            if "KONFLIKT" in evidence:
                parts = evidence.split(" vs ")
                if len(parts) == 2:
                    pdf_val = parts[0].replace("KONFLIKT - PDF: ", "").strip("'")
                    mfr_val = parts[1].replace("Produsent: ", "").strip("'")
            ws.cell(row=row_idx, column=5, value=pdf_val)
            ws.cell(row=row_idx, column=6, value=mfr_val)
            ws.cell(row=row_idx, column=7, value=evidence)
            ws.cell(row=row_idx, column=8, value="Manuell vurdering p\u00e5krevd")
            # Color the row
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = PatternFill(
                    start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
                )
            row_idx += 1

    if row_idx == 2:
        ws.cell(row=2, column=1, value="Ingen kildekonflikter funnet")

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = max(16, len(headers[col - 1]) + 4)

    ws.freeze_panes = "A2"


def _create_summary_sheet(ws, results: list[ProductAnalysis],
                          analysis_mode: str = "full_enrichment",
                          focus_areas: list[str] | None = None) -> None:
    """Create the Summary sheet with high-level two-source metrics."""
    title_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=12, color="4472C4")
    label_font = Font(bold=True, size=11)
    value_font = Font(size=11)
    indent_font = Font(size=11, color="404040")

    mode_titles = {
        "full_enrichment": "Masterdata Kvalitetsrapport — Full berikelse",
        "audit_only": "Masterdata Kvalitetsrapport — Kvalitetsrevisjon",
        "focused_scan": "Masterdata Kvalitetsrapport — Fokusert sjekk",
    }
    ws.cell(row=1, column=1, value=mode_titles.get(analysis_mode, "Masterdata Kvalitetsrapport")).font = title_font
    ws.cell(row=2, column=1, value=f"Generert: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = indent_font
    if focus_areas:
        from backend.scoring import AREA_LABELS
        area_labels = [AREA_LABELS.get(a, a) for a in focus_areas]
        ws.cell(row=3, column=1, value=f"Fokusomr\u00e5der: {', '.join(area_labels)}").font = indent_font

    total = len(results)
    found = sum(1 for r in results if r.product_data.found_on_onemed)
    not_found = total - found
    has_jeeves = sum(1 for r in results if r.jeeves_data)
    avg_score = sum(r.total_score for r in results) / total if total else 0

    # Per-field: count how many have website data
    web_desc = sum(1 for r in results if r.product_data.description)
    web_spec = sum(1 for r in results if r.product_data.specification or r.product_data.technical_details)
    web_pkg = sum(1 for r in results if r.product_data.packaging_info or r.product_data.packaging_unit)
    web_img = sum(1 for r in results if r.product_data.image_url)
    web_cat = sum(1 for r in results if r.product_data.category or r.product_data.category_breadcrumb)

    # Suggestions
    products_with_suggestions = sum(1 for r in results if r.enrichment_suggestions)
    total_suggestions = sum(len(r.enrichment_suggestions) for r in results)
    manual_review = sum(
        1 for r in results
        if r.manual_review_needed or any(es.review_required for es in r.enrichment_suggestions)
    )

    # Status distribution
    status_counts = {}
    for r in results:
        status = r.overall_status.value
        status_counts[status] = status_counts.get(status, 0) + 1

    # Per-field status distribution
    field_status_counts: dict[str, dict[str, int]] = {}
    for r in results:
        for fa in r.field_analyses:
            if fa.field_name == "Konsistens mellom felter":
                continue
            if fa.field_name not in field_status_counts:
                field_status_counts[fa.field_name] = {}
            s = fa.status.value
            field_status_counts[fa.field_name][s] = field_status_counts[fa.field_name].get(s, 0) + 1

    # Priority distribution from area scoring
    priority_counts = {"Kritisk": 0, "Høy": 0, "Middels": 0, "Lav": 0}
    area_score_sums: dict[str, list[float]] = {}
    for r in results:
        asd = (r.ai_score or {}).get("area_scores", {})
        pl = asd.get("priority_level", "")
        if pl in priority_counts:
            priority_counts[pl] += 1
        for a in asd.get("area_scores", []):
            label = a.get("area_label", "")
            if label:
                area_score_sums.setdefault(label, []).append(a.get("score", 0))

    # Build summary rows as (label, value, indent_level)
    rows: list[tuple[str, str | int | float, int]] = [
        # Run metadata
        ("KJØREMETADATA", "", 0),
        ("Analysemodus", mode_titles.get(analysis_mode, analysis_mode), 1),
        ("Dato/tid", datetime.now().strftime('%Y-%m-%d %H:%M'), 1),
        ("Totalt antall produkter", total, 1),
        ("Produkter i Jeeves", has_jeeves, 1),
        ("Gjennomsnittlig kvalitetsscore", f"{avg_score:.1f}%", 1),
        ("", "", 0),
        # Priority distribution
        ("PRIORITETSFORDELING", "", 0),
        ("Kritisk", priority_counts["Kritisk"], 1),
        ("Høy", priority_counts["Høy"], 1),
        ("Middels", priority_counts["Middels"], 1),
        ("Lav", priority_counts["Lav"], 1),
        ("", "", 0),
        # Average area scores
        ("GJENNOMSNITTLIG OMRÅDESCORE", "", 0),
    ]
    for area_label, scores in sorted(area_score_sums.items()):
        avg = sum(scores) / len(scores) if scores else 0
        rows.append((area_label, f"{avg:.1f}/100", 1))

    rows.append(("", "", 0))

    # Issue summary (missing fields)
    rows.append(("MANGLENDE FELTER", "", 0))
    rows.append(("Mangler beskrivelse", total - web_desc, 1))
    rows.append(("Mangler spesifikasjon", total - web_spec, 1))
    rows.append(("Mangler bilde", total - web_img, 1))
    rows.append(("Mangler kategori", total - web_cat, 1))
    rows.append(("Mangler pakningsinformasjon", total - web_pkg, 1))
    rows.append(("", "", 0))

    # Website coverage
    rows.append(("NETTSIDE-DEKNING", "", 0))
    rows.append(("Funnet på onemed.no", found, 1))
    rows.append(("Ikke funnet på onemed.no", not_found, 1))
    rows.append(("", "", 0))

    if analysis_mode == "full_enrichment":
        # Enrichment
        rows.append(("BERIKELSE OG FORSLAG", "", 0))
        rows.append(("Produkter med forslag", products_with_suggestions, 1))
        rows.append(("Totalt antall forslag", total_suggestions, 1))
        rows.append(("Krever manuell gjennomgang", manual_review, 1))
        rows.append(("", "", 0))

    # Status distribution
    rows.append(("STATUSFORDELING (OVERORDNET)", "", 0))
    for status_name, count in sorted(status_counts.items()):
        rows.append((status_name, count, 1))

    rows.append(("", "", 0))
    rows.append(("STATUSFORDELING PER FELT", "", 0))
    for field_name in [
        "Produktnavn", "Beskrivelse", "Spesifikasjon", "Produsent",
        "Produsentens varenummer", "Merkevare", "Kategori",
        "Pakningsinformasjon", "Bildekvalitet",
    ]:
        counts = field_status_counts.get(field_name, {})
        if not counts:
            continue
        ok = counts.get("OK", 0)
        missing = counts.get("Mangler", 0)
        improve = counts.get("Bør forbedres", 0)
        error = counts.get("Sannsynlig feil", 0)
        rows.append((field_name, f"OK: {ok} | Mangler: {missing} | Forbedres: {improve} | Feil: {error}", 1))

    # Write rows
    row_num = 4
    for label, value, indent in rows:
        if not label and not value:
            row_num += 1
            continue
        cell_label = ws.cell(row=row_num, column=1, value=label)
        cell_value = ws.cell(row=row_num, column=2, value=value)
        if indent == 0:
            cell_label.font = section_font
        elif indent == 1:
            cell_label.font = label_font
            cell_value.font = value_font
        row_num += 1

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 55


def _create_area_scores_sheet(ws, results: list[ProductAnalysis],
                              focus_areas: list[str] | None = None) -> None:
    """Create the Area Scores sheet for audit/focused modes.

    One row per product with per-area scores, status, issues, and recommendations.
    """
    from backend.scoring import AREA_LABELS, ALL_AREAS

    # Determine which areas to show
    areas_to_show = focus_areas if focus_areas else ALL_AREAS

    # Build headers: Article Number, Product Name, Priority, Overall Score, Why Low, then per-area columns
    headers = ["Artikkelnummer", "Produktnavn", "Prioritet", "Totalscore", "Alvorlighetsgrad", "Forklaring"]
    for area in areas_to_show:
        label = AREA_LABELS.get(area, area)
        headers.append(f"{label} Score")
        headers.append(f"{label} Status")
        headers.append(f"{label} Problemer")
        headers.append(f"{label} Anbefaling")

    headers.append("Manglende omr\u00e5der")
    headers.append("Oppsummering")

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    _style_header(ws, 1, len(headers))

    for row_idx, result in enumerate(results, 2):
        _write_id_cell(ws, row_idx, 1, result.article_number)
        ws.cell(row=row_idx, column=2, value=result.product_data.product_name or "")

        # Get area scores from the ai_score dict (populated by scoring framework)
        area_score_data = {}
        overall_score = result.total_score
        overall_severity = ""
        missing_areas = []
        issue_summary = ""
        priority_level = ""
        why_low = ""

        if result.ai_score and "area_scores" in result.ai_score:
            asd = result.ai_score["area_scores"]
            overall_score = asd.get("overall_score", result.total_score)
            overall_severity = asd.get("overall_severity", "")
            missing_areas = asd.get("missing_areas", [])
            issue_summary = asd.get("issue_summary", "")
            priority_level = asd.get("priority_level", "")
            why_low = asd.get("why_low", "")
            for a in asd.get("area_scores", []):
                area_score_data[a["area"]] = a

        # Priority column with color-coding
        priority_cell = ws.cell(row=row_idx, column=3, value=priority_level)
        priority_colors = {
            "Kritisk": ("FF6B6B", "FFFFFF"),
            "Høy": ("FFC7CE", "9C0006"),
            "Middels": ("FFEB9C", "9C6500"),
            "Lav": ("C6EFCE", "006100"),
        }
        if priority_level in priority_colors:
            bg, fg = priority_colors[priority_level]
            priority_cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
            priority_cell.font = Font(color=fg, bold=True)

        # Overall score with color-coding
        score_cell = ws.cell(row=row_idx, column=4, value=round(overall_score, 1))
        if overall_score >= 75:
            score_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif overall_score >= 50:
            score_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        else:
            score_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        ws.cell(row=row_idx, column=5, value=overall_severity)
        ws.cell(row=row_idx, column=6, value=why_low)

        col = 7
        for area in areas_to_show:
            a_data = area_score_data.get(area, {})
            a_score = a_data.get("score", 0)
            a_status = a_data.get("status", "Mangler")
            a_issues = "; ".join(i["description"] for i in a_data.get("issues", []))
            a_action = a_data.get("recommended_action", "")

            score_cell = ws.cell(row=row_idx, column=col, value=round(a_score, 1))
            if a_score >= 75:
                score_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif a_score >= 40:
                score_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            elif a_score > 0:
                score_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            else:
                score_cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                score_cell.font = Font(color="FFFFFF")

            ws.cell(row=row_idx, column=col + 1, value=a_status)
            ws.cell(row=row_idx, column=col + 2, value=a_issues[:200] if a_issues else "")
            ws.cell(row=row_idx, column=col + 3, value=a_action)
            col += 4

        ws.cell(row=row_idx, column=col, value=", ".join(missing_areas) if missing_areas else "")
        ws.cell(row=row_idx, column=col + 1, value=issue_summary)

    # Auto-width for key columns
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 35
    # Freeze header row and set auto-filter for usability
    ws.freeze_panes = "A2"
    if len(headers) > 0:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _create_debug_log_sheet(ws, results: list[ProductAnalysis]) -> None:
    """Create the Debug_Log sheet for traceability.

    One row per product. Shows raw extraction details so a reviewer can
    understand why the product got its status and suggestions.
    """
    headers = [
        "Artikkelnummer",
        "Website_URL",
        "Product_Found_On_Website",
        "Accordion_Expanded",
        "Extracted_Title_Raw",
        "Extracted_Breadcrumb_Raw",
        "Extracted_Description_Raw",
        "Extracted_Specification_Raw",
        "Extracted_Packaging_Raw",
        "Image_Found",
        "Datasheet_URL_Found",
        "Jeeves_Fields_Used",
        "Website_Fields_Used",
        "Suggestion_Source",
        "Review_Required",
        "Quality_Score",
        "Debug_Comment",
    ]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    _style_header(ws, 1, len(headers))

    wrap_align = Alignment(wrap_text=True, vertical="top")
    top_align = Alignment(vertical="top")

    for row_idx, result in enumerate(results, 2):
        product = result.product_data
        jeeves = result.jeeves_data

        # Determine which Jeeves fields are populated
        jeeves_fields_used = []
        if jeeves:
            if jeeves.item_description:
                jeeves_fields_used.append("Item description")
            if jeeves.specification:
                jeeves_fields_used.append("Specification")
            if jeeves.supplier:
                jeeves_fields_used.append("Supplier")
            if jeeves.supplier_item_no:
                jeeves_fields_used.append("Supplier Item.no")
            if jeeves.product_brand:
                jeeves_fields_used.append("Product Brand")
            if jeeves.web_title:
                jeeves_fields_used.append("Web Title")
            if jeeves.web_text:
                jeeves_fields_used.append("Web Text")

        # Determine which website fields are populated
        website_fields_used = []
        if product.product_name:
            website_fields_used.append("Title")
        if product.description:
            website_fields_used.append("Description")
        if product.specification or product.technical_details:
            website_fields_used.append("Specification")
        if product.category or product.category_breadcrumb:
            website_fields_used.append("Category")
        if product.packaging_info or product.packaging_unit:
            website_fields_used.append("Packaging")
        if product.image_url:
            website_fields_used.append("Image")
        if product.manufacturer:
            website_fields_used.append("Manufacturer")

        # Accordion: website description from accordion selector
        accordion_status = "N/A"
        if product.found_on_onemed:
            if product.description and len(product.description) > 10:
                accordion_status = "Yes"
            else:
                accordion_status = "No content found"

        # Build suggestion source summary
        suggestion_sources = set()
        any_review = False
        for es in result.enrichment_suggestions:
            if es.source:
                suggestion_sources.add(es.source)
            if es.review_required:
                any_review = True
        for fa in result.field_analyses:
            if fa.source and fa.suggested_value:
                suggestion_sources.add(fa.source)

        # Debug comment: explain notable findings
        debug_parts = []
        if not product.found_on_onemed:
            debug_parts.append("Not found on website")
        if not jeeves:
            debug_parts.append("Not in Jeeves")

        # Note fields where Jeeves fills gaps
        for fa in result.field_analyses:
            if fa.source and "Jeeves kun" in fa.source:
                debug_parts.append(f"{fa.field_name} from Jeeves only")
            elif fa.status == QualityStatus.MISSING:
                debug_parts.append(f"{fa.field_name} missing in both")

        # Note weak/poor fields
        for fa in result.field_analyses:
            if fa.status in (QualityStatus.WEAK, QualityStatus.SHOULD_IMPROVE):
                debug_parts.append(f"{fa.field_name}: {fa.comment}")

        # Raw website extraction values
        raw_spec = product.specification or ""
        if not raw_spec and product.technical_details:
            raw_spec = "; ".join(f"{k}: {v}" for k, v in product.technical_details.items())
        raw_breadcrumb = " > ".join(product.category_breadcrumb) if product.category_breadcrumb else ""
        raw_packaging = product.packaging_info or product.packaging_unit or ""

        c = 1
        _write_id_cell(ws, row_idx, c, result.article_number, top_align); c += 1
        ws.cell(row=row_idx, column=c, value=product.product_url or "").alignment = top_align; c += 1
        ws.cell(row=row_idx, column=c, value="Ja" if product.found_on_onemed else "Nei").alignment = top_align; c += 1
        ws.cell(row=row_idx, column=c, value=accordion_status).alignment = top_align; c += 1
        ws.cell(row=row_idx, column=c, value=product.product_name or "").alignment = top_align; c += 1
        ws.cell(row=row_idx, column=c, value=raw_breadcrumb).alignment = top_align; c += 1
        ws.cell(row=row_idx, column=c, value=product.description or "").alignment = wrap_align; c += 1
        ws.cell(row=row_idx, column=c, value=raw_spec).alignment = wrap_align; c += 1
        ws.cell(row=row_idx, column=c, value=raw_packaging).alignment = top_align; c += 1
        ws.cell(row=row_idx, column=c, value="Ja" if product.image_url else "Nei").alignment = top_align; c += 1
        ws.cell(row=row_idx, column=c, value=result.pdf_url or "").alignment = top_align; c += 1
        ws.cell(row=row_idx, column=c, value=", ".join(jeeves_fields_used) if jeeves_fields_used else "None").alignment = wrap_align; c += 1
        ws.cell(row=row_idx, column=c, value=", ".join(website_fields_used) if website_fields_used else "None").alignment = wrap_align; c += 1
        ws.cell(row=row_idx, column=c, value=", ".join(sorted(suggestion_sources)) if suggestion_sources else "None").alignment = top_align; c += 1
        ws.cell(row=row_idx, column=c, value="Ja" if any_review else "Nei").alignment = top_align; c += 1
        ws.cell(row=row_idx, column=c, value=result.total_score).alignment = top_align; c += 1
        ws.cell(row=row_idx, column=c, value="; ".join(debug_parts) if debug_parts else "No issues").alignment = wrap_align; c += 1

    # Column widths
    col_widths = {
        1: 16, 2: 35, 3: 14, 4: 14, 5: 25, 6: 30, 7: 35,
        8: 35, 9: 25, 10: 12, 11: 35, 12: 30, 13: 30,
        14: 25, 15: 14, 16: 12, 17: 55,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "B2"
    if len(results) > 0:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(results) + 1}"


# --- Inriver Import colors ---
INRIVER_STATUS_COLORS = {
    "Ready for Inriver": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "Needs Review": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "No Change": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
    "Missing Source Data": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
}

INRIVER_STATUS_FONTS = {
    "Ready for Inriver": Font(color="006100"),
    "Needs Review": Font(color="9C6500"),
    "No Change": Font(color="808080"),
    "Missing Source Data": Font(color="9C0006"),
}


def _get_suggestion_for_field(result: ProductAnalysis, field_name: str) -> Optional[str]:
    """Get the best suggestion for a field from all enrichment sources.

    Priority order (source-grounded first, AI as fallback):
    1. Enrichment suggestions (AI-reviewed, source-grounded, quality-gated)
    2. Field analysis suggestions (from PDF/manufacturer enrichment)
    3. Enrichment results (raw source data)
    4. AI enrichment (independent AI suggestions — used only as last resort)

    AI enrichment is demoted to Priority 4 because it operates on raw product
    data without source evidence, making it less reliable than source-grounded
    suggestions that have been through the quality gate.
    """
    # Priority 1: Enrichment suggestions (AI-reviewed, quality-gated)
    for es in result.enrichment_suggestions:
        if es.field_name == field_name and es.suggested_value:
            return es.suggested_value

    # Priority 2: Field analysis suggestions (from PDF/manufacturer enrichment)
    for fa in result.field_analyses:
        if fa.field_name == field_name and fa.suggested_value:
            return fa.suggested_value

    # Priority 3: Enrichment results (raw source extractions)
    enrichment_field_map = {
        "Produktnavn": "product_name",
        "Beskrivelse": "description",
        "Produsent": "manufacturer",
        "Produsentens varenummer": "manufacturer_article_number",
        "Pakningsinformasjon": "packaging_info",
    }
    er_key = enrichment_field_map.get(field_name)
    if er_key:
        for er in result.enrichment_results:
            if er.field_name == er_key and er.suggested_value and er.match_status != "NOT_FOUND":
                # Only use if confidence is sufficient
                if er.confidence >= 0.5:
                    return er.suggested_value

    # Priority 4 (LAST RESORT): AI enrichment — independent AI suggestions
    # These are not source-grounded so require more caution
    ai = result.ai_enrichment or {}
    ai_map = {
        "Beskrivelse": ai.get("improved_description"),
        "Kategori": ai.get("suggested_category"),
        "Pakningsinformasjon": ai.get("packaging_suggestions"),
    }
    ai_val = ai_map.get(field_name)
    if ai_val and isinstance(ai_val, str) and len(ai_val.strip()) > 5:
        return ai_val

    return None


def _get_suggestion_source(result: ProductAnalysis, field_name: str) -> str:
    """Determine the source of the suggestion for a field.

    Matches the priority order in _get_suggestion_for_field().
    Returns the source of whichever suggestion would actually be used.
    """
    # Priority 1: Enrichment suggestions (AI-reviewed, quality-gated)
    for es in result.enrichment_suggestions:
        if es.field_name == field_name and es.suggested_value:
            return es.source or "enrichment"

    # Priority 2: Field analysis suggestions
    for fa in result.field_analyses:
        if fa.field_name == field_name and fa.suggested_value and fa.source:
            return fa.source

    # Priority 3: Enrichment results
    enrichment_field_map = {
        "Produktnavn": "product_name",
        "Beskrivelse": "description",
        "Produsent": "manufacturer",
        "Produsentens varenummer": "manufacturer_article_number",
        "Pakningsinformasjon": "packaging_info",
    }
    er_key = enrichment_field_map.get(field_name)
    if er_key:
        for er in result.enrichment_results:
            if er.field_name == er_key and er.suggested_value and er.match_status != "NOT_FOUND":
                if er.confidence >= 0.5:
                    if er.source_level == "internal_product_sheet":
                        return "product datasheet"
                    elif er.source_level == "manufacturer_source":
                        return "manufacturer website"

    # Priority 4: AI enrichment
    ai = result.ai_enrichment or {}
    ai_map = {
        "Beskrivelse": ai.get("improved_description"),
        "Kategori": ai.get("suggested_category"),
        "Pakningsinformasjon": ai.get("packaging_suggestions"),
    }
    if field_name in ai_map and ai_map[field_name]:
        return "AI suggestion (verifiser manuelt)"

    if result.product_data.found_on_onemed:
        return "onemed.no"
    return "existing catalog"


def _determine_enrichment_status(result: ProductAnalysis) -> tuple[str, bool, str]:
    """Determine Enrichment_Status, Review_Required, and comment for a product.

    Returns (enrichment_status, review_required, comment).
    """
    pd = result.product_data
    comments = []

    if not pd.found_on_onemed:
        return "Missing Source Data", True, "Produkt ikke funnet i kilde"

    # Count how many fields have suggestions
    suggestion_fields = []
    ai_fields = []
    for field_name in ["Produktnavn", "Beskrivelse", "Spesifikasjon", "Kategori", "Pakningsinformasjon"]:
        suggestion = _get_suggestion_for_field(result, field_name)
        if suggestion:
            suggestion_fields.append(field_name)
            source = _get_suggestion_source(result, field_name)
            if "AI" in source:
                ai_fields.append(field_name)

    # Count missing/poor fields
    missing_fields = [
        fa.field_name for fa in result.field_analyses
        if fa.status in (QualityStatus.MISSING, QualityStatus.PROBABLE_ERROR)
        and fa.field_name not in ("Bildekvalitet", "Konsistens mellom felter")
    ]
    improve_fields = [
        fa.field_name for fa in result.field_analyses
        if fa.status in (QualityStatus.WEAK, QualityStatus.SHOULD_IMPROVE)
        and fa.field_name not in ("Bildekvalitet", "Konsistens mellom felter")
    ]

    # No suggestions and no problems = No Change
    if not suggestion_fields and not missing_fields:
        return "No Change", False, "Produkt har akseptabel kvalitet"

    # Missing source data if too many critical fields are missing with no suggestions
    critical_missing = [f for f in missing_fields if f in ("Produktnavn", "Beskrivelse", "Spesifikasjon")]
    if len(critical_missing) >= 2 and not suggestion_fields:
        comments.append(f"Mangler: {', '.join(critical_missing)}")
        return "Missing Source Data", True, "; ".join(comments)

    # If AI generated content or low-confidence suggestions → Needs Review
    if ai_fields:
        comments.append(f"AI-generert innhold for: {', '.join(ai_fields)}")

    # Check for conflicts in enrichment
    conflicts = [
        er for er in result.enrichment_results
        if er.match_status == "FOUND_IN_BOTH_CONFLICT"
    ]
    if conflicts:
        comments.append(f"{len(conflicts)} kildekonflikt(er)")

    # Check confidence levels
    low_confidence_fields = []
    for fa in result.field_analyses:
        if fa.suggested_value and fa.confidence and fa.confidence < 0.7:
            low_confidence_fields.append(fa.field_name)
    if low_confidence_fields:
        comments.append(f"Lav confidence: {', '.join(low_confidence_fields)}")

    needs_review = bool(ai_fields or conflicts or low_confidence_fields or missing_fields)

    if suggestion_fields and not needs_review:
        # High confidence, structured data, no AI guessing
        comments.append(f"Forslag for: {', '.join(suggestion_fields)}")
        return "Ready for Inriver", False, "; ".join(comments)

    if suggestion_fields:
        if missing_fields:
            comments.append(f"Mangler fortsatt: {', '.join(missing_fields)}")
        return "Needs Review", True, "; ".join(comments)

    if missing_fields:
        comments.append(f"Mangler: {', '.join(missing_fields)}")
        return "Needs Review", True, "; ".join(comments)

    if improve_fields:
        comments.append(f"Bør forbedres: {', '.join(improve_fields)}")
        return "Needs Review", True, "; ".join(comments)

    return "No Change", False, "Ingen endringer nødvendig"


def _is_material_change(current: str | None, suggested: str | None) -> bool:
    """Check if a suggested value is materially different from the current value.

    Returns False if values are identical, near-identical, or if suggestion is empty.
    """
    if not suggested:
        return False
    if not current:
        return True  # New value where none existed
    # Normalize for comparison
    cv = current.strip().lower()
    sv = suggested.strip().lower()
    if cv == sv:
        return False
    # Check if one is a substring of the other (minor variant)
    if len(cv) > 10 and len(sv) > 10:
        if cv in sv or sv in cv:
            # Only material if significantly more content
            if abs(len(cv) - len(sv)) < min(len(cv), len(sv)) * 0.2:
                return False
    return True


# ═══════════════════════════════════════════════════════════
# Inriver Import — strict change-only export helpers
# ═══════════════════════════════════════════════════════════

# Minimum confidence to include a suggestion in Inriver Import.
# Lower-confidence suggestions belong in Forbedringsforslag (analysis),
# NOT in the production import sheet.
_INRIVER_MIN_CONFIDENCE = 0.50


def _get_suggestion_with_confidence(
    result: "ProductAnalysis", field_name: str
) -> tuple[Optional[str], float, str]:
    """Get the best suggestion, its confidence, and source for a field.

    Returns (suggested_value, confidence, source).
    Returns (None, 0.0, "") if no suggestion exists.

    Priority: enrichment_suggestions > field_analyses > enrichment_results.
    AI enrichment (Priority 4) is EXCLUDED from Inriver — too unreliable.
    """
    # Priority 1: Enrichment suggestions (quality-gated, source-grounded)
    for es in result.enrichment_suggestions:
        if es.field_name == field_name and es.suggested_value:
            return es.suggested_value, es.confidence or 0.0, es.source or "enrichment"

    # Priority 2: Field analysis suggestions
    for fa in result.field_analyses:
        if fa.field_name == field_name and fa.suggested_value:
            return fa.suggested_value, fa.confidence or 0.0, fa.source or "analysis"

    # Priority 3: Enrichment results (raw extractions with sufficient confidence)
    enrichment_field_map = {
        "Produktnavn": "product_name",
        "Beskrivelse": "description",
        "Produsent": "manufacturer",
        "Produsentens varenummer": "manufacturer_article_number",
        "Pakningsinformasjon": "packaging_info",
    }
    er_key = enrichment_field_map.get(field_name)
    if er_key:
        for er in result.enrichment_results:
            if er.field_name == er_key and er.suggested_value and er.match_status != "NOT_FOUND":
                if er.confidence >= _INRIVER_MIN_CONFIDENCE:
                    source_label = "produktdatablad" if er.source_level == "internal_product_sheet" else "produsentens nettside"
                    return er.suggested_value, er.confidence, source_label

    # AI enrichment is deliberately EXCLUDED from Inriver Import.
    # It belongs in analysis sheets, not in production data.
    return None, 0.0, ""


def _validate_inriver_field_suggestion(
    raw_suggestion: str, field_name: str, sku: str, confidence: float
) -> tuple[bool, str, str]:
    """Validate and clean a single field suggestion for Inriver export.

    Returns (is_valid, cleaned_value, reject_reason).

    A suggestion is REJECTED (not exported) if:
    - It is empty or whitespace-only
    - It contains contact information (phone, email, address)
    - It contains article numbers for other products
    - It contains PDF noise (page numbers, headers, footers)
    - Its confidence is below the Inriver minimum threshold
    - It fails content validation
    """
    if not raw_suggestion or not raw_suggestion.strip():
        return False, "", "Tom verdi"

    raw_suggestion = raw_suggestion.strip()

    # Confidence gate — Inriver demands higher certainty than analysis sheets
    if confidence < _INRIVER_MIN_CONFIDENCE:
        return False, "", f"Confidence {confidence:.2f} er under grensen ({_INRIVER_MIN_CONFIDENCE}) for Inriver Import"

    # Content validation — contact info, other SKUs, PDF noise, variant tables
    ok, reject_reason = validate_suggestion_output(raw_suggestion, field_name, sku)
    if not ok:
        return False, "", reject_reason

    # Medical safety gate — block sensitive attributes with insufficient confidence
    from backend.medical_safety import screen_suggestion as _medical_screen
    med_result = _medical_screen(field_name, raw_suggestion, confidence, "inriver_export")
    if med_result.blocked:
        return False, "", f"Medisinsk sikkerhet: {med_result.reason}"

    # Translate sv/da → Norwegian before export
    translated, _lang, _changed = translate_to_norwegian_if_needed(raw_suggestion)

    # Final sanity: the translated value must still have substance
    if not translated or not translated.strip():
        return False, "", "Verdi ble tom etter oversettelse"

    return True, translated.strip(), ""


def _evaluate_inriver_row(
    result: "ProductAnalysis",
    import_fields: list[tuple[str, callable]],
) -> tuple[bool, list[dict]]:
    """Evaluate whether a product should appear in Inriver Import.

    Returns (should_export, field_evaluations).

    should_export is True ONLY if at least one field has a VALID,
    MATERIAL, CONFIDENT suggestion that passed all checks.

    Each field_evaluation dict contains:
        field_name, current_value, suggested_value, is_changed,
        source, confidence, reject_reason
    """
    sku = result.article_number
    field_evaluations = []
    valid_change_count = 0

    # Build a lookup of field confidence scores from the analysis
    field_confidence_map = {}
    for fa in result.field_analyses:
        if fa.confidence is not None:
            field_confidence_map[fa.field_name] = fa.confidence

    for field_name, current_fn in import_fields:
        current_val = current_fn(result)

        # Get the best suggestion with confidence
        raw_suggestion, confidence, source = _get_suggestion_with_confidence(result, field_name)

        # Get the field's own confidence from the analyzer
        field_conf = field_confidence_map.get(field_name, None)

        eval_entry = {
            "field_name": field_name,
            "current_value": current_val,
            "suggested_value": "",
            "is_changed": False,
            "source": source,
            "confidence": confidence,
            "field_confidence": field_conf,
            "reject_reason": "",
        }

        if not raw_suggestion:
            eval_entry["reject_reason"] = "Ingen forslag"
            field_evaluations.append(eval_entry)
            continue

        # Is this a material change from current?
        if not _is_material_change(current_val, raw_suggestion):
            eval_entry["reject_reason"] = "Forslaget er identisk med nåværende verdi"
            field_evaluations.append(eval_entry)
            continue

        # Validate, clean, and translate the suggestion
        is_valid, cleaned_value, reject_reason = _validate_inriver_field_suggestion(
            raw_suggestion, field_name, sku, confidence
        )

        if not is_valid:
            eval_entry["reject_reason"] = reject_reason
            logger.debug(f"[inriver] {sku}/{field_name}: avvist — {reject_reason}")
            field_evaluations.append(eval_entry)
            continue

        # This field has a valid, exportable change
        eval_entry["suggested_value"] = cleaned_value
        eval_entry["is_changed"] = True
        valid_change_count += 1
        field_evaluations.append(eval_entry)

    should_export = valid_change_count > 0
    return should_export, field_evaluations


def _collect_inriver_sources(result: "ProductAnalysis") -> set[str]:
    """Collect source labels for a product's enrichment data."""
    sources = set()
    pd = result.product_data
    if pd.found_on_onemed:
        sources.add("onemed.no")
    if result.pdf_available:
        sources.add("produktdatablad")
    for er in result.enrichment_results:
        if er.match_status != "NOT_FOUND":
            if er.source_level == "internal_product_sheet":
                sources.add("produktdatablad")
            elif er.source_level == "manufacturer_source":
                sources.add("produsentens nettside")
    return sources or {"eksisterende katalog"}


def _create_inriver_import_sheet(ws, results: list[ProductAnalysis]) -> None:
    """Create the Inriver Import sheet — a strict, change-only export.

    This is NOT an analysis sheet. It is a production staging sheet for
    direct import into Inriver PIM. Every row represents a concrete,
    validated change that a human can approve or reject.

    Strict rules:
    1. A product appears ONLY if it has ≥1 valid proposed change.
       "Has current data" is NOT a reason to include a row.
    2. For each field, ONLY fields with a validated change get a value
       in _forslag. All other fields show "ikke relevant" in _godkjent.
    3. A change is "valid" only if:
       - A suggestion exists from a grounded source (not pure AI)
       - The suggestion is materially different from the current value
       - The suggestion passes content validation (no contact info,
         no other SKUs, no PDF noise, no variant tables)
       - The confidence meets the Inriver threshold (≥0.50)
       - The suggestion is in Norwegian (translated if sv/da)
    4. If ALL fields for a product fail validation, the row is excluded.
    5. AI-only suggestions (no source evidence) are excluded entirely —
       they belong in Forbedringsforslag, not in Inriver Import.
    """
    import_fields = [
        ("Produktnavn", lambda r: r.product_data.product_name or ""),
        ("Beskrivelse", lambda r: r.product_data.description or ""),
        ("Spesifikasjon", lambda r: r.product_data.specification or (
            "; ".join(f"{k}: {v}" for k, v in r.product_data.technical_details.items())
            if r.product_data.technical_details else "")),
        ("Kategori", lambda r: r.product_data.category or (
            " > ".join(r.product_data.category_breadcrumb)
            if r.product_data.category_breadcrumb else "")),
        ("Pakningsinformasjon", lambda r: r.product_data.packaging_info or r.product_data.packaging_unit or ""),
        ("Produsent", lambda r: r.product_data.manufacturer or ""),
        ("Produsentens varenummer", lambda r: r.product_data.manufacturer_article_number or ""),
    ]

    headers = ["Artikkelnummer"]
    for field_name, _ in import_fields:
        headers.append(f"{field_name}_nåværende")
        headers.append(f"{field_name}_forslag")
        headers.append(f"{field_name}_godkjent")
    headers.extend([
        "Antall endringer",
        "Endringstyper",
        "Kilde",
        "Enrichment_Status",
        "Sist_oppdatert",
    ])

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    _style_header(ws, 1, len(headers))

    # Styles
    change_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # Light green
    approval_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Light yellow
    not_relevant_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # Gray
    not_relevant_font = Font(color="999999", italic=True)
    empty_sugg_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    top_alignment = Alignment(vertical="top")

    row_idx = 2
    exported_count = 0
    skipped_count = 0

    for result in results:
        sku = result.article_number

        # ── Core decision: should this product be in Inriver Import? ──
        should_export, field_evals = _evaluate_inriver_row(result, import_fields)

        if not should_export:
            skipped_count += 1
            continue

        # Count how many fields actually have changes
        change_count = sum(1 for fe in field_evals if fe["is_changed"])

        # Determine enrichment status and sources
        enrichment_status, _, _ = _determine_enrichment_status(result)
        sources = _collect_inriver_sources(result)

        # ── Write row ──
        _write_id_cell(ws, row_idx, 1, sku, top_alignment)

        col = 2
        for fe in field_evals:
            # _nåværende
            ws.cell(row=row_idx, column=col, value=fe["current_value"]).alignment = wrap_alignment
            col += 1

            # _forslag
            if fe["is_changed"]:
                sugg_cell = ws.cell(row=row_idx, column=col, value=fe["suggested_value"])
                sugg_cell.alignment = wrap_alignment
                sugg_cell.fill = change_fill
            else:
                sugg_cell = ws.cell(row=row_idx, column=col, value="")
                sugg_cell.fill = empty_sugg_fill
            col += 1

            # _godkjent
            if fe["is_changed"]:
                approval_cell = ws.cell(row=row_idx, column=col, value="Ikke godkjent")
                approval_cell.alignment = top_alignment
                approval_cell.fill = approval_fill
            else:
                approval_cell = ws.cell(row=row_idx, column=col, value="ikke relevant")
                approval_cell.alignment = top_alignment
                approval_cell.fill = not_relevant_fill
                approval_cell.font = not_relevant_font
            col += 1

        # Trailing metadata columns
        ws.cell(row=row_idx, column=col, value=change_count).alignment = top_alignment
        col += 1
        # Compact change type summary for all changed fields
        change_types_summary = "; ".join(sorted({
            f"{fe['field_name']}: {summarize_change_type(fe['current_value'], fe['suggested_value'])}"
            for fe in field_evals
            if fe["is_changed"] and fe.get("suggested_value")
        }))
        ws.cell(row=row_idx, column=col, value=change_types_summary).alignment = wrap_alignment
        col += 1
        ws.cell(row=row_idx, column=col, value="; ".join(sorted(sources))).alignment = top_alignment
        col += 1
        status_cell = ws.cell(row=row_idx, column=col, value=enrichment_status)
        status_cell.alignment = top_alignment
        if enrichment_status in INRIVER_STATUS_COLORS:
            status_cell.fill = INRIVER_STATUS_COLORS[enrichment_status]
            status_cell.font = INRIVER_STATUS_FONTS.get(enrichment_status, Font())
        col += 1
        ws.cell(row=row_idx, column=col, value=now_str).alignment = top_alignment
        col += 1

        row_idx += 1
        exported_count += 1

    # Empty state message
    if row_idx == 2:
        ws.cell(row=2, column=1, value="Ingen produkter med eksportklare endringsforslag")

    # Log summary
    logger.info(
        f"[inriver] Eksport ferdig: {exported_count} produkter med endringer, "
        f"{skipped_count} produkter filtrert bort (ingen gyldige endringer)"
    )

    # Column widths
    for col_idx in range(1, len(headers) + 1):
        header_text = headers[col_idx - 1]
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, min(40, len(header_text) + 4))

    ws.freeze_panes = "A2"
    if row_idx > 2:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row_idx - 1}"



# NOTE: _create_family_sheet was removed — family/relationship analysis
# is a separate module with its own export via /api/families/{source_id}/export.
# Standard masterdata Excel output must NOT contain family results.
