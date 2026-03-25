"""Loader for Jeeves ERP Excel export (Masterdata 2103.xlsx).

Reads the Excel file once, builds a lookup dict keyed by article number,
and returns JeevesData objects for individual products.

CANONICAL FIELD MODEL (from Excel columns A-I):
  Column A: Vårt art.nr      → article_number
  Column B: Vårt GID nr      → gid
  Column C: Varebeskrivelse   → item_description
  Column D: Spesifikasjon     → specification
  Column E: Produsent         → supplier (producer/manufacturer name)
  Column F: Produsent art.nr  → supplier_item_no (producer article number)
  Column G: Product brand     → product_brand
  Column H: Web Title         → web_title
  Column I: Web Text          → web_text
"""

import logging
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

from backend.identifiers import normalize_identifier
from backend.models import JeevesData

logger = logging.getLogger(__name__)

# ── Canonical column mapping ──
# Maps known header variants (case-insensitive) to JeevesData field names.
# Multiple variants are supported per field to handle different Excel versions.
# The FIRST match wins, so more specific patterns come first.

_HEADER_VARIANTS: dict[str, list[str]] = {
    "article_number": [
        "item. no", "item.no", "item no", "vårt art.nr", "vårt art nr",
        "artikkelnummer", "art.nr", "artnr",
    ],
    "gid": [
        "gid", "vårt gid nr", "vårt gid", "gid nr",
    ],
    "item_description": [
        "item description", "varebeskrivelse", "beskrivelse", "description",
    ],
    "specification": [
        "specification", "spesifikasjon", "spec",
    ],
    "supplier": [
        "supplier", "produsent", "leverandør", "manufacturer", "producer",
        "produsent (supplier)",
    ],
    "supplier_item_no": [
        "supplier item.no", "supplier item no", "supplier item.nr",
        "produsent art.nr", "produsent art nr", "produsentens varenummer",
        "produsent art.nr (supplier item number)",
        "supplier item number", "manufacturer article number",
        "leverandør art.nr",
    ],
    "product_brand": [
        "product brand", "brand", "merke", "produktmerke",
    ],
    "web_title": [
        "web title", "webtitle", "nettside tittel",
    ],
    "web_text": [
        "web text", "webtext", "nettside tekst",
    ],
}

# Fallback: If header matching fails completely, use column positions (A=0..I=8).
# This matches the canonical structure described in the module docstring.
_POSITION_FALLBACK: dict[int, str] = {
    0: "article_number",
    1: "gid",
    2: "item_description",
    3: "specification",
    4: "supplier",
    5: "supplier_item_no",
    6: "product_brand",
    7: "web_title",
    8: "web_text",
}


def _normalize_header(header) -> str:
    """Normalize a header value for fuzzy matching."""
    if header is None:
        return ""
    return str(header).strip().lower().replace("_", " ").replace("-", " ")


def _match_header(header_text: str) -> Optional[str]:
    """Match a header to a canonical field name using known variants.

    Returns the field name (e.g. "supplier") or None if no match.
    Uses longest-match-first to avoid greedy partial matches
    (e.g. "Supplier Item.no" must match supplier_item_no, not supplier).
    """
    norm = _normalize_header(header_text)
    if not norm:
        return None

    # Collect all (variant, field_name) pairs, sorted by variant length DESC
    # so more specific patterns match first
    all_pairs = []
    for field_name, variants in _HEADER_VARIANTS.items():
        for variant in variants:
            all_pairs.append((variant, field_name))
    all_pairs.sort(key=lambda p: len(p[0]), reverse=True)

    for variant, field_name in all_pairs:
        if norm == variant or norm.startswith(variant):
            return field_name
    return None


class JeevesIndex:
    """In-memory index of Jeeves product data, keyed by article number."""

    def __init__(self) -> None:
        self._data: dict[str, JeevesData] = {}
        self._loaded = False
        self._source_path: Optional[str] = None

    @property
    def loaded(self) -> bool:
        return self._loaded

    @property
    def count(self) -> int:
        return len(self._data)

    def load(self, excel_path: str) -> int:
        """Load the Jeeves Excel file and build the article number index.

        Column mapping strategy:
          1. Try to match headers by name (case-insensitive, multiple variants)
          2. If key columns are missing, fall back to column positions (A-I)
          3. Log exactly which columns map to which fields

        Returns the number of products loaded.
        """
        path = Path(excel_path)
        if not path.exists():
            raise FileNotFoundError(f"Jeeves Excel file not found: {excel_path}")

        logger.info(f"[jeeves] Loading from {excel_path}")
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb.active

        # Read header row
        rows = ws.iter_rows()
        header_row = next(rows)
        headers = [cell.value for cell in header_row]

        # ── Step 1: Match headers by name ──
        col_to_field: dict[int, str] = {}
        matched_fields: set[str] = set()

        for col_idx, header in enumerate(headers):
            if header is None:
                continue
            field = _match_header(header)
            if field and field not in matched_fields:
                col_to_field[col_idx] = field
                matched_fields.add(field)

        # ── Step 2: Log header mapping ──
        for col_idx, field in sorted(col_to_field.items()):
            header_val = headers[col_idx]
            logger.info(
                f"[jeeves] Kolonne {chr(65 + col_idx)} ({col_idx}): "
                f"'{header_val}' → {field}"
            )

        # ── Step 3: Validate critical fields, fallback to positions if needed ──
        critical_fields = {"article_number", "supplier", "supplier_item_no"}
        missing_critical = critical_fields - matched_fields

        if missing_critical:
            logger.warning(
                f"[jeeves] ⚠ Kunne ikke matche header for: {missing_critical}. "
                f"Headers funnet: {[h for h in headers if h]}. "
                f"Bruker kolonneposisjon som fallback."
            )
            # Apply position-based fallback for missing fields
            for pos, field in _POSITION_FALLBACK.items():
                if field in missing_critical and pos < len(headers):
                    col_to_field[pos] = field
                    matched_fields.add(field)
                    logger.info(
                        f"[jeeves] Fallback: Kolonne {chr(65 + pos)} ({pos}): "
                        f"'{headers[pos]}' → {field} (posisjon)"
                    )

        # Final check: article_number MUST be mapped
        if "article_number" not in matched_fields:
            wb.close()
            raise ValueError(
                f"Kunne ikke finne artikkelnummer-kolonne i {excel_path}. "
                f"Sjekk at Excel-filen har riktig format. "
                f"Headers: {[h for h in headers if h]}"
            )

        # Log warning if supplier fields are still missing
        if "supplier" not in matched_fields:
            logger.error(
                f"[jeeves] ✗ Produsent-kolonne (E) ikke funnet! "
                f"Headers: {headers[:10]}"
            )
        if "supplier_item_no" not in matched_fields:
            logger.error(
                f"[jeeves] ✗ Produsent art.nr-kolonne (F) ikke funnet! "
                f"Headers: {headers[:10]}"
            )

        # ── Step 4: Read data rows ──
        # Identifier fields that must be normalized (not free text)
        _IDENTIFIER_FIELDS = {"article_number", "gid", "supplier_item_no"}

        count = 0
        supplier_found = 0
        supplier_item_found = 0

        for row in rows:
            values: dict[str, Optional[str]] = {}
            for col_idx, field_name in col_to_field.items():
                if col_idx >= len(row):
                    continue
                cell_val = row[col_idx].value
                if cell_val is not None:
                    if field_name in _IDENTIFIER_FIELDS:
                        normalized = normalize_identifier(cell_val)
                        if normalized:
                            values[field_name] = normalized
                    else:
                        values[field_name] = str(cell_val).strip()

            artnr = values.get("article_number")
            if not artnr:
                continue

            # Track supplier coverage for diagnostics
            if values.get("supplier"):
                supplier_found += 1
            if values.get("supplier_item_no"):
                supplier_item_found += 1

            self._data[artnr] = JeevesData(**values)
            count += 1

        wb.close()
        self._loaded = True
        self._source_path = str(excel_path)

        # ── Step 5: Summary logging ──
        logger.info(
            f"[jeeves] ✓ Lastet {count} produkter fra {path.name}"
        )
        logger.info(
            f"[jeeves] Produsent-dekning: "
            f"{supplier_found}/{count} ({round(supplier_found/max(count,1)*100)}%) har produsent, "
            f"{supplier_item_found}/{count} ({round(supplier_item_found/max(count,1)*100)}%) har produsent art.nr"
        )

        # Hard validation: warn if supplier coverage is suspiciously low
        if count > 0 and supplier_found == 0:
            logger.error(
                f"[jeeves] ✗ INGEN produkter har produsent! "
                f"Sjekk at kolonne E er korrekt mappet. "
                f"Header i kolonne E: '{headers[4] if len(headers) > 4 else 'N/A'}'"
            )

        return count

    def get(self, article_number: str) -> Optional[JeevesData]:
        """Look up Jeeves data for a given article number."""
        key = normalize_identifier(article_number)
        if key is None:
            return None
        return self._data.get(key)

    def has(self, article_number: str) -> bool:
        """Check if an article number exists in Jeeves."""
        key = normalize_identifier(article_number)
        return key is not None and key in self._data

    def all_article_numbers(self) -> list[str]:
        """Return all article numbers in the index."""
        return list(self._data.keys())

    def articles_with_supplier(self) -> list[str]:
        """Return article numbers where supplier (manufacturer) is known.

        Filters out rows where supplier is empty, None, or a placeholder value.
        Used for "run quality check / relation builder for all products with manufacturer".
        """
        from backend.content_validator import is_valid_supplier
        return [
            art_no
            for art_no, jd in self._data.items()
            if is_valid_supplier(jd.supplier)
        ]

    def supplier_stats(self) -> dict[str, int]:
        """Return statistics about supplier coverage in the catalog."""
        from backend.content_validator import is_valid_supplier
        total = len(self._data)
        with_supplier = sum(1 for jd in self._data.values() if is_valid_supplier(jd.supplier))
        with_item_no = sum(1 for jd in self._data.values() if is_valid_supplier(jd.supplier_item_no))
        return {
            "total": total,
            "with_supplier": with_supplier,
            "with_supplier_item_no": with_item_no,
            "without_supplier": total - with_supplier,
        }


# Module-level singleton for convenience
_default_index: Optional[JeevesIndex] = None


def load_jeeves(excel_path: str) -> JeevesIndex:
    """Load Jeeves data and return the index. Caches the result."""
    global _default_index
    if _default_index and _default_index.loaded:
        return _default_index
    _default_index = JeevesIndex()
    _default_index.load(excel_path)
    return _default_index


def get_jeeves_data(article_number: str) -> Optional[JeevesData]:
    """Look up Jeeves data for an article number using the default index."""
    if _default_index:
        return _default_index.get(article_number)
    return None
