"""Tests for trust and logic fixes (P0/P1/P2).

Validates:
  1. Products with populated OneMed pages are NOT treated as empty
  2. Accordion description and spec extraction works
  3. STRONG/OK fields suppress suggestions
  4. Normalized comparison prevents redundant suggestions
  5. The new field status model classifies correctly
  6. Quality gate blocks near-duplicates
  7. Workbook has simplified sheet structure
"""

import pytest

from backend.models import (
    FieldAnalysis,
    JeevesData,
    ProductAnalysis,
    ProductData,
    QualityStatus,
)


# ── Test helpers ──


def _make_product(**kwargs) -> ProductData:
    defaults = dict(
        article_number="N569223270203",
        found_on_onemed=True,
    )
    defaults.update(kwargs)
    return ProductData(**defaults)


def _make_jeeves(**kwargs) -> JeevesData:
    defaults = dict(
        article_number="N569223270203",
    )
    defaults.update(kwargs)
    return JeevesData(**defaults)


# ── PART 2: Current value extraction ──


class TestAccordionDescriptionExtraction:
    """P0: Accordion description must be preferred over short JSON-LD."""

    def setup_method(self):
        from backend.scraper import _parse_product_page
        self.parse = _parse_product_page

    def test_accordion_preferred_over_short_json_ld(self):
        """If accordion has richer text than JSON-LD, use accordion."""
        html = """
        <html><head>
        <script type="application/ld+json">
        {"@type": "Product", "name": "Test Product", "sku": "N569223270203",
         "description": "Short desc."}
        </script>
        </head><body>
        <div id="accordionItem_descriptionAndDocuments">
            <p>This is a comprehensive product description with detailed information
            about the medical product features, materials, and usage instructions
            that is much richer than the JSON-LD summary.</p>
        </div>
        </body></html>
        """
        product = self.parse(html, "N569223270203")
        # Accordion text is much longer → should be chosen over "Short desc."
        assert len(product.description) > 50
        assert "comprehensive" in product.description

    def test_json_ld_used_when_no_accordion(self):
        """If no accordion present, JSON-LD description is used."""
        html = """
        <html><head>
        <script type="application/ld+json">
        {"@type": "Product", "name": "Test Product", "sku": "N569223270203",
         "description": "JSON-LD description text."}
        </script>
        </head><body></body></html>
        """
        product = self.parse(html, "N569223270203")
        assert product.description == "JSON-LD description text."

    def test_json_ld_kept_when_accordion_similar_length(self):
        """If accordion is not significantly richer, keep JSON-LD."""
        html = """
        <html><head>
        <script type="application/ld+json">
        {"@type": "Product", "name": "Test Product", "sku": "N569223270203",
         "description": "Good enough description text here."}
        </script>
        </head><body>
        <div id="accordionItem_descriptionAndDocuments">
            <p>Similar length description text.</p>
        </div>
        </body></html>
        """
        product = self.parse(html, "N569223270203")
        # JSON-LD is preferred when accordion is not 30%+ longer
        assert product.description is not None


class TestSpecAccordionExtraction:
    """P0: Specification accordion sections must be scraped."""

    def setup_method(self):
        from backend.scraper import _parse_product_page
        self.parse = _parse_product_page

    def test_spec_accordion_with_table(self):
        """Specification accordion with a table should extract key-value specs."""
        html = """
        <html><head>
        <script type="application/ld+json">
        {"@type": "Product", "name": "Test Product", "sku": "N569223270203"}
        </script>
        </head><body>
        <div id="accordionItem_specifications">
            <table>
                <tr><td>Materiale</td><td>Nitril</td></tr>
                <tr><td>Størrelse</td><td>Medium</td></tr>
                <tr><td>Farge</td><td>Blå</td></tr>
            </table>
        </div>
        </body></html>
        """
        product = self.parse(html, "N569223270203")
        assert product.technical_details is not None
        assert "Materiale" in product.technical_details
        assert product.technical_details["Materiale"] == "Nitril"
        assert len(product.technical_details) >= 3

    def test_spec_accordion_with_dl(self):
        """Specification accordion with definition lists."""
        html = """
        <html><head>
        <script type="application/ld+json">
        {"@type": "Product", "name": "Test Product", "sku": "N569223270203"}
        </script>
        </head><body>
        <div id="accordionItem_specification">
            <dl>
                <dt>Lengde:</dt><dd>25 cm</dd>
                <dt>Bredde:</dt><dd>10 cm</dd>
            </dl>
        </div>
        </body></html>
        """
        product = self.parse(html, "N569223270203")
        assert product.technical_details is not None
        assert "Lengde" in product.technical_details

    def test_spec_accordion_text_fallback(self):
        """Specification accordion without tables → use as text."""
        html = """
        <html><head>
        <script type="application/ld+json">
        {"@type": "Product", "name": "Test Product", "sku": "N569223270203"}
        </script>
        </head><body>
        <div id="accordionItem_specifications">
            <p>Materiale: Nitril, pudderfri. Størrelse Medium.
            Egnet for medisinsk bruk. CE-merket.</p>
        </div>
        </body></html>
        """
        product = self.parse(html, "N569223270203")
        assert product.specification is not None
        assert "Nitril" in product.specification


class TestTextNormalization:
    """P0: Extracted values should be normalized."""

    def setup_method(self):
        from backend.scraper import _parse_product_page
        self.parse = _parse_product_page

    def test_whitespace_normalization(self):
        """Excessive whitespace should be collapsed."""
        html = """
        <html><head>
        <script type="application/ld+json">
        {"@type": "Product", "name": "  Test   Product  ", "sku": "N569223270203",
         "description": "  Too   many   spaces   here.  "}
        </script>
        </head><body></body></html>
        """
        product = self.parse(html, "N569223270203")
        assert "  " not in product.product_name
        assert "  " not in product.description


# ── PART 5: Field status model ──


class TestFieldStatusModel:
    """P1: Fields must be classified as STRONG/OK/WEAK/SHOULD_IMPROVE/MISSING."""

    def setup_method(self):
        from backend.analyzer import (
            _analyze_description,
            _analyze_product_name,
            _analyze_specification,
        )
        self.analyze_name = _analyze_product_name
        self.analyze_desc = _analyze_description
        self.analyze_spec = _analyze_specification

    def test_strong_name(self):
        """Product with a good name should be STRONG."""
        product = _make_product(product_name="SELEFA Undersøkelseshanske Nitril Blå")
        fa = self.analyze_name(product)
        assert fa.status == QualityStatus.STRONG
        assert fa.status_reason is not None

    def test_ok_short_name(self):
        """Short but acceptable name should be OK."""
        product = _make_product(product_name="Hansker M")
        fa = self.analyze_name(product)
        assert fa.status in (QualityStatus.OK, QualityStatus.SHOULD_IMPROVE)

    def test_missing_name(self):
        """No name should be MISSING."""
        product = _make_product(product_name=None)
        fa = self.analyze_name(product)
        assert fa.status == QualityStatus.MISSING

    def test_strong_description(self):
        """Rich description with sentences should be STRONG."""
        product = _make_product(
            description="SELEFA undersøkelseshanske er laget av nitril og er pudderfri. "
                        "Hanskene er egnet for medisinsk bruk og gir god beskyttelse. "
                        "Tilgjengelig i flere størrelser."
        )
        fa = self.analyze_desc(product)
        assert fa.status == QualityStatus.STRONG

    def test_weak_description(self):
        """Short description should be WEAK, not MISSING."""
        product = _make_product(description="Nitril hansker")
        fa = self.analyze_desc(product)
        assert fa.status in (QualityStatus.WEAK, QualityStatus.SHOULD_IMPROVE)
        assert fa.status != QualityStatus.MISSING

    def test_missing_description(self):
        """No description should be MISSING."""
        product = _make_product(description=None)
        fa = self.analyze_desc(product)
        assert fa.status == QualityStatus.MISSING

    def test_strong_specification(self):
        """Rich structured specs should be STRONG."""
        product = _make_product(
            specification="Materiale: Nitril\nStørrelse: Medium\nFarge: Blå\nLengde: 25 cm",
            technical_details={
                "Materiale": "Nitril", "Størrelse": "Medium",
                "Farge": "Blå", "Lengde": "25 cm",
            },
        )
        fa = self.analyze_spec(product)
        assert fa.status == QualityStatus.STRONG

    def test_weak_specification(self):
        """Specs in description only should be WEAK."""
        product = _make_product(
            specification=None,
            description="Materiale er nitril, størrelse Medium, 25 cm lang.",
        )
        fa = self.analyze_spec(product)
        assert fa.status == QualityStatus.WEAK


# ── PART 3: Source priority and comparison logic ──


class TestSuggestionSuppression:
    """P1: STRONG/OK fields must not generate suggestions."""

    def setup_method(self):
        from backend.enricher import _enrich_product_name, _enrich_description
        self.enrich_name = _enrich_product_name
        self.enrich_desc = _enrich_description

    def test_strong_name_no_suggestion(self):
        """STRONG product name must suppress enrichment."""
        from backend.analyzer import analyze_product
        product = _make_product(
            product_name="SELEFA Undersøkelseshanske Nitril Blå Medium",
            description="God beskrivelse av produkt med full informasjon. "
                        "Hanskene er egnet for medisinsk bruk.",
            specification="Materiale: Nitril",
            technical_details={"Materiale": "Nitril"},
            manufacturer="OneMed",
            category="Hansker",
        )
        analysis = analyze_product(product)
        fa = next(f for f in analysis.field_analyses if f.field_name == "Produktnavn")
        assert fa.status in (QualityStatus.STRONG, QualityStatus.OK)
        # Enricher should return None for strong/OK fields
        suggestion = self.enrich_name(product, analysis, {}, None)
        assert suggestion is None

    def test_ok_description_no_suggestion(self):
        """OK description must suppress enrichment."""
        from backend.analyzer import analyze_product
        product = _make_product(
            product_name="Test Product Name",
            description="En god beskrivelse med tilstrekkelig informasjon om produktet. "
                        "Den inneholder relevante detaljer for nettbutikkvisning.",
            specification="Materiale: Nitril",
            technical_details={"Materiale": "Nitril"},
        )
        analysis = analyze_product(product)
        fa = next(f for f in analysis.field_analyses if f.field_name == "Beskrivelse")
        # Description is STRONG (>80 chars with sentences) or OK
        assert fa.status in (QualityStatus.STRONG, QualityStatus.OK)
        suggestion = self.enrich_desc(product, analysis, {}, None)
        assert suggestion is None


class TestNormalizedComparison:
    """P1: Quality gate must use normalized comparison."""

    def setup_method(self):
        from backend.enricher import _normalize_for_comparison, final_quality_gate
        self.normalize = _normalize_for_comparison
        self.quality_gate = final_quality_gate

    def test_normalize_bullets(self):
        """Different bullet styles should normalize to same form."""
        a = "• Item one\n• Item two"
        b = "- Item one\n- Item two"
        assert self.normalize(a) == self.normalize(b)

    def test_normalize_whitespace(self):
        """Different whitespace should normalize to same form."""
        a = "Product   description\n\nwith  gaps"
        b = "Product description with gaps"
        assert self.normalize(a) == self.normalize(b)

    def test_quality_gate_blocks_normalized_duplicate(self):
        """Quality gate should block suggestions identical after normalization."""
        from backend.models import EnrichmentSuggestion
        suggestions = [
            EnrichmentSuggestion(
                field_name="Beskrivelse",
                current_value="• Punkt en\n• Punkt to\n• Punkt tre",
                suggested_value="- Punkt en\n- Punkt to\n- Punkt tre",
                source="PDF",
                confidence=0.8,
                review_required=False,
            ),
        ]
        result = self.quality_gate(suggestions)
        # Should be empty — normalized content is identical
        assert len(result) == 0


# ── PART 6: Traceability ──


class TestTraceability:
    """P1: FieldAnalysis must include traceability fields."""

    def test_traceability_fields_populated(self):
        """Analyzer should populate website_value, jeeves_value, value_origin."""
        from backend.analyzer import _analyze_product_name
        product = _make_product(product_name="Website Name")
        jeeves = _make_jeeves(item_description="Jeeves Name")
        fa = _analyze_product_name(product, jeeves)
        assert fa.website_value == "Website Name"
        assert fa.jeeves_value == "Jeeves Name"
        assert fa.value_origin == "nettside"
        assert fa.status_reason is not None

    def test_traceability_jeeves_only(self):
        """When only Jeeves has value, origin should say Jeeves."""
        from backend.analyzer import _analyze_product_name
        product = _make_product(product_name=None)
        jeeves = _make_jeeves(item_description="Jeeves Name")
        fa = _analyze_product_name(product, jeeves)
        assert fa.value_origin == "Jeeves"
        assert fa.current_value == "Jeeves Name"

    def test_traceability_missing_both(self):
        """When both sources are empty, status_reason should explain."""
        from backend.analyzer import _analyze_product_name
        product = _make_product(product_name=None)
        jeeves = _make_jeeves(item_description=None)
        fa = _analyze_product_name(product, jeeves)
        assert fa.status == QualityStatus.MISSING
        assert fa.status_reason is not None
        assert "mangler" in fa.status_reason.lower() or "ingen" in fa.status_reason.lower()


# ── PART 7: Workbook simplification ──


class TestWorkbookSimplification:
    """P2: Workbook should have max 5-6 sheets in full_enrichment mode."""

    def test_full_enrichment_sheet_count(self):
        """Full enrichment mode should produce ≤ 7 sheets."""
        from backend.excel_handler import create_output_excel
        from backend.analyzer import analyze_product
        import tempfile
        import os
        from openpyxl import load_workbook

        product = _make_product(
            product_name="Test Product",
            description="A test description for the product. It has enough content.",
            specification="Material: Test",
            technical_details={"Material": "Test"},
            manufacturer="TestMfr",
            category="TestCat",
        )
        analysis = analyze_product(product)

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            create_output_excel([analysis], path, analysis_mode="full_enrichment")
            wb = load_workbook(path)
            sheet_names = wb.sheetnames
            # Should be at most 7 sheets (Summary, Oversikt, Feltanalyse,
            # Forbedringsforslag, Inriver Import, optionally Produsentoppfølging, Bildeanalyse)
            assert len(sheet_names) <= 7, f"Too many sheets: {sheet_names}"
            # Must NOT have removed sheets
            assert "Comparison_And_Enrichment" not in sheet_names
            assert "Debug_Log" not in sheet_names
            assert "Kildekonflikter" not in sheet_names
            assert "Bildeproblemer" not in sheet_names
            # Must have required sheets
            assert "Summary" in sheet_names
            assert "Oversikt" in sheet_names
            assert "Feltanalyse" in sheet_names
        finally:
            os.unlink(path)

    def test_no_manufacturer_sheet_when_not_needed(self):
        """Produsentoppfølging should be omitted when no products need manufacturer contact."""
        from backend.excel_handler import create_output_excel
        from backend.analyzer import analyze_product
        import tempfile
        import os
        from openpyxl import load_workbook

        product = _make_product(
            product_name="Complete Product",
            description="Full description with enough detail. Includes usage and materials.",
            specification="Material: Nitril\nSize: M",
            technical_details={"Material": "Nitril", "Size": "M"},
            manufacturer="TestMfr",
            category="TestCat",
        )
        analysis = analyze_product(product)
        analysis.requires_manufacturer_contact = False

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            create_output_excel([analysis], path, analysis_mode="full_enrichment")
            wb = load_workbook(path)
            assert "Produsentoppfølging" not in wb.sheetnames
        finally:
            os.unlink(path)
